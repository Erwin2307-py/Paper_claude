"""
Voraussetzungen
---------------
pip install openpyxl requests
( Claude API-Key bleibt optional; ohne Key werden Standard-Texte eingesetzt.)
"""

import os, shutil, re, json, requests, xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import load_workbook


# ------------------------------------------------------------------
# 1. Vorlage 1-zu-1 kopieren
# ------------------------------------------------------------------
def copy_template(gene):
    src = "vorlage_paperqa2.xlsx"               # angehängte Original­datei
    if not os.path.exists(src):
        raise FileNotFoundError(f"Vorlage '{src}' nicht gefunden!")
    dst = f"ausgefuellt_{gene}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    shutil.copy2(src, dst)                      # byte­genaue Kopie
    if os.path.getsize(src) != os.path.getsize(dst):
        raise IOError("Kopierfehler – Dateigrößen stimmen nicht überein.")
    return dst


# ------------------------------------------------------------------
# 2. rs-Nummern via dbSNP (max 3)
# ------------------------------------------------------------------
def fetch_rs_numbers(gene):
    ids_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    ids = requests.get(ids_url, params={
        "db": "snp", "term": f"{gene}[Gene Name]", "retmax": 3, "retmode": "xml"
    }, timeout=10).text
    snp_ids = [id_.text for id_ in ET.fromstring(ids).iterfind(".//Id")]
    if not snp_ids:
        return [f"rs{gene.lower()}1"]

    fetch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    xml = requests.get(fetch_url, params={
        "db": "snp", "id": ",".join(snp_ids), "retmode": "xml"
    }, timeout=10).text
    return [f"rs{rs.get('rsId')}" for rs in ET.fromstring(xml).iterfind(".//Rs")][:3]


# ------------------------------------------------------------------
# 3. Claude-API: Genotyp-/Phänotyp-Texte erstellen (optional)
# ------------------------------------------------------------------
def ask_claude(gene, rs_numbers, api_key):
    if not api_key:
        return None                            # ohne Key sofort abbrechen
    prompt = f"""
Analysiere das Gen {gene}. Nutze die rs-Nummern {', '.join(rs_numbers)}.
Erstelle JSON:
{{
 "genotypes": ["AA","AG","GG"],
 "frequencies": ["45%","35%","20%"],
 "phenotypes": ["Text zu AA","Text zu AG","Text zu GG"]
}}"""
    headers = {
        "Content-Type": "application/json",
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01"
    }
    payload = {
        "model": "claude-3-haiku-20240307",
        "max_tokens": 800,
        "messages": [{"role": "user", "content": prompt}]
    }
    try:
        rsp = requests.post("https://api.anthropic.com/v1/messages",
                            headers=headers, json=payload, timeout=30)
        rsp.raise_for_status()
        txt = rsp.json()["content"][0]["text"]
        data = json.loads(re.search(r"{.*}", txt, re.S).group())
        return data
    except Exception:
        return None                            # bei Fehler Fallback nutzen


# ------------------------------------------------------------------
# 4. Excel-Kopie ausfüllen (Format bleibt erhalten)
# ------------------------------------------------------------------
def fill_excel(xlsx, gene, rs_num, data):
    wb, ws = load_workbook(xlsx), load_workbook(xlsx).active

    # Grund­daten
    ws["D5"] = gene
    ws["D6"] = rs_num
    ws["I2"] = datetime.now().strftime("%d.%m.%Y")

    # Genotyp-Block
    genos  = data.get("genotypes",   ["AA", "AG", "GG"])
    freqs  = data.get("frequencies", ["45%", "35%", "20%"])
    phenos = data.get("phenotypes",  ["-", "-", "-"])

    for i in range(3):                      # Zeilen 10-12
        row = 10 + i
        ws[f"D{row}"] = genos[i]
        ws[f"E{row}"] = freqs[i]
        ws[f"F{row}"] = phenos[i]

    wb.save(xlsx)


# ------------------------------------------------------------------
# 5. Haupt­routine
# ------------------------------------------------------------------
def run(gene, api_key=None):
    xlsx      = copy_template(gene)
    rs_nums   = fetch_rs_numbers(gene)
    claude    = ask_claude(gene, rs_nums, api_key) or {}
    fill_excel(xlsx, gene, rs_nums[0], claude)
    print("✅ Fertig – ausgefüllte Datei:", xlsx)


# ------------------------------------------------------------------
if __name__ == "__main__":
    g_name = input("Gen-Name: ").strip()
    c_key  = input("Claude API-Key (leer = Standardtexte): ").strip() or None
    run(g_name, c_key)
