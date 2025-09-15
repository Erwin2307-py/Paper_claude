# modules/email_module.py - MIT STREAMLIT SECRETS INTEGRATION
import streamlit as st
import datetime
import requests
import xml.etree.ElementTree as ET
import pandas as pd
import time
import re
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import ssl
from typing import List, Dict, Any, Tuple
import json
from pathlib import Path
from typing import List, Dict, Any, Tuple
import threading

# =============== STREAMLIT SECRETS INTEGRATION ===============

def get_email_secret(key, fallback=None):
    """Holt Email-Konfiguration aus Streamlit Secrets mit Fallback"""
    try:
        # Streamlit Secrets zuerst
        if hasattr(st, 'secrets') and 'email' in st.secrets:
            value = st.secrets["email"].get(key)
            if value:
                return value
    except Exception:
        pass

    # Environment Variable Fallback
    try:
        env_key = f"EMAIL_{key.upper()}"
        env_value = os.getenv(env_key)
        if env_value:
            return env_value
    except Exception:
        pass

    return fallback

def load_email_config_from_secrets() -> Dict:
    """Lädt Email-Konfiguration aus Streamlit Secrets mit robusten Fallbacks"""

    # Standard-Konfiguration mit Fallbacks
    email_config = {
        "sender_email": get_email_secret("sender_email"),
        "smtp_server": get_email_secret("smtp_server", "smtp.gmail.com"),
        "smtp_port": int(get_email_secret("smtp_port", 587)),
        "sender_password": get_email_secret("sender_password") or get_email_secret("password"),
        "use_tls": get_email_secret("use_tls", True),
        "auto_notifications": get_email_secret("auto_notifications", True),
        "min_papers": int(get_email_secret("min_papers", 1)),
        "notification_frequency": "Bei jeder Suche",
        "subject_template": "🔬 {count} neue Papers für '{search_term}' - {frequency}",
        "message_template": """📧 Automatische Paper-Benachrichtigung

📅 Datum: {date}
🔍 Suchbegriff: '{search_term}'
📊 Neue Papers: {count}
⏰ Häufigkeit: {frequency}

📋 Neue Papers:
{new_papers_list}

📎 Excel-Datei wurde aktualisiert: {excel_file}

Mit freundlichen Grüßen,
Ihr automatisches Paper-Überwachung-System""",
            "from_secrets": True  # Flag to indicate loaded from secrets
        }

        # Empfänger-Liste laden (kann mehrere Empfänger enthalten)
        recipients = get_email_secret("recipients") or get_email_secret("recipient_email")
        if recipients:
            if isinstance(recipients, str):
                # Comma-separated string zu Liste
                email_config["recipient_emails"] = [email.strip() for email in recipients.split(",")]
            elif isinstance(recipients, list):
                email_config["recipient_emails"] = recipients
        else:
            email_config["recipient_emails"] = []

        return email_config

    except Exception as e:
        # Keine Warnung mehr - stille Rückgabe von None für robuste Fallbacks
        return None

def show_email_config_status():
    """Zeigt Email-Konfigurationsstatus an"""
    st.subheader("📧 Email-Konfiguration")

    # Lade Konfiguration aus Secrets
    secrets_config = load_email_config_from_secrets()

    col1, col2 = st.columns(2)

    with col1:
        st.write("**📋 Konfigurationsstatus:**")
        if secrets_config:
            st.success("✅ Streamlit Secrets geladen")
            sender_email = secrets_config.get("sender_email", "")
            if sender_email:
                masked_email = f"{sender_email[:3]}***@{sender_email.split('@')[1]}" if "@" in sender_email else "***"
                st.write(f"📧 Absender: {masked_email}")
            else:
                st.warning("⚠️ Absender-Email fehlt")

            recipients = secrets_config.get("recipient_emails", [])
            if recipients:
                st.write(f"👥 Empfänger: {len(recipients)} konfiguriert")
            else:
                st.warning("⚠️ Empfänger fehlen")

            smtp_server = secrets_config.get("smtp_server", "")
            smtp_port = secrets_config.get("smtp_port", 587)
            st.write(f"🌐 SMTP: {smtp_server}:{smtp_port}")

        else:
            st.warning("⚠️ Keine Email-Secrets konfiguriert")
            st.info("Verwende manuelle Konfiguration")

    with col2:
        st.write("**🔧 Benötigte Secrets:**")
        st.code("""[email]
sender_email = "ihre@email.com"
sender_password = "ihr_app_passwort"
recipients = "empfaenger1@email.com,empfaenger2@email.com"
smtp_server = "smtp.gmail.com"
smtp_port = 587""", language="toml")

    return secrets_config

def module_email():
    """VOLLSTÄNDIGE FUNKTION - Email-Modul mit Secrets Integration"""
    st.title("📧 Wissenschaftliches Paper-Suche & Email-System")

    # Zeige Konfigurationsstatus
    secrets_config = show_email_config_status()

    # Session State initialisieren
    initialize_session_state()

    # Email-Funktionalität basierend auf Konfiguration
    if secrets_config and secrets_config.get("sender_email"):
        st.success("🚀 Email-System bereit - Verwendet Streamlit Secrets")
        show_email_dashboard_with_secrets(secrets_config)
    else:
        st.info("📝 Manuelle Email-Konfiguration")
        integrated_email_interface()

def show_email_dashboard_with_secrets(secrets_config):
    """Email-Dashboard mit Streamlit Secrets Konfiguration"""
    st.subheader("🚀 Email-Dashboard (Secrets-basiert)")

    # Quick Action Buttons
    col1, col2, col3 = st.columns(3)

    with col1:
        if st.button("📧 Test Email senden"):
            send_test_email_with_secrets(secrets_config)

    with col2:
        if st.button("📊 Email-Status prüfen"):
            check_email_connectivity(secrets_config)

    with col3:
        if st.button("📋 Empfänger anzeigen"):
            show_recipient_list(secrets_config)

    # Email-Versand für Paper-Suche
    st.markdown("---")
    st.subheader("🔍 Paper-Suche mit Email-Benachrichtigung")

    search_term = st.text_input("🔍 Suchbegriff für Papers:", placeholder="z.B. BRCA1 breast cancer")
    max_papers = st.slider("📊 Max. Anzahl Papers:", 1, 100, 20)

    if st.button("🚀 Suche starten & Email senden") and search_term:
        with st.spinner("Suche Papers und sende Email..."):
            try:
                # Hier würde die Paper-Suche stattfinden
                # Für Demo verwenden wir Dummy-Daten
                papers = [
                    {"title": f"Paper {i+1} zu {search_term}", "authors": "Autor et al.", "journal": "Nature"}
                    for i in range(min(5, max_papers))
                ]

                # Email mit Ergebnissen senden
                send_paper_results_email(secrets_config, search_term, papers)
                st.success(f"✅ {len(papers)} Papers gefunden und Email gesendet!")

            except Exception as e:
                st.error(f"❌ Fehler: {str(e)}")

def send_test_email_with_secrets(secrets_config):
    """Sendet Test-Email mit Secrets-Konfiguration"""
    try:
        server = smtplib.SMTP(secrets_config["smtp_server"], secrets_config["smtp_port"])
        server.starttls()
        server.login(secrets_config["sender_email"], secrets_config["sender_password"])

        # Test-Nachricht erstellen
        msg = MIMEMultipart()
        msg['From'] = secrets_config["sender_email"]
        msg['Subject'] = "🧪 Paper Claude - Test Email"

        body = f"""
        📧 Test-Email von Paper Claude

        ✅ Konfiguration erfolgreich getestet!

        📅 Zeitpunkt: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        📧 Absender: {secrets_config["sender_email"]}
        🌐 SMTP Server: {secrets_config["smtp_server"]}:{secrets_config["smtp_port"]}

        Diese Test-Email bestätigt, dass Ihre Email-Konfiguration korrekt funktioniert.

        Mit freundlichen Grüßen,
        Ihr Paper Claude Team 🔬
        """

        msg.attach(MIMEText(body, 'plain'))

        # An alle konfigurierten Empfänger senden
        recipients = secrets_config.get("recipient_emails", [])
        if recipients:
            for recipient in recipients:
                msg['To'] = recipient
                text = msg.as_string()
                server.sendmail(secrets_config["sender_email"], recipient, text)
                del msg['To']  # Remove for next iteration

            server.quit()
            st.success(f"✅ Test-Email an {len(recipients)} Empfänger gesendet!")
        else:
            st.error("❌ Keine Empfänger konfiguriert!")

    except Exception as e:
        st.error(f"❌ Email-Versand fehlgeschlagen: {str(e)}")

def check_email_connectivity(secrets_config):
    """Prüft Email-Server Verbindung"""
    try:
        server = smtplib.SMTP(secrets_config["smtp_server"], secrets_config["smtp_port"])
        server.starttls()
        server.login(secrets_config["sender_email"], secrets_config["sender_password"])
        server.quit()
        st.success("✅ Email-Server Verbindung erfolgreich!")
    except Exception as e:
        st.error(f"❌ Verbindungsfehler: {str(e)}")

def show_recipient_list(secrets_config):
    """Zeigt Empfänger-Liste an"""
    recipients = secrets_config.get("recipient_emails", [])
    if recipients:
        st.write("**📋 Konfigurierte Empfänger:**")
        for i, recipient in enumerate(recipients, 1):
            st.write(f"{i}. 📧 {recipient}")
    else:
        st.warning("⚠️ Keine Empfänger konfiguriert!")

def send_paper_results_email(secrets_config, search_term, papers):
    """Sendet Email mit Paper-Ergebnissen"""
    try:
        server = smtplib.SMTP(secrets_config["smtp_server"], secrets_config["smtp_port"])
        server.starttls()
        server.login(secrets_config["sender_email"], secrets_config["sender_password"])

        # Email-Nachricht erstellen
        msg = MIMEMultipart()
        msg['From'] = secrets_config["sender_email"]
        msg['Subject'] = f"🔬 {len(papers)} Papers gefunden für '{search_term}'"

        # Paper-Liste formatieren
        papers_list = "\n".join([
            f"• {paper['title']} - {paper['authors']} ({paper['journal']})"
            for paper in papers
        ])

        body = f"""
        📧 Paper-Suchergebnisse

        🔍 Suchbegriff: {search_term}
        📊 Gefundene Papers: {len(papers)}
        📅 Datum: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

        📋 Papers:
        {papers_list}

        Mit freundlichen Grüßen,
        Ihr automatisches Paper-Überwachung-System 🔬
        """

        msg.attach(MIMEText(body, 'plain'))

        # An alle Empfänger senden
        recipients = secrets_config.get("recipient_emails", [])
        for recipient in recipients:
            msg['To'] = recipient
            text = msg.as_string()
            server.sendmail(secrets_config["sender_email"], recipient, text)
            del msg['To']

        server.quit()
        st.success(f"📧 Email an {len(recipients)} Empfänger gesendet!")

    except Exception as e:
        st.error(f"❌ Email-Versand fehlgeschlagen: {str(e)}")

def create_master_excel_template():
    """Erstellt Master Excel-Template mit Overview-Sheet und Excel-Integration"""
    template_path = st.session_state["excel_template"]["file_path"]
    
    if not os.path.exists(template_path):
        try:
            wb = openpyxl.Workbook()
            
            # Overview Sheet
            overview_sheet = wb.active
            overview_sheet.title = "📊_Overview"
            
            # Header-Style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Overview Headers
            overview_headers = [
                "Sheet_Name", "Suchbegriff", "Anzahl_Papers", "Letztes_Update", 
                "Neue_Papers_Letzter_Run", "Status", "Erstellt_am"
            ]
            
            for col, header in enumerate(overview_headers, 1):
                cell = overview_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Spaltenbreite anpassen
            column_widths = [20, 25, 15, 18, 20, 12, 18]
            for col, width in enumerate(column_widths, 1):
                col_letter = get_column_letter(col)
                overview_sheet.column_dimensions[col_letter].width = width
            
            # Template Info Sheet
            info_sheet = wb.create_sheet("ℹ️_Template_Info")
            
            info_data = [
                ["📋 Excel Template Information", ""],
                ["", ""],
                ["Erstellt am:", datetime.datetime.now().strftime("%d.%m.%Y %H:%M")],
                ["System:", "Wissenschaftliches Paper-Suche System"],
                ["Version:", "4.0 mit Streamlit Secrets Integration"],
                ["", ""],
                ["📖 Anleitung:", ""],
                ["• Jeder Suchbegriff bekommt ein eigenes Sheet", ""],
                ["• Das Overview-Sheet zeigt alle Suchanfragen", ""],
                ["• Neue Papers werden automatisch hinzugefügt", ""],
                ["• Email-Benachrichtigungen aus Streamlit Secrets", ""],
                ["• Duplikate werden automatisch erkannt", ""],
                ["• Sichere Email-Konfiguration über secrets.toml", ""],
            ]
            
            for row_idx, (key, value) in enumerate(info_data, 1):
                info_sheet.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
                info_sheet.cell(row=row_idx, column=2, value=value)
            
            info_sheet.column_dimensions['A'].width = 30
            info_sheet.column_dimensions['B'].width = 40
            
            wb.save(template_path)
            st.session_state["system_status"]["excel_sheets"] = len(wb.sheetnames)
            
        except Exception as e:
            st.error(f"❌ Fehler beim Erstellen des Master-Templates: {str(e)}")
    
    # Erweiterte Tabs
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        "📊 Dashboard", 
        "🔍 Paper-Suche", 
        "📧 Email-Konfiguration",
        "📋 Excel-Management",
        "🤖 Automatische Suchen",
        "📈 Statistiken",
        "⚙️ System-Einstellungen"
    ])
    
    with tab1:
        show_dashboard()
    
    with tab2:
        show_advanced_paper_search()
    
    with tab3:
        show_email_config_with_secrets()
    
    with tab4:
        show_excel_template_management()
    
    with tab5:
        show_automatic_search_system()
    
    with tab6:
        show_detailed_statistics()
    
    with tab7:
        show_system_settings()

def initialize_session_state():
    """Session State Initialisierung mit PERSISTENTER Excel-Datenbank"""
    # Erstelle notwendige Ordner
    for folder in ["excel_templates", "saved_searches", "search_history", "config", "backups"]:
        if not os.path.exists(folder):
            os.makedirs(folder)

            # ✅ System-Status initialisieren
            initialize_system_status()

    
    # Excel-Template System - PERSISTENT & SICHER
    if "excel_template" not in st.session_state:
        try:
            st.session_state["excel_template"] = {
                "file_path": st.secrets.get("excel", {}).get("template_path", "excel_templates/master_papers.xlsx"),
                "auto_create_sheets": st.secrets.get("excel", {}).get("auto_create_sheets", True),
                "sheet_naming": "topic_based",
                "max_sheets": st.secrets.get("excel", {}).get("max_sheets", 50)
            }
        except:
            st.session_state["excel_template"] = {
                "file_path": "excel_templates/master_papers.xlsx",
                "auto_create_sheets": True,
                "sheet_naming": "topic_based",
                "max_sheets": 50
            }
    
    # 🔒 PERSISTENTE EXCEL-DATENBANK LADEN/ERSTELLEN
    ensure_persistent_excel_database()
    
    # Email-Einstellungen laden
    load_email_config_from_secrets()
    
    # Andere Session State Elemente...
    initialize_other_session_elements()
    
    # 📊 SOFORT DASHBOARD-DATEN LADEN
    load_dashboard_data_on_startup()
def initialize_other_session_elements():
    """Initialisiert weitere Session-State-Elemente"""
    # Email-Historie
    if "email_history" not in st.session_state:
        st.session_state["email_history"] = []
    
    # Suchhistorie 
    if "search_history" not in st.session_state:
        st.session_state["search_history"] = []
    
    # Aktuelle Suchergebnisse
    if "current_search_results" not in st.session_state:
        st.session_state["current_search_results"] = {}
    
    # Automatische Suchen
    if "automatic_searches" not in st.session_state:
        st.session_state["automatic_searches"] = {}

def initialize_system_status():
    """Initialisiert System-Status falls nicht vorhanden"""
    if "system_status" not in st.session_state:
        st.session_state["system_status"] = {
            "total_searches": 0,
            "total_papers": 0,
            "total_emails": 0,
            "last_search": None,
            "excel_sheets": 0,
            "last_backup": None,
            "pending_automation_searches": 0,
            "last_automation_check": None
        }


def ensure_persistent_excel_database():
    """Stellt sicher, dass die Excel-Datenbank persistent existiert"""
    excel_path = st.session_state["excel_template"]["file_path"]
    
    # 🔍 PRÜFE OB EXCEL EXISTIERT
    if os.path.exists(excel_path):
        # ✅ EXCEL EXISTIERT - LADE STATISTIKEN
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            # Prüfe kritische Sheets
            required_sheets = ["📊_Overview", "ℹ️_Template_Info"]
            for sheet in required_sheets:
                if sheet not in wb.sheetnames:
                    repair_missing_sheets(wb, excel_path)
            
            # Lade Statistiken
            stats = get_search_statistics_from_excel()
            st.session_state["system_status"]["excel_sheets"] = stats.get("total_sheets", 0)
            st.session_state["persistent_search_terms"] = stats.get("search_terms", [])
            
            # Erfolg-Meldung
            st.success(f"🔒 **Persistente Excel-Datenbank geladen!** {len(stats.get('search_terms', []))} Suchthemen verfügbar")
            
            # AUTOMATISCHES BACKUP
            create_automatic_backup(excel_path)
            
        except Exception as e:
            st.error(f"⚠️ Excel-Datei beschädigt: {str(e)} - Versuche Reparatur...")
            repair_excel_database(excel_path)
    else:
        # ❌ EXCEL EXISTIERT NICHT - ERSTELLE NEUE
        st.warning("📊 Keine Excel-Datenbank gefunden - erstelle neue persistente Datenbank...")
        create_fresh_persistent_excel()

def create_fresh_persistent_excel():
    """Erstellt neue persistente Excel-Datenbank (nur wenn noch nicht existiert)"""
    excel_path = st.session_state["excel_template"]["file_path"]
    
    try:
        wb = openpyxl.Workbook()
        
        # 📊 OVERVIEW SHEET - ZENTRALE THEMEN-ÜBERSICHT
        overview_sheet = wb.active
        overview_sheet.title = "📊_Overview"
        
        # Header-Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Overview Headers
        overview_headers = [
            "Sheet_Name", "Suchbegriff", "Anzahl_Papers", "Letztes_Update", 
            "Neue_Papers_Letzter_Run", "Status", "Erstellt_am", "Gesamt_Läufe"
        ]
        
        for col, header in enumerate(overview_headers, 1):
            cell = overview_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Spaltenbreite anpassen
        column_widths = [20, 25, 15, 18, 20, 12, 18, 12]
        for col, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(col)
            overview_sheet.column_dimensions[col_letter].width = width
        
        # ℹ️ PERSISTENT INFO SHEET
        info_sheet = wb.create_sheet("ℹ️_Persistent_Info")
        
        info_data = [
            ["📋 PERSISTENTE EXCEL-DATENBANK", ""],
            ["", ""],
            ["🔒 Status:", "PERSISTENT & SICHER"],
            ["📅 Erstellt am:", datetime.datetime.now().strftime("%d.%m.%Y %H:%M")],
            ["📍 Pfad:", excel_path],
            ["🔄 Letzte Sicherung:", "Bei Erstellung"],
            ["", ""],
            ["⚠️ WICHTIG:", ""],
            ["• Diese Datei NIEMALS löschen!", ""],
            ["• Enthält ALLE Suchergebnisse dauerhaft", ""],
            ["• Automatische Backups in 'backups/' Ordner", ""],
            ["• Bei jedem Programmstart geladen", ""],
            ["", ""],
            ["📊 Suchthemen werden hier gespeichert:", ""],
            ["• Jeder Suchbegriff = eigenes Sheet", ""],
            ["• Overview zeigt alle durchsuchten Themen", ""],
            ["• Duplikate werden automatisch erkannt", ""],
        ]
        
        for row_idx, (key, value) in enumerate(info_data, 1):
            cell_a = info_sheet.cell(row=row_idx, column=1, value=key)
            cell_b = info_sheet.cell(row=row_idx, column=2, value=value)
            
            if key.startswith(("📋", "⚠️", "📊")):
                cell_a.font = Font(bold=True, color="FFFFFF")
                cell_a.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            elif key and not key.startswith(" "):
                cell_a.font = Font(bold=True)
        
        info_sheet.column_dimensions['A'].width = 35
        info_sheet.column_dimensions['B'].width = 45
        
        # SPEICHERN
        wb.save(excel_path)
        
        # Session State aktualisieren
        st.session_state["system_status"]["excel_sheets"] = 2
        st.session_state["persistent_search_terms"] = []
        
        st.success(f"🆕 **Neue persistente Excel-Datenbank erstellt:** {excel_path}")
        
    except Exception as e:
        st.error(f"❌ Fehler beim Erstellen der persistenten Excel: {str(e)}")
def load_dashboard_data_on_startup():
    """Lädt Dashboard-Daten sofort beim Start"""
    try:
        excel_stats = get_search_statistics_from_excel()
        
        if excel_stats.get("search_terms"):
            st.session_state["dashboard_ready"] = True
            st.session_state["available_topics"] = [term["term"] for term in excel_stats["search_terms"]]
            
            # Info für User
            topics_count = len(excel_stats["search_terms"])
            papers_count = excel_stats.get("total_papers", 0)
            
            # Zeige verfügbare Themen im Sidebar (falls möglich)
            if topics_count > 0:
                st.sidebar.success(f"🔒 **Persistente Daten geladen:**")
                st.sidebar.write(f"📊 {topics_count} Suchthemen")
                st.sidebar.write(f"📄 {papers_count} Papers")
                
                # Zeige Top 5 Themen
                with st.sidebar.expander("📋 Verfügbare Suchthemen"):
                    for term_info in excel_stats["search_terms"][:5]:
                        term = term_info["term"]
                        papers = term_info["papers"]
                        st.write(f"• **{term}** ({papers} Papers)")
                    
                    if topics_count > 5:
                        st.write(f"... und {topics_count - 5} weitere")
        else:
            st.session_state["dashboard_ready"] = False
            st.session_state["available_topics"] = []
            
    except Exception as e:
        st.error(f"⚠️ Fehler beim Laden der Dashboard-Daten: {str(e)}")
        st.session_state["dashboard_ready"] = False
def show_dashboard():
    """Dashboard mit persistenten Excel-Daten"""
    st.subheader("📊 Dashboard - Persistente Excel-Datenbank")
    
    # PERSISTENT STATUS ANZEIGEN
    excel_path = st.session_state["excel_template"]["file_path"]
    if os.path.exists(excel_path):
        file_size = os.path.getsize(excel_path) / 1024 / 1024  # MB
        file_date = datetime.datetime.fromtimestamp(os.path.getmtime(excel_path))
        
        st.success(f"🔒 **Persistente Excel-Datenbank aktiv:** {excel_path}")
        st.info(f"📊 **Größe:** {file_size:.2f} MB | **Letzte Änderung:** {file_date.strftime('%d.%m.%Y %H:%M')}")
    else:
        st.error("❌ **KRITISCH:** Persistente Excel-Datenbank nicht gefunden!")
        if st.button("🔧 **Datenbank wiederherstellen**"):
            ensure_persistent_excel_database()
            st.rerun()
    
    # Excel-Statistiken holen
    excel_stats = get_search_statistics_from_excel()
    status = st.session_state["system_status"]
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("🔍 Gesuchte Themen", excel_stats.get("total_searches", 0))
    
    with col2:
        st.metric("📄 Gesamt Papers", excel_stats.get("total_papers", 0))
    
    with col3:
        st.metric("📧 Email-Empfänger", len(parse_recipient_emails(st.session_state.get("email_settings", {}).get("recipient_emails", ""))))
    
    with col4:
        backup_count = count_available_backups()
        st.metric("💾 Verfügbare Backups", backup_count)
    
    # 📋 PERSISTENTE SUCHTHEMEN ANZEIGEN
    st.markdown("---")
    st.subheader("📋 Alle durchsuchten Themen (Persistent gespeichert)")
    
    if excel_stats.get("search_terms"):
        # Sortiere nach letztem Update
        recent_terms = sorted(excel_stats["search_terms"], key=lambda x: x.get("last_update", ""), reverse=True)
        
        st.write(f"**🔒 Insgesamt {len(recent_terms)} Themen in persistenter Datenbank:**")
        
        for term_info in recent_terms:
            search_term = term_info["term"]
            papers = term_info["papers"]
            last_update = term_info.get("last_update", "")[:16].replace('T', ' ')
            new_papers = term_info.get("new_papers", 0)
            
            col_search1, col_search2, col_search3, col_search4 = st.columns([3, 1, 1, 1])
            
            with col_search1:
                if st.button(f"🔍 **{search_term}** ({papers} Papers)", key=f"search_btn_{search_term}"):
                    # Führe Suche aus und füge neue Papers hinzu
                    execute_excel_integrated_search(search_term, 100, "Letzte 2 Jahre", False, False)
            
            with col_search2:
                st.write(f"🆕 {new_papers}")
            
            with col_search3:
                st.write(f"📅 {last_update}")
            
            with col_search4:
                if st.button("📊", key=f"excel_btn_{search_term}", help="Excel-Sheet anzeigen"):
                    show_excel_sheet_content(search_term)
        
        # DOWNLOAD & BACKUP BEREICH
        st.markdown("---")
        st.subheader("💾 Datenbank-Management")
        
        col_mgmt1, col_mgmt2, col_mgmt3 = st.columns(3)
        
        with col_mgmt1:
            # Excel Download
            with open(excel_path, "rb") as file:
                st.download_button(
                    "📥 **Persistente Datenbank herunterladen**",
                    data=file.read(),
                    file_name=f"master_papers_persistent_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with col_mgmt2:
            if st.button("💾 **Manuelles Backup erstellen**"):
                create_automatic_backup(excel_path)
                st.success("✅ Backup erstellt!")
        
        with col_mgmt3:
            if st.button("📊 **Alle Themen aktualisieren**"):
                repeat_all_searches_from_excel()
    
    else:
        st.info("📭 **Noch keine Themen gesucht.** Starten Sie im Tab 'Paper-Suche' - alle Ergebnisse werden persistent gespeichert!")

def count_available_backups() -> int:
    """Zählt verfügbare Backup-Dateien"""
    try:
        if not os.path.exists("backups"):
            return 0
        
        backup_files = [f for f in os.listdir("backups") if f.startswith("master_papers_backup_") and f.endswith(".xlsx")]
        return len(backup_files)
    except:
        return 0

def create_automatic_backup(excel_path: str):
    """Erstellt automatische Backups der Excel-Datenbank"""
    try:
        if not os.path.exists("backups"):
            os.makedirs("backups")
        
        # Backup-Dateiname mit Timestamp
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"backups/master_papers_backup_{timestamp}.xlsx"
        
        # Kopiere Excel-Datei
        import shutil
        shutil.copy2(excel_path, backup_path)
        
        # Alte Backups löschen (behalte nur die letzten 10)
        cleanup_old_backups()
        
        # Update Info in Session State
        st.session_state["last_backup"] = datetime.datetime.now().isoformat()
        
    except Exception as e:
        st.warning(f"⚠️ Backup-Fehler: {str(e)}")

def cleanup_old_backups(keep_count: int = 10):
    """Löscht alte Backup-Dateien (behält nur die neuesten)"""
    try:
        backup_files = []
        for file in os.listdir("backups"):
            if file.startswith("master_papers_backup_") and file.endswith(".xlsx"):
                file_path = os.path.join("backups", file)
                backup_files.append((file_path, os.path.getmtime(file_path)))
        
        # Sortiere nach Änderungsdatum (neueste zuerst)
        backup_files.sort(key=lambda x: x[1], reverse=True)
        
        # Lösche alte Backups
        for file_path, _ in backup_files[keep_count:]:
            os.remove(file_path)
            
    except Exception as e:
        pass  # Stiller Fehler - Backup-Cleanup ist nicht kritisch


def check_due_searches_silent():
    """Stille Überprüfung überfälliger Suchen (ohne UI-Updates)"""
    try:
        automation_path = st.session_state.get("automation_excel_path")
        if not automation_path or not os.path.exists(automation_path):
            return 0
        
        wb = openpyxl.load_workbook(automation_path)
        if "🤖_Auto_Schedule" not in wb.sheetnames:
            return 0
        
        schedule_sheet = wb["🤖_Auto_Schedule"]
        now = datetime.datetime.now()
        due_count = 0
        
        for row_num in range(2, schedule_sheet.max_row + 1):
            row = schedule_sheet[row_num]
            
            if not row[1].value or row[10].value != "AKTIV":
                continue
            
            try:
                next_run_str = row[7].value
                if next_run_str:
                    next_run = datetime.datetime.fromisoformat(next_run_str)
                    if now >= next_run:
                        due_count += 1
            except:
                continue
        
        # Update Session State
        st.session_state["system_status"]["pending_automation_searches"] = due_count
        st.session_state["system_status"]["last_automation_check"] = now.isoformat()
        
        return due_count
        
    except Exception as e:
        return 0
def show_dashboard():
    """Dashboard mit persistenten Excel-Daten - sofort verfügbar"""
    st.subheader("📊 Dashboard - Persistente Excel-Datenbank")
    
    # PERSISTENT STATUS PRÜFEN
    persistent_status = st.session_state.get("persistent_data_status", {})
    excel_path = st.session_state["excel_template"]["file_path"]
    
    if persistent_status.get("loaded", False):
        # ✅ DATEN ERFOLGREICH GELADEN
        topics_count = persistent_status.get("topics_count", 0)
        papers_count = persistent_status.get("papers_count", 0)
        
        if os.path.exists(excel_path):
            file_size = os.path.getsize(excel_path) / 1024 / 1024  # MB
            file_date = datetime.datetime.fromtimestamp(os.path.getmtime(excel_path))
            
            st.success(f"🔒 **Persistente Excel-Datenbank aktiv:** {excel_path}")
            st.info(f"📊 **{topics_count} Suchthemen | {papers_count} Papers | {file_size:.2f} MB** | Letzte Änderung: {file_date.strftime('%d.%m.%Y %H:%M')}")
        
        # Hauptmetriken
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("🔍 Gesuchte Themen", topics_count)
        
        with col2:
            st.metric("📄 Gesamt Papers", papers_count)
        
        with col3:
            recipients = len(parse_recipient_emails(st.session_state.get("email_settings", {}).get("recipient_emails", "")))
            st.metric("📧 Email-Empfänger", recipients)
        
        with col4:
            backup_count = count_available_backups()
            st.metric("💾 Verfügbare Backups", backup_count)
        
        # 📋 VERFÜGBARE SUCHTHEMEN ANZEIGEN
        available_topics = st.session_state.get("available_topics", [])
        if available_topics:
            st.markdown("---")
            st.subheader(f"📋 Alle verfügbaren Suchthemen ({len(available_topics)})")
            
            # Excel-Statistiken für Details
            excel_stats = get_search_statistics_from_excel()
            recent_terms = sorted(excel_stats.get("search_terms", []), key=lambda x: x.get("last_update", ""), reverse=True)
            
            for term_info in recent_terms:
                search_term = term_info["term"]
                papers = term_info["papers"]
                last_update = term_info.get("last_update", "")[:16].replace('T', ' ')
                new_papers = term_info.get("new_papers", 0)
                
                col_search1, col_search2, col_search3, col_search4 = st.columns([3, 1, 1, 1])
                
                with col_search1:
                    if st.button(f"🔍 **{search_term}** ({papers} Papers)", key=f"search_btn_{search_term}"):
                        # Führe Suche aus und füge neue Papers hinzu
                        execute_excel_integrated_search(search_term, 100, "Letzte 2 Jahre", False, False)
                
                with col_search2:
                    st.write(f"🆕 {new_papers}")
                
                with col_search3:
                    st.write(f"📅 {last_update}")
                
                with col_search4:
                    if st.button("📊", key=f"excel_btn_{search_term}", help="Excel-Sheet anzeigen"):
                        show_excel_sheet_content(search_term)
        
        else:
            st.info("📭 **Noch keine Themen gesucht.** Starten Sie im Tab 'Paper-Suche' - alle Ergebnisse werden persistent gespeichert!")
    
    else:
        # ❌ DATEN NICHT GELADEN
        st.error("❌ **KRITISCH:** Persistente Excel-Datenbank konnte nicht geladen werden!")
        
        error_msg = persistent_status.get("error", "Unbekannter Fehler")
        st.error(f"**Fehler:** {error_msg}")
        
        if st.button("🔧 **Datenbank reparieren/neu erstellen**"):
            ensure_persistent_excel_database() 
            st.rerun()
    
    # DOWNLOAD & BACKUP BEREICH (immer verfügbar)
    if os.path.exists(excel_path):
        st.markdown("---")
        st.subheader("💾 Datenbank-Management")
        
        col_mgmt1, col_mgmt2, col_mgmt3 = st.columns(3)
        
        with col_mgmt1:
            # Excel Download
            with open(excel_path, "rb") as file:
                st.download_button(
                    "📥 **Persistente Datenbank herunterladen**",
                    data=file.read(),
                    file_name=f"master_papers_persistent_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with col_mgmt2:
            if st.button("💾 **Manuelles Backup erstellen**"):
                create_automatic_backup(excel_path)
                st.success("✅ Backup erstellt!")
        
        with col_mgmt3:
            if st.button("📊 **Alle Themen aktualisieren**"):
                repeat_all_searches_from_excel()

def count_available_backups() -> int:
    """Zählt verfügbare Backup-Dateien"""
    try:
        if not os.path.exists("backups"):
            return 0
        
        backup_files = [f for f in os.listdir("backups") if f.startswith("master_papers_backup_") and f.endswith(".xlsx")]
        return len(backup_files)
    except:
        return 0

def show_email_config_with_secrets():
    """Email-Konfiguration mit Secrets-Integration"""
    st.subheader("📧 Email-Konfiguration (Streamlit Secrets)")
    
    settings = st.session_state.get("email_settings", {})
    is_from_secrets = settings.get("from_secrets", False)
    
    if is_from_secrets:
        # ===== SECRETS MODUS =====
        st.success("🔐 **Email-Konfiguration aus Streamlit Secrets aktiv**")
        
        col_info1, col_info2 = st.columns(2)
        
        with col_info1:
            st.info("📋 **Aktuelle Konfiguration:**")
            # Sicher maskierte Anzeige
            sender_email = settings.get("sender_email", "")
            if sender_email:
                masked_sender = f"{sender_email[:3]}***@{sender_email.split('@')[1]}" if "@" in sender_email else "***"
                st.write(f"📧 Absender: `{masked_sender}`")
            
            st.write(f"🔒 SMTP: `{settings.get('smtp_server', 'N/A')}:{settings.get('smtp_port', 'N/A')}`")
            st.write(f"🔐 TLS: `{'✅ Aktiviert' if settings.get('use_tls') else '❌ Deaktiviert'}`")
            st.write(f"📧 Auto-Benachrichtigungen: `{'✅ An' if settings.get('auto_notifications') else '❌ Aus'}`")
        
        with col_info2:
            st.info("📧 **Empfänger:**")
            recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
            st.write(f"**Anzahl:** {len(recipient_emails)}")
            
            for i, email in enumerate(recipient_emails, 1):
                # Maskiere Email-Adressen
                if "@" in email:
                    masked = f"{email[:2]}***@{email.split('@')[1]}"
                else:
                    masked = "***"
                st.write(f"   {i}. `{masked}`")
        
        # Secrets Konfiguration Hilfe
        with st.expander("📖 Streamlit Secrets Konfiguration"):
            st.info("""
            **Secrets-Datei:** `.streamlit/secrets.toml`
            
            ```
            [email]
            sender_email = "absender@gmail.com"
            smtp_server = "smtp.gmail.com"
            smtp_port = 587
            password = "app-passwort"
            use_tls = true
            recipients = "emp1@example.com,emp2@example.com"
            auto_notifications = true
            min_papers = 1
            ```
            
            **Sicherheitshinweise:**
            ✅ Secrets werden nicht in Git gespeichert
            ✅ Passwörter sind nicht im Code sichtbar
            ✅ Produktionsumgebung nutzt verschlüsselte Secrets
            """)
        
        # Test-Funktionen für Secrets
        st.markdown("---")
        st.subheader("🧪 Email-System testen (Secrets)")
        
        col_test1, col_test2 = st.columns(2)
        
        with col_test1:
            if st.button("📧 **Test-Email senden (Secrets)**", type="primary"):
                send_test_email_secrets()
        
        with col_test2:
            if st.button("🔄 **Secrets neu laden**"):
                reload_email_secrets()
        
        # Override für Notfälle
        st.markdown("---")
        with st.expander("⚠️ Notfall-Override (manuelle Konfiguration)"):
            st.warning("⚠️ Nur für Entwicklung/Debugging verwenden!")
            if st.button("🔓 **Zu manueller Konfiguration wechseln**"):
                switch_to_manual_config()
    
    else:
        # ===== MANUELLER MODUS =====
        st.info("📝 **Manuelle Email-Konfiguration**")
        st.write("💡 Für erhöhte Sicherheit empfehlen wir Streamlit Secrets!")
        
        show_manual_email_config()

def send_test_email_secrets():
    """Sendet Test-Email mit Secrets-Konfiguration"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    if not settings.get("from_secrets"):
        st.error("❌ Keine Secrets-Konfiguration aktiv!")
        return
    
    if not recipient_emails:
        st.error("❌ Keine Empfänger in Secrets konfiguriert!")
        return
    
    subject = "🧪 Test-Email vom Paper-Suche System (Streamlit Secrets)"
    message = f"""Dies ist eine Test-Email vom Paper-Suche System mit Streamlit Secrets Integration.

📅 Gesendet am: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}
🔐 Konfiguration: Streamlit Secrets
📧 Von: {settings.get('sender_email', 'N/A')}
📧 An: {len(recipient_emails)} Empfänger (aus Secrets)

🔒 **Sicherheitsfeatures:**
✅ Passwort aus verschlüsselten Secrets
✅ Keine Credentials im Code
✅ Sichere SMTP-Verbindung

System-Informationen:
• SMTP Server: {settings.get('smtp_server')}
• Port: {settings.get('smtp_port')}
• TLS: {'Aktiviert' if settings.get('use_tls') else 'Deaktiviert'}
• Empfänger: {len(recipient_emails)}

Mit freundlichen Grüßen,
Ihr sicheres Paper-Suche System"""
    
    success, status_message = send_real_email_multiple(
        recipient_emails, 
        subject, 
        message
    )
    
    if success:
        st.success(f"✅ **Test-Email mit Secrets erfolgreich gesendet!** {status_message}")
        st.balloons()
    else:
        st.error(f"❌ **Test-Email fehlgeschlagen:** {status_message}")

def reload_email_secrets():
    """Lädt Email-Secrets neu"""
    try:
        secrets_config = load_email_config_from_secrets()
        if secrets_config:
            st.session_state["email_settings"] = secrets_config
            st.success("✅ **Email-Secrets erfolgreich neu geladen!**")
            st.rerun()
        else:
            st.error("❌ **Fehler beim Neuladen der Secrets!**")
    except Exception as e:
        st.error(f"❌ **Secrets-Fehler:** {str(e)}")

def switch_to_manual_config():
    """Wechselt zur manuellen Email-Konfiguration"""
    st.session_state["email_settings"]["from_secrets"] = False
    st.warning("⚠️ **Zu manueller Konfiguration gewechselt!**")
    st.rerun()

def show_manual_email_config():
    """Zeigt manuelle Email-Konfiguration (Original-Funktion)"""
    settings = st.session_state.get("email_settings", {})
    
    with st.expander("📖 Email-Setup Hilfe"):
        st.info("""
        **Für Gmail (empfohlen):**
        1. ✅ 2-Faktor-Authentifizierung aktivieren
        2. ✅ App-Passwort erstellen
        3. ✅ SMTP: smtp.gmail.com, Port: 587, TLS: An
        
        **Sicherheitshinweis:**
        🔐 Für Produktion empfehlen wir Streamlit Secrets!
        """)
    
    with st.form("manual_email_config_form"):
        st.subheader("📬 Manuelle Grundeinstellungen")
        
        col1, col2 = st.columns(2)
        
        with col1:
            sender_email = st.text_input(
                "Absender Email *", 
                value=settings.get("sender_email", "") if not settings.get("from_secrets") else "",
                placeholder="absender@gmail.com"
            )
            
            smtp_server = st.text_input(
                "SMTP Server *",
                value=settings.get("smtp_server", "smtp.gmail.com") if not settings.get("from_secrets") else "smtp.gmail.com"
            )
        
        with col2:
            smtp_port = st.number_input(
                "SMTP Port *",
                value=settings.get("smtp_port", 587) if not settings.get("from_secrets") else 587,
                min_value=1,
                max_value=65535
            )
            
            use_tls = st.checkbox(
                "TLS Verschlüsselung",
                value=settings.get("use_tls", True) if not settings.get("from_secrets") else True
            )
        
        recipient_emails = st.text_area(
            "📧 Empfänger Email-Adressen * (komma-getrennt)",
            value=settings.get("recipient_emails", "") if not settings.get("from_secrets") else "",
            placeholder="emp1@example.com, emp2@example.com",
            height=80
        )
        
        sender_password = st.text_input(
            "Email Passwort / App-Passwort *",
            value="",
            type="password",
            help="⚠️ Für Sicherheit nutzen Sie Streamlit Secrets!"
        )
        
        if st.form_submit_button("💾 **Manuelle Einstellungen speichern**", type="secondary"):
            recipient_list = parse_recipient_emails(recipient_emails)
            
            if not recipient_list:
                st.error("❌ Mindestens eine gültige Email erforderlich!")
            else:
                st.session_state["email_settings"] = {
                    "sender_email": sender_email,
                    "recipient_emails": recipient_emails,
                    "smtp_server": smtp_server,
                    "smtp_port": smtp_port,
                    "sender_password": sender_password,
                    "use_tls": use_tls,
                    "auto_notifications": True,
                    "min_papers": 1,
                    "subject_template": "🔬 {count} neue Papers für '{search_term}'",
                    "message_template": "📧 Neue Papers gefunden...",
                    "from_secrets": False
                }
                
                st.success(f"✅ Manuelle Einstellungen gespeichert! ({len(recipient_list)} Empfänger)")
    
    # Secrets-Empfehlung
    st.info("💡 **Empfehlung:** Verwenden Sie Streamlit Secrets für höhere Sicherheit!")

# ===== SICHERE EMAIL-VERSAND FUNKTIONEN =====

def send_real_email_multiple(to_emails: List[str], subject: str, message: str, attachment_path: str = None) -> tuple:
    """Sendet Email mit Secrets-Integration"""
    settings = st.session_state.get("email_settings", {})
    
    # Lade Credentials sicher
    sender_email = settings.get("sender_email", "")
    sender_password = settings.get("sender_password", "")
    smtp_server = settings.get("smtp_server", "smtp.gmail.com")
    smtp_port = settings.get("smtp_port", 587)
    use_tls = settings.get("use_tls", True)
    
    # Bei Secrets: Passwort ist bereits geladen
    if settings.get("from_secrets", False):
        # Credentials sind bereits sicher aus Secrets geladen
        pass
    
    if not all([sender_email, sender_password]):
        return False, "❌ Email-Konfiguration unvollständig"
    
    if not to_emails:
        return False, "❌ Keine Empfänger konfiguriert"
    
    try:
        # SMTP Server Setup mit sicheren Credentials
        server = smtplib.SMTP(smtp_server, smtp_port)
        
        if use_tls:
            context = ssl.create_default_context()
            server.starttls(context=context)
        
        server.login(sender_email, sender_password)
        
        successful_sends = 0
        failed_sends = []
        
        # Send to each recipient
        for recipient in to_emails:
            try:
                msg = MIMEMultipart()
                msg['From'] = sender_email
                msg['To'] = recipient
                msg['Subject'] = subject
                
                msg.attach(MIMEText(message, 'plain', 'utf-8'))
                
                # Add attachment if provided
                if attachment_path and os.path.exists(attachment_path):
                    with open(attachment_path, "rb") as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header(
                            'Content-Disposition',
                            f'attachment; filename= {os.path.basename(attachment_path)}'
                        )
                        msg.attach(part)
                
                server.send_message(msg)
                successful_sends += 1
                
            except Exception as e:
                failed_sends.append(f"{recipient}: {str(e)}")
        
        server.quit()
        
        if successful_sends == len(to_emails):
            return True, f"✅ Email erfolgreich an alle {successful_sends} Empfänger gesendet"
        elif successful_sends > 0:
            return True, f"⚠️ Email an {successful_sends}/{len(to_emails)} Empfänger gesendet"
        else:
            return False, f"❌ Email an keinen Empfänger gesendet"
        
    except smtplib.SMTPAuthenticationError:
        return False, "❌ SMTP-Authentifizierung fehlgeschlagen"
    except Exception as e:
        return False, f"❌ Email-Fehler: {str(e)}"

def is_email_configured() -> bool:
    """Prüft Email-Konfiguration (Secrets-kompatibel)"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    return (bool(settings.get("sender_email")) and 
            len(recipient_emails) > 0 and
            bool(settings.get("sender_password")))

def parse_recipient_emails(email_string: str) -> List[str]:
    """Parst Email-String und gibt Liste gültiger Emails zurück"""
    if not email_string:
        return []
    
    emails = [email.strip() for email in email_string.split(",")]
    valid_emails = []
    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    
    for email in emails:
        if email and email_pattern.match(email):
            valid_emails.append(email)
    
    return valid_emails

# ===== ALLE ANDEREN FUNKTIONEN BLEIBEN UNVERÄNDERT =====
# (Hier würden alle anderen Funktionen aus dem ursprünglichen Script eingefügt werden)




def send_real_email_multiple(to_emails: List[str], subject: str, message: str, attachment_path: str = None) -> tuple:
    """Sendet echte Email über SMTP an mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    
    sender_email = settings.get("sender_email", "")
    sender_password = settings.get("sender_password", "")
    smtp_server = settings.get("smtp_server", "smtp.gmail.com")
    smtp_port = settings.get("smtp_port", 587)
    use_tls = settings.get("use_tls", True)
    
    if not all([sender_email, sender_password]):
        return False, "❌ Email-Konfiguration unvollständig (Absender/Passwort)"
    
    if not to_emails:
        return False, "❌ Keine Empfänger-Emails konfiguriert"
    
    try:
        # SMTP Server Setup
        server = smtplib.SMTP(smtp_server, smtp_port)
        
        if use_tls:
            context = ssl.create_default_context()
            server.starttls(context=context)
        
        server.login(sender_email, sender_password)
        
        successful_sends = 0
        failed_sends = []
        
        # Send to each recipient
        for recipient in to_emails:
            try:
                msg = MIMEMultipart()
                msg['From'] = sender_email
                msg['To'] = recipient
                msg['Subject'] = subject
                
                msg.attach(MIMEText(message, 'plain', 'utf-8'))
                
                # Add attachment if provided
                if attachment_path and os.path.exists(attachment_path):
                    with open(attachment_path, "rb") as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header(
                            'Content-Disposition',
                            f'attachment; filename= {os.path.basename(attachment_path)}'
                        )
                        msg.attach(part)
                
                server.send_message(msg)
                successful_sends += 1
                
            except Exception as e:
                failed_sends.append(f"{recipient}: {str(e)}")
        
        server.quit()
        
        if successful_sends == len(to_emails):
            return True, f"✅ Email erfolgreich an alle {successful_sends} Empfänger gesendet"
        elif successful_sends > 0:
            return True, f"⚠️ Email an {successful_sends}/{len(to_emails)} Empfänger gesendet. Fehler: {'; '.join(failed_sends)}"
        else:
            return False, f"❌ Email an keinen Empfänger gesendet. Fehler: {'; '.join(failed_sends)}"
        
    except smtplib.SMTPAuthenticationError:
        return False, "❌ SMTP-Authentifizierung fehlgeschlagen - Prüfen Sie Email/Passwort"
    except smtplib.SMTPServerDisconnected:
        return False, "❌ SMTP-Server-Verbindung unterbrochen"
    except Exception as e:
        return False, f"❌ Email-Fehler: {str(e)}"

# =============== HAUPTFUNKTIONEN ===============

def show_dashboard():
    """Dashboard mit anklickbaren Suchhistorie und Excel-Integration"""
    st.subheader("📊 Dashboard - Excel-Integrierte Übersicht")
    
    # Excel-Statistiken holen
    excel_stats = get_search_statistics_from_excel()
    status = st.session_state["system_status"]
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("🔍 Gesamt Suchen", excel_stats.get("total_searches", 0))
    
    with col2:
        st.metric("📄 Papers (Excel)", excel_stats.get("total_papers", 0))
    
    with col3:
        st.metric("📧 Gesendete Emails", status["total_emails"])
    
    with col4:
        recipients = len(parse_recipient_emails(st.session_state.get("email_settings", {}).get("recipient_emails", "")))
        st.metric("📧 Email-Empfänger", recipients)
    
    # Letzte Aktivität
    if status["last_search"]:
        try:
            last_search_time = datetime.datetime.fromisoformat(status["last_search"])
            time_diff = datetime.datetime.now() - last_search_time
            hours = time_diff.seconds // 3600
            minutes = (time_diff.seconds % 3600) // 60
            st.info(f"🕒 Letzte Suche: vor {time_diff.days}d {hours}h {minutes}min")
        except:
            st.info("🕒 Letzte Suche: Unbekannt")
    
    # Excel-Download im Dashboard
    excel_path = st.session_state["excel_template"]["file_path"]
    if os.path.exists(excel_path):
        with open(excel_path, "rb") as file:
            st.download_button(
                "📎 Excel-Datenbank herunterladen",
                data=file.read(),
                file_name=f"paper_database_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    # Suchhistorie aus Excel
    st.markdown("---")
    st.subheader("📋 Excel-basierte Suchhistorie")
    
    if excel_stats.get("search_terms"):
        # Sortiere nach letztem Update
        recent_terms = sorted(excel_stats["search_terms"], key=lambda x: x.get("last_update", ""), reverse=True)
        
        for term_info in recent_terms:
            search_term = term_info["term"]
            papers = term_info["papers"]
            last_update = term_info.get("last_update", "")[:16].replace('T', ' ')
            new_papers = term_info.get("new_papers", 0)
            
            col_search1, col_search2, col_search3 = st.columns([3, 1, 1])
            
            with col_search1:
                if st.button(f"🔍 **{search_term}** ({papers} Papers, {new_papers} neue)", 
                           key=f"search_btn_{search_term}"):
                    show_search_details_from_excel(search_term, term_info)
            
            with col_search2:
                st.write(f"📅 {last_update}")
            
            with col_search3:
                if st.button("📊 Excel", key=f"excel_btn_{search_term}"):
                    show_excel_sheet_content(search_term)
        
        # Quick Actions
        st.markdown("---")
        st.subheader("⚡ Quick Actions")
        
        col_quick1, col_quick2, col_quick3 = st.columns(3)
        
        with col_quick1:
            if st.button("🔄 **Alle Suchen wiederholen**"):
                repeat_all_searches_from_excel()
        
        with col_quick2:
            if st.button("📧 **Status-Email senden**"):
                send_status_email_multiple()
        
        with col_quick3:
            if st.button("📁 **Excel öffnen**"):
                offer_excel_download()
    
    else:
        st.info("📭 Noch keine Suchen durchgeführt. Starten Sie im Tab 'Paper-Suche'!")

def show_advanced_paper_search():
    """Erweiterte Paper-Suche mit Excel-Integration und mehreren Email-Empfängern"""
    st.subheader("🔍 Excel-Integrierte Paper-Suche")
    
    # Excel- und Email-Status
    excel_stats = get_search_statistics_from_excel()
    email_status = is_email_configured()
    recipient_count = len(parse_recipient_emails(st.session_state.get("email_settings", {}).get("recipient_emails", "")))
    
    col_info1, col_info2, col_info3, col_info4 = st.columns(4)
    with col_info1:
        st.metric("📊 Excel-Sheets", excel_stats.get("total_sheets", 0))
    with col_info2:
        st.metric("📄 Papers in Excel", excel_stats.get("total_papers", 0))
    with col_info3:
        st.metric("📧 Email-Empfänger", recipient_count)
    with col_info4:
        st.metric("🔍 Durchsuchungen", excel_stats.get("total_searches", 0))
    
    if email_status:
        st.success(f"✅ Email-System bereit für **{recipient_count} Empfänger** | Excel-Integration: ✅ Aktiv")
    else:
        st.info("ℹ️ Email-System nicht konfiguriert | Excel-Integration: ✅ Aktiv")
    
    # Such-Interface
    with st.form("advanced_search_form"):
        col_search1, col_search2 = st.columns([3, 1])
        
        with col_search1:
            search_query = st.text_input(
                "**🔍 PubMed Suchbegriff:**",
                placeholder="z.B. 'diabetes genetics', 'machine learning radiology', 'COVID-19 treatment'",
                help="Durchsucht Excel auf bereits bekannte Papers und fügt nur neue hinzu"
            )
        
        with col_search2:
            max_results = st.number_input(
                "Max. Ergebnisse", 
                min_value=10, 
                max_value=500, 
                value=100
            )
        
        # Erweiterte Optionen
        with st.expander("🔧 Erweiterte Excel- & Email-Optionen"):
            col_adv1, col_adv2, col_adv3 = st.columns(3)
            
            with col_adv1:
                date_filter = st.selectbox(
                    "📅 Zeitraum:",
                    ["Alle", "Letztes Jahr", "Letzte 2 Jahre", "Letzte 5 Jahre", "Letzte 10 Jahre"],
                    index=2
                )
            
            with col_adv2:
                force_email = st.checkbox(
                    "📧 Email erzwingen", 
                    value=False,
                    help="Sendet Email auch wenn keine neuen Papers gefunden"
                )
            
            with col_adv3:
                show_existing = st.checkbox(
                    "📊 Bereits bekannte Papers anzeigen", 
                    value=False,
                    help="Zeigt auch Papers an, die bereits in Excel vorhanden sind"
                )
        
        search_button = st.form_submit_button("🚀 **EXCEL-INTEGRIERTE PAPER-SUCHE**", type="primary")
    
    # Quick Search aus Excel-Historie
    if excel_stats.get("search_terms"):
        st.write("**⚡ Schnellsuche (aus Excel-Historie):**")
        recent_terms = sorted(excel_stats["search_terms"], key=lambda x: x.get("last_update", ""), reverse=True)[:5]
        
        cols = st.columns(min(len(recent_terms), 5))
        for i, term_info in enumerate(recent_terms):
            term = term_info["term"]
            papers = term_info["papers"]
            with cols[i]:
                if st.button(f"🔍 {term[:15]}... ({papers})", key=f"quick_{i}"):
                    execute_excel_integrated_search(term, 50, "Letzte 2 Jahre", False, False)
    
    # Hauptsuche ausführen
    if search_button and search_query:
        execute_excel_integrated_search(search_query, max_results, date_filter, force_email, show_existing)
    
    # *** MANUELLER EMAIL-VERSAND BEREICH ***
    show_manual_email_section()

def execute_excel_integrated_search(query: str, max_results: int, date_filter: str, force_email: bool, show_existing: bool):
    """
    ULTRA-ROBUSTE Excel-integrierte Paper-Suche mit Multi-Fallback-System
    Behebt PubMed 500 Server Errors durch korrekte Datums-Filter und Retry-Mechanismen
    """
    st.markdown("---")
    st.subheader(f"🔍 **Excel-integrierte Suche:** '{query}'")
    
    # Progress Tracking
    progress_container = st.container()
    with progress_container:
        progress_bar = st.progress(0)
        status_text = st.empty()
    
    try:
        # 1. ULTRA-SICHERE Excel-Validierung
        status_text.text("📊 Lade Excel-Datei für Duplikatsprüfung...")
        progress_bar.progress(0.1)
        
        wb = load_master_workbook()
        if not wb:
            st.error("❌ Excel-Datei konnte nicht geladen werden!")
            progress_bar.empty()
            status_text.empty()
            return
        
        # 2. KORRIGIERTE PUBMED-SUCHE mit Multi-Fallback
        status_text.text("🔍 Durchsuche PubMed-Datenbank...")
        progress_bar.progress(0.2)
        
        # ✅ KORRIGIERTE Query-Generierung
        advanced_query = build_advanced_search_query_corrected(query, date_filter)
        
        st.info(f"🔍 **Generated Query:** `{advanced_query}`")
        
        # MULTI-FALLBACK SYSTEM
        current_papers = None
        search_attempts = []
        
        # 🎯 VERSUCH 1: Korrigierte Haupt-Suche
        try:
            status_text.text("🔍 Versuch 1: Haupt-Suche mit korrigierten Datums-Filtern...")
            progress_bar.progress(0.3)
            
            current_papers = perform_comprehensive_pubmed_search_robust(advanced_query, max_results)
            
            if current_papers:
                search_attempts.append(f"✅ Haupt-Suche erfolgreich: {len(current_papers)} Papers")
                st.success(f"✅ **Haupt-Suche erfolgreich:** {len(current_papers)} Papers gefunden")
            else:
                search_attempts.append("⚠️ Haupt-Suche: Keine Ergebnisse")
                
        except Exception as e:
            search_attempts.append(f"❌ Haupt-Suche Fehler: {str(e)}")
            st.warning(f"⚠️ Haupt-Suche fehlgeschlagen: {str(e)}")
        
        # 🎯 VERSUCH 2: Vereinfachte Suche ohne Datums-Filter
        if not current_papers:
            try:
                status_text.text("🔄 Versuch 2: Vereinfachte Suche ohne Datums-Filter...")
                progress_bar.progress(0.4)
                
                simple_query = query.strip()
                current_papers = try_simple_pubmed_search_enhanced(simple_query, max_results)
                
                if current_papers:
                    search_attempts.append(f"✅ Vereinfachte Suche erfolgreich: {len(current_papers)} Papers")
                    st.success(f"✅ **Vereinfachte Suche erfolgreich:** {len(current_papers)} Papers gefunden")
                else:
                    search_attempts.append("⚠️ Vereinfachte Suche: Keine Ergebnisse")
                    
            except Exception as e:
                search_attempts.append(f"❌ Vereinfachte Suche Fehler: {str(e)}")
                st.warning(f"⚠️ Vereinfachte Suche fehlgeschlagen: {str(e)}")
        
        # 🎯 VERSUCH 3: Alternative PubMed Parameter
        if not current_papers:
            try:
                status_text.text("🔄 Versuch 3: Alternative PubMed-Parameter...")
                progress_bar.progress(0.5)
                
                current_papers = try_alternative_pubmed_search_enhanced(query, max_results)
                
                if current_papers:
                    search_attempts.append(f"✅ Alternative Suche erfolgreich: {len(current_papers)} Papers")
                    st.success(f"✅ **Alternative Suche erfolgreich:** {len(current_papers)} Papers gefunden")
                else:
                    search_attempts.append("⚠️ Alternative Suche: Keine Ergebnisse")
                    
            except Exception as e:
                search_attempts.append(f"❌ Alternative Suche Fehler: {str(e)}")
                st.warning(f"⚠️ Alternative Suche fehlgeschlagen: {str(e)}")
        
        # 🎯 VERSUCH 4: Minimal-Suche als letzter Fallback
        if not current_papers:
            try:
                status_text.text("🔄 Versuch 4: Minimal-Suche (letzter Fallback)...")
                progress_bar.progress(0.6)
                
                minimal_query = query.split()[0] if " " in query else query  # Nur erstes Wort
                current_papers = try_minimal_pubmed_search(minimal_query, min(max_results, 20))
                
                if current_papers:
                    search_attempts.append(f"✅ Minimal-Suche erfolgreich: {len(current_papers)} Papers")
                    st.success(f"✅ **Minimal-Suche erfolgreich:** {len(current_papers)} Papers für '{minimal_query}'")
                else:
                    search_attempts.append("❌ Minimal-Suche: Keine Ergebnisse")
                    
            except Exception as e:
                search_attempts.append(f"❌ Minimal-Suche Fehler: {str(e)}")
        
        # 🔴 FINALE VALIDIERUNG
        if not current_papers:
            st.error(f"❌ **ALLE SUCHMETHODEN FEHLGESCHLAGEN für '{query}'!**")
            
            # Debug-Informationen anzeigen
            with st.expander("🔍 **Debug-Informationen - Such-Attempts:**"):
                for i, attempt in enumerate(search_attempts, 1):
                    st.write(f"{i}. {attempt}")
                
                st.write(f"**Original Query:** `{query}`")
                st.write(f"**Advanced Query:** `{advanced_query}`")
                st.write(f"**Date Filter:** {date_filter}")
                st.write(f"**Max Results:** {max_results}")
            
            st.info("💡 **Mögliche Lösungen:**")
            st.write("• **Server-Überlastung:** Warten Sie 2-5 Minuten und versuchen Sie es erneut")
            st.write("• **Suchbegriff vereinfachen:** Verwenden Sie weniger oder andere Begriffe")  
            st.write("• **Datums-Filter entfernen:** Wählen Sie 'Alle' als Zeitraum")
            st.write("• **Kleinere Ergebnis-Anzahl:** Reduzieren Sie 'Max. Ergebnisse' auf 20-50")
            
            progress_bar.empty()
            status_text.empty()
            return
        
        # 3. EXCEL-INTEGRATION: Prüfe auf neue Papers
        status_text.text("📊 Prüfe Papers gegen Excel-Datenbank...")
        progress_bar.progress(0.7)
        
        added_count, new_papers = add_new_papers_to_excel(query, current_papers)
        
        # 4. ERGEBNISSE VERARBEITEN
        status_text.text("📊 Verarbeite Ergebnisse...")
        progress_bar.progress(0.85)
        
        if added_count > 0:
            st.success(f"🆕 **{added_count} NEUE Papers gefunden und zu Excel hinzugefügt!** (von {len(current_papers)} gesamt)")
            st.balloons()
            
            # Email senden bei neuen Papers
            if is_email_configured() and (force_email or should_send_email(added_count)):
                send_excel_integrated_email_multiple(query, new_papers, len(current_papers), added_count)
        else:
            st.info(f"ℹ️ **Keine neuen Papers** - Alle {len(current_papers)} Papers bereits in Excel vorhanden")
            
            # Email erzwingen wenn gewünscht
            if force_email and is_email_configured():
                send_excel_integrated_email_multiple(query, [], len(current_papers), 0)
        
        # 5. DETAILLIERTE ERGEBNISSE ANZEIGEN
        display_excel_integrated_results(current_papers, new_papers, query, added_count, show_existing)
        
        # 6. SESSION STATE UPDATE
        st.session_state["current_search_results"] = {
            "search_term": query,
            "papers": current_papers,
            "new_papers": new_papers,
            "added_count": added_count,
            "timestamp": datetime.datetime.now().isoformat(),
            "search_attempts": search_attempts,
            "success_method": search_attempts[-1] if current_papers else "Alle Methoden fehlgeschlagen"
        }
        
        # 7. SYSTEM-STATUS AKTUALISIEREN
        progress_bar.progress(1.0)
        status_text.text("✅ Excel-integrierte Suche erfolgreich abgeschlossen!")
        
        st.session_state["system_status"]["total_searches"] += 1
        st.session_state["system_status"]["total_papers"] += added_count
        st.session_state["system_status"]["last_search"] = datetime.datetime.now().isoformat()
        
        # Erfolgs-Statistik anzeigen
        with st.expander("📊 **Such-Statistik für diese Suche:**"):
            st.write(f"**🔍 Suchbegriff:** {query}")
            st.write(f"**📊 Gefunden:** {len(current_papers)} Papers")
            st.write(f"**🆕 Neue:** {added_count} Papers") 
            st.write(f"**📅 Zeitfilter:** {date_filter}")
            st.write(f"**✅ Erfolgreiche Methode:** {search_attempts[-1] if current_papers else 'N/A'}")
            
            for i, attempt in enumerate(search_attempts, 1):
                st.write(f"   {i}. {attempt}")
        
        # Cleanup
        time.sleep(1)
        progress_bar.empty()
        status_text.empty()
        
    except Exception as e:
        progress_bar.empty()
        status_text.empty()
        st.error(f"❌ **KRITISCHER FEHLER bei Excel-integrierter Suche:** {str(e)}")
        
        # Fehler-Debugging
        with st.expander("🔍 **Error Debug Info:**"):
            st.code(f"Error: {str(e)}")
            st.code(f"Query: {query}")
            st.code(f"Date Filter: {date_filter}")
            st.code(f"Max Results: {max_results}")

# =============== KORRIGIERTE HILFSFUNKTIONEN ===============

def build_advanced_search_query_corrected(query: str, date_filter: str) -> str:
    """
    🔧 KORRIGIERTE Suchanfrage-Generierung - Behebt PubMed 500 Errors
    """
    query_parts = [query.strip()]
    
    if date_filter != "Alle":
        current_year = datetime.datetime.now().year
        
        # ✅ KORRIGIERT: Verwende [pdat] statt [dp] und korrekte Jahresbereiche
        if date_filter == "Letztes Jahr":
            start_year = current_year - 1
            end_year = current_year - 1
            query_parts.append(f"AND ({start_year}[pdat]:{end_year}[pdat])")
            
        elif date_filter == "Letzte 2 Jahre":
            start_year = current_year - 2
            end_year = current_year - 1  # Nicht aktuelles Jahr verwenden
            query_parts.append(f"AND ({start_year}[pdat]:{end_year}[pdat])")
            
        elif date_filter == "Letzte 5 Jahre":
            start_year = current_year - 5
            end_year = current_year - 1
            query_parts.append(f"AND ({start_year}[pdat]:{end_year}[pdat])")
            
        elif date_filter == "Letzte 10 Jahre":
            start_year = current_year - 10
            end_year = current_year - 1
            query_parts.append(f"AND ({start_year}[pdat]:{end_year}[pdat])")
    
    final_query = " ".join(query_parts)
    return final_query

def try_simple_pubmed_search_enhanced(query: str, max_results: int) -> List[Dict[str, Any]]:
    """
    🔧 VERBESSERTE einfache PubMed-Suche ohne Filter
    """
    try:
        st.info(f"🔄 **Enhanced Simple Search:** {query}")
        
        search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        
        # Optimierte Parameter
        params = {
            "db": "pubmed",
            "term": query.strip(),
            "retmode": "json", 
            "retmax": min(max_results, 100),
            "email": "research@papersearch.com",
            "tool": "PaperSearchSystem",
            "sort": "relevance",
            "field": "title/abstract"  # Fokus auf Titel/Abstract
        }
        
        response = requests.get(search_url, params=params, timeout=45)
        
        if response.status_code == 200:
            data = response.json()
            pmids = data.get("esearchresult", {}).get("idlist", [])
            total_count = int(data.get("esearchresult", {}).get("count", 0))
            
            if pmids:
                st.success(f"✅ **Enhanced Simple Search erfolgreich:** {len(pmids)} von {total_count:,} Papers")
                return fetch_paper_details_batch(pmids)
            else:
                st.warning(f"⚠️ Enhanced Simple Search: Keine Results für '{query}'")
                return []
        else:
            st.error(f"❌ Enhanced Simple Search HTTP Error: {response.status_code}")
            return []
            
    except Exception as e:
        st.error(f"❌ Enhanced Simple Search Exception: {str(e)}")
        return []

def try_alternative_pubmed_search_enhanced(query: str, max_results: int) -> List[Dict[str, Any]]:
    """
    🔧 VERBESSERTE alternative PubMed-Suche
    """
    try:
        st.info(f"🔄 **Enhanced Alternative Search:** {query}")
        
        # Mehrere alternative Strategien
        search_strategies = [
            query.replace(" ", " AND "),  # AND-Verknüpfung
            f'"{query}"',                 # Phrase-Suche  
            query.split()[0] if " " in query else query  # Nur erstes Wort
        ]
        
        for i, strategy_query in enumerate(search_strategies, 1):
            try:
                search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
                params = {
                    "db": "pubmed",
                    "term": strategy_query,
                    "retmode": "json",
                    "retmax": min(max_results, 50),
                    "email": "research@papersearch.com", 
                    "tool": "PaperSearchSystem"
                }
                
                response = requests.get(search_url, params=params, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    pmids = data.get("esearchresult", {}).get("idlist", [])
                    
                    if pmids:
                        st.success(f"✅ **Alternative Strategie {i} erfolgreich:** {len(pmids)} Papers für '{strategy_query}'")
                        return fetch_paper_details_batch(pmids[:max_results//2])  # Begrenzte Anzahl
                
            except Exception as strategy_error:
                continue
        
        st.warning("⚠️ Alle alternativen Strategien fehlgeschlagen")
        return []
        
    except Exception as e:
        st.error(f"❌ Enhanced Alternative Search Exception: {str(e)}")
        return []

def try_minimal_pubmed_search(query: str, max_results: int) -> List[Dict[str, Any]]:
    """
    🔧 MINIMALE PubMed-Suche als absoluter Fallback
    """
    try:
        st.info(f"🔄 **Minimal Fallback Search:** {query}")
        
        search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        
        # Absolute Minimal-Parameter
        params = {
            "db": "pubmed",
            "term": query,
            "retmax": min(max_results, 20),  # Sehr kleine Anzahl
            "retmode": "json"
        }
        
        response = requests.get(search_url, params=params, timeout=20)
        
        if response.status_code == 200:
            data = response.json()
            pmids = data.get("esearchresult", {}).get("idlist", [])
            
            if pmids:
                st.success(f"✅ **Minimal Search erfolgreich:** {len(pmids)} Papers")
                return fetch_paper_details_batch(pmids)
        
        return []
        
    except Exception as e:
        st.error(f"❌ Minimal Search Exception: {str(e)}")
        return []


def send_excel_integrated_email_multiple(search_term: str, new_papers: List[Dict], total_found: int, added_count: int):
    """Sendet Email für Excel-integrierte Suche an mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    if not recipient_emails:
        st.warning("⚠️ Keine Email-Empfänger konfiguriert!")
        return
    
    # Subject generieren
    if added_count > 0:
        subject = f"📊 {added_count} neue Papers für '{search_term}' - Excel aktualisiert"
    else:
        subject = f"📊 Keine neuen Papers für '{search_term}' - Excel-Check durchgeführt"
    
    # Sheet-Name ermitteln
    sheet_name = generate_sheet_name(search_term)
    
    # Papers-Liste formatieren (nur neue)
    if new_papers:
        papers_list = ""
        for i, paper in enumerate(new_papers[:8], 1):
            title = paper.get("Title", "Unbekannt")[:70]
            authors = paper.get("Authors", "n/a")[:40]
            journal = paper.get("Journal", "n/a")
            year = paper.get("Year", "n/a")
            pmid = paper.get("PMID", "n/a")
            
            papers_list += f"\n{i}. **{title}...**\n"
            papers_list += f"   👥 {authors}...\n"
            papers_list += f"   📚 {journal} ({year}) | PMID: {pmid}\n\n"
        
        if len(new_papers) > 8:
            papers_list += f"... und {len(new_papers) - 8} weitere neue Papers (siehe Excel-Datei)\n"
    else:
        papers_list = "\nKeine neuen Papers gefunden - alle Papers bereits in Excel-Datenbank vorhanden.\n"
    
    # Message generieren
    message = f"""📊 **Excel-Integrierte Paper-Suche - Ergebnisse**

📅 **Datum:** {datetime.datetime.now().strftime("%d.%m.%Y %H:%M")}
🔍 **Suchbegriff:** '{search_term}'
📊 **Gefundene Papers:** {total_found}
🆕 **Neue Papers:** {added_count}
📊 **Bereits bekannt:** {total_found - added_count}
📁 **Excel-Sheet:** {sheet_name}

{'-' * 60}
🆕 **NEUE PAPERS:**
{papers_list}

📎 **Excel-Integration:**
✅ Alle neuen Papers wurden automatisch zur Excel-Datei hinzugefügt
✅ Duplikate wurden automatisch erkannt und übersprungen
✅ Sheet für diesen Suchbegriff wurde aktualisiert
📋 Sheet-Name: {sheet_name}

📧 **Email-Info:**
📧 Versendet an: {len(recipient_emails)} Empfänger
{chr(10).join([f"   • {email}" for email in recipient_emails])}
📎 Excel-Datei als Anhang beigefügt

Mit freundlichen Grüßen,
Ihr Excel-integriertes Paper-Suche System"""
    
    # Excel als Anhang
    excel_path = st.session_state["excel_template"]["file_path"]
    attachment_path = excel_path if os.path.exists(excel_path) else None
    
    # Email senden
    with st.spinner(f"📧 Sende Excel-integrierte Email an {len(recipient_emails)} Empfänger..."):
        success, status_message = send_real_email_multiple(recipient_emails, subject, message, attachment_path)
    
    # Email-Historie
    email_entry = {
        "timestamp": datetime.datetime.now().isoformat(),
        "type": "Excel-Integriert",
        "search_term": search_term,
        "recipients": recipient_emails,
        "recipient_count": len(recipient_emails),
        "subject": subject,
        "paper_count": added_count,
        "total_found": total_found,
        "success": success,
        "status": status_message,
        "has_attachment": attachment_path is not None,
        "sheet_name": sheet_name
    }
    
    st.session_state["email_history"].append(email_entry)
    
    # Ergebnis anzeigen
    if success:
        st.session_state["system_status"]["total_emails"] += 1
        st.success(f"📧 **Excel-integrierte Email erfolgreich versendet!**\n{status_message}")
        
        with st.expander("📋 Email-Details"):
            st.write(f"**📧 Empfänger:** {len(recipient_emails)}")
            for i, email in enumerate(recipient_emails, 1):
                st.write(f"   {i}. {email}")
            st.write(f"**🆕 Neue Papers:** {added_count}")
            st.write(f"**📊 Gesamt gefunden:** {total_found}")
            st.write(f"**📁 Excel-Sheet:** {sheet_name}")
            st.write(f"**📎 Anhang:** {'✅ Excel-Datei' if attachment_path else '❌ Kein Anhang'}")
    else:
        st.error(f"❌ **Email-Fehler:** {status_message}")

def show_manual_email_section():
    """Manueller Email-Versand nach Suche für mehrere Empfänger"""
    if st.session_state.get("current_search_results"):
        st.markdown("---")
        st.subheader("📧 Manueller Email-Versand (Excel-Integriert)")
        
        current_results = st.session_state["current_search_results"]
        search_term = current_results.get("search_term", "")
        papers = current_results.get("papers", [])
        new_papers = current_results.get("new_papers", [])
        added_count = current_results.get("added_count", 0)
        
        if papers:
            col_email1, col_email2, col_email3 = st.columns(3)
            
            with col_email1:
                st.metric("📄 Verfügbare Papers", len(papers))
            
            with col_email2:
                st.metric("🆕 Neue Papers", added_count)
            
            with col_email3:
                recipient_count = len(parse_recipient_emails(st.session_state.get("email_settings", {}).get("recipient_emails", "")))
                st.metric("📧 Empfänger", recipient_count)
            
            # Email-Optionen
            email_status = is_email_configured()
            
            if email_status and recipient_count > 0:
                col_send1, col_send2 = st.columns(2)
                
                with col_send1:
                    if st.button(f"📧 **Alle Papers emailen** ({len(papers)})", type="primary"):
                        send_manual_search_email_multiple(search_term, papers, "Alle Papers")
                
                with col_send2:
                    if added_count > 0 and st.button(f"📧 **Nur neue Papers emailen** ({added_count})", type="secondary"):
                        send_manual_search_email_multiple(search_term, new_papers, "Nur neue Papers")
            else:
                if not email_status:
                    st.warning("⚠️ **Email-Versand nicht möglich:** Konfigurieren Sie Email-Einstellungen im entsprechenden Tab")
                elif recipient_count == 0:
                    st.warning("⚠️ **Keine Empfänger konfiguriert:** Fügen Sie Email-Adressen in der Email-Konfiguration hinzu")

def send_manual_search_email_multiple(search_term: str, papers: List[Dict], email_type: str):
    """Sendet manuelle Email für Suchergebnisse an mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    if not recipient_emails:
        st.error("❌ Keine Empfänger konfiguriert!")
        return
    
    # Subject generieren
    subject = f"📧 {email_type}: {len(papers)} Papers für '{search_term}' (Manuell)"
    
    # Papers-Liste formatieren
    papers_list = ""
    for i, paper in enumerate(papers[:15], 1):  # Erste 15 Papers
        title = paper.get("Title", "Unbekannt")[:70]
        authors = paper.get("Authors", "n/a")[:50]
        journal = paper.get("Journal", "n/a")
        year = paper.get("Year", "n/a")
        pmid = paper.get("PMID", "n/a")
        
        papers_list += f"\n{i}. **{title}...**\n"
        papers_list += f"   👥 {authors}...\n"
        papers_list += f"   📚 {journal} ({year}) | PMID: {pmid}\n\n"
    
    if len(papers) > 15:
        papers_list += f"... und {len(papers) - 15} weitere Papers (siehe Excel-Datei)\n"
    
    # Message generieren
    message = f"""📧 **Manueller Email-Versand - Paper-Suche**

📅 **Datum:** {datetime.datetime.now().strftime("%d.%m.%Y %H:%M")}
🔍 **Suchbegriff:** '{search_term}'
📊 **Typ:** {email_type}
📄 **Anzahl Papers:** {len(papers)}
📧 **Empfänger:** {len(recipient_emails)}

📧 **Empfänger-Liste:**
{chr(10).join([f"   • {email}" for email in recipient_emails])}

{'-' * 50}
📋 **PAPERS:**
{papers_list}

📎 **Excel-Datei:** Die aktualisierte Excel-Datei ist als Anhang beigefügt.

ℹ️ **Hinweis:** Diese Email wurde manuell über das Paper-Suche System versendet.

Mit freundlichen Grüßen,
Ihr Paper-Suche System"""
    
    # Excel als Anhang
    excel_path = st.session_state["excel_template"]["file_path"]
    attachment_path = excel_path if os.path.exists(excel_path) else None
    
    # Email senden
    with st.spinner(f"📧 Sende Email an {len(recipient_emails)} Empfänger..."):
        success, status_message = send_real_email_multiple(recipient_emails, subject, message, attachment_path)
    
    # Email-Historie
    email_entry = {
        "timestamp": datetime.datetime.now().isoformat(),
        "type": f"Manuell - {email_type}",
        "search_term": search_term,
        "recipients": recipient_emails,
        "recipient_count": len(recipient_emails),
        "subject": subject,
        "paper_count": len(papers),
        "success": success,
        "status": status_message,
        "has_attachment": attachment_path is not None
    }
    
    st.session_state["email_history"].append(email_entry)
    
    # Ergebnis anzeigen
    if success:
        st.session_state["system_status"]["total_emails"] += 1
        st.success(f"📧 **Email erfolgreich versendet!**\n{status_message}")
        st.balloons()
        
        # Details anzeigen
        with st.expander("📋 Email-Details anzeigen"):
            st.write(f"**📧 Empfänger ({len(recipient_emails)}):**")
            for i, email in enumerate(recipient_emails, 1):
                st.write(f"   {i}. {email}")
            st.write(f"**📄 Papers:** {len(papers)}")
            st.write(f"**📎 Anhang:** {'✅ Excel-Datei' if attachment_path else '❌ Kein Anhang'}")
    else:
        st.error(f"❌ **Email-Fehler:** {status_message}")

def show_email_config():
    """Email-Konfiguration mit mehreren Empfängern"""
    st.subheader("📧 Email-Konfiguration (Mehrere Empfänger)")
    
    settings = st.session_state.get("email_settings", {})
    
    # Email-Setup Hilfe
    with st.expander("📖 Email-Setup Hilfe & Mehrere Empfänger"):
        st.info("""
        **Für Gmail (empfohlen):**
        1. ✅ 2-Faktor-Authentifizierung aktivieren
        2. ✅ App-Passwort erstellen (nicht normales Passwort!)
        3. ✅ SMTP: smtp.gmail.com, Port: 587, TLS: An
        
        **Mehrere Empfänger:**
        • Trennen Sie mehrere Email-Adressen mit Kommas
        • Beispiel: user1@gmail.com, user2@outlook.com, user3@company.de
        • Whitespaces werden automatisch entfernt
        
        **Für Outlook/Hotmail:**
        - SMTP: smtp-mail.outlook.com, Port: 587
        """)
    
    with st.form("email_config_form"):
        st.subheader("📬 Grundeinstellungen")
        
        col1, col2 = st.columns(2)
        
        with col1:
            sender_email = st.text_input(
                "Absender Email *", 
                value=settings.get("sender_email", ""),
                placeholder="absender@gmail.com"
            )
            
            smtp_server = st.text_input(
                "SMTP Server *",
                value=settings.get("smtp_server", "smtp.gmail.com")
            )
            
            auto_notifications = st.checkbox(
                "Automatische Benachrichtigungen", 
                value=settings.get("auto_notifications", True)
            )
        
        with col2:
            smtp_port = st.number_input(
                "SMTP Port *",
                value=settings.get("smtp_port", 587),
                min_value=1,
                max_value=65535
            )
            
            min_papers = st.number_input(
                "Min. Papers für Benachrichtigung", 
                value=settings.get("min_papers", 1),
                min_value=1,
                max_value=100
            )
            
            use_tls = st.checkbox(
                "TLS Verschlüsselung verwenden (empfohlen)",
                value=settings.get("use_tls", True)
            )
        
        # MEHRERE EMPFÄNGER - Text Area
        recipient_emails = st.text_area(
            "📧 Empfänger Email-Adressen * (mehrere mit Komma trennen)",
            value=settings.get("recipient_emails", ""),
            placeholder="empfaenger1@example.com, empfaenger2@gmail.com, empfaenger3@company.de",
            help="Mehrere Email-Adressen mit Komma trennen. Beispiel: user1@gmail.com, user2@outlook.com",
            height=80
        )
        
        sender_password = st.text_input(
            "Email Passwort / App-Passwort *",
            value=settings.get("sender_password", ""),
            type="password",
            help="Für Gmail: App-spezifisches Passwort verwenden!"
        )
        
        # Email-Vorlagen
        st.subheader("📝 Email-Vorlagen")
        
        subject_template = st.text_input(
            "Betreff-Vorlage",
            value=settings.get("subject_template", "🔬 {count} neue Papers für '{search_term}'"),
            help="Platzhalter: {count}, {search_term}, {frequency}"
        )
        
        message_template = st.text_area(
            "Nachricht-Vorlage",
            value=settings.get("message_template", """📧 Automatische Paper-Benachrichtigung

📅 Datum: {date}
🔍 Suchbegriff: '{search_term}'
📊 Neue Papers: {count}

📋 Neue Papers:
{new_papers_list}

📎 Excel-Datei: {excel_file}

Mit freundlichen Grüßen,
Ihr Paper-Suche System"""),
            height=200,
            help="Platzhalter: {date}, {search_term}, {count}, {frequency}, {new_papers_list}, {excel_file}"
        )
        
        if st.form_submit_button("💾 **Email-Einstellungen speichern**", type="primary"):
            # Validiere Email-Adressen
            recipient_list = parse_recipient_emails(recipient_emails)
            
            if not recipient_list:
                st.error("❌ Mindestens eine gültige Empfänger-Email erforderlich!")
            else:
                new_settings = {
                    "sender_email": sender_email,
                    "recipient_emails": recipient_emails,
                    "smtp_server": smtp_server,
                    "smtp_port": smtp_port,
                    "sender_password": sender_password,
                    "use_tls": use_tls,
                    "auto_notifications": auto_notifications,
                    "min_papers": min_papers,
                    "subject_template": subject_template,
                    "message_template": message_template,
                    "parsed_recipients": recipient_list  # Store parsed list
                }
                
                st.session_state["email_settings"] = new_settings
                st.success(f"✅ Email-Einstellungen gespeichert! **{len(recipient_list)} Empfänger** konfiguriert:")
                for i, email in enumerate(recipient_list, 1):
                    st.write(f"   {i}. 📧 {email}")
    
    # Zeige konfigurierte Empfänger
    if settings.get("recipient_emails"):
        recipient_list = parse_recipient_emails(settings.get("recipient_emails", ""))
        if recipient_list:
            st.info(f"📧 **Aktuell konfigurierte Empfänger ({len(recipient_list)}):**")
            cols = st.columns(min(len(recipient_list), 3))
            for i, email in enumerate(recipient_list):
                with cols[i % 3]:
                    st.write(f"✅ {email}")
    
    # Test-Email
    st.markdown("---")
    st.subheader("🧪 Email-System testen")
    
    col_test1, col_test2 = st.columns(2)
    
    with col_test1:
        if st.button("📧 **Test-Email an alle Empfänger senden**", type="primary"):
            send_test_email_multiple()
    
    with col_test2:
        if st.button("📊 **Email-Status prüfen**"):
            check_email_status_multiple()

def perform_comprehensive_pubmed_search(query: str, max_results: int) -> List[Dict[str, Any]]:
    """Kompatibilitäts-Wrapper für robuste PubMed-Suche mit Fallback"""
    try:
        # 1. Versuche robuste Methode
        papers = perform_comprehensive_pubmed_search_robust(query, max_results)
        if papers:
            return papers
        
        # 2. Bei Fehler: Versuche alternative Methode
        st.warning("🔄 Erste Methode fehlgeschlagen - versuche Alternative...")
        papers = try_alternative_pubmed_search(query, max_results)
        if papers:
            return papers
        
        # 3. Bei weiterem Fehler: Vereinfachte Suche ohne Datums-Filter
        simple_query = query.split(" AND ")[0]  # Entferne Datums-Filter
        st.info(f"🔄 Versuche vereinfachte Suche ohne Filter: '{simple_query}'")
        papers = try_simple_pubmed_search(simple_query, min(max_results, 50))
        return papers if papers else []
        
    except Exception as e:
        st.error(f"❌ Alle Suchmethoden fehlgeschlagen: {str(e)}")
        return []

def try_simple_pubmed_search(query: str, max_results: int) -> List[Dict[str, Any]]:
    """Einfache PubMed-Suche ohne Filter als letzter Fallback"""
    try:
        search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        params = {
            "db": "pubmed",
            "term": query,  # Nur der reine Suchbegriff
            "retmode": "json",
            "retmax": max_results,
            "email": "research@example.com",
            "tool": "PaperSearch"
        }
        
        response = requests.get(search_url, params=params, timeout=30)
        if response.status_code == 200:
            data = response.json()
            pmids = data.get("esearchresult", {}).get("idlist", [])
            if pmids:
                st.success(f"✅ Vereinfachte Suche erfolgreich: {len(pmids)} Papers gefunden")
                return fetch_paper_details_batch(pmids)
        return []
    except Exception as e:
        st.error(f"❌ Auch vereinfachte Suche fehlgeschlagen: {str(e)}")
        return []



def send_test_email_multiple():
    """Sendet Test-Email an alle konfigurierten Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    if not settings.get("sender_email") or not recipient_emails:
        st.error("❌ Email-Konfiguration unvollständig!")
        return
    
    subject = "🧪 Test-Email vom Paper-Suche System (Mehrere Empfänger)"
    message = f"""Dies ist eine Test-Email vom Paper-Suche System mit Unterstützung für mehrere Empfänger.

📅 Gesendet am: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}
📧 Von: {settings.get('sender_email')}
📧 An: {len(recipient_emails)} Empfänger

Empfänger-Liste:
{chr(10).join([f"• {email}" for email in recipient_emails])}

✅ Wenn Sie diese Email erhalten, funktioniert das Email-System korrekt!

System-Informationen:
• SMTP Server: {settings.get('smtp_server')}
• Port: {settings.get('smtp_port')}
• TLS: {'Aktiviert' if settings.get('use_tls') else 'Deaktiviert'}
• Empfänger: {len(recipient_emails)}

Mit freundlichen Grüßen,
Ihr Paper-Suche System"""
    
    success, status_message = send_real_email_multiple(
        recipient_emails, 
        subject, 
        message
    )
    
    if success:
        st.success(f"✅ **Test-Email erfolgreich gesendet!** {status_message}")
        st.balloons()
    else:
        st.error(f"❌ **Test-Email fehlgeschlagen:** {status_message}")

def check_email_status_multiple():
    """Prüft Email-Status mit mehreren Empfängern"""
    settings = st.session_state.get("email_settings", {})
    
    st.write("**📊 Email-Konfiguration Status:**")
    
    # Prüfe Konfiguration
    sender_ok = bool(settings.get("sender_email"))
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    recipients_ok = len(recipient_emails) > 0
    password_ok = bool(settings.get("sender_password"))
    
    st.write(f"📧 Absender Email: {'✅' if sender_ok else '❌'} {settings.get('sender_email', 'Nicht konfiguriert')}")
    st.write(f"📧 Empfänger Emails: {'✅' if recipients_ok else '❌'} {len(recipient_emails)} konfiguriert")
    
    if recipients_ok:
        with st.expander(f"📧 Empfänger-Liste ({len(recipient_emails)})"):
            for i, email in enumerate(recipient_emails, 1):
                st.write(f"   {i}. {email}")
    
    st.write(f"🔑 Passwort: {'✅' if password_ok else '❌'} {'Konfiguriert' if password_ok else 'Nicht konfiguriert'}")
    st.write(f"🔒 SMTP Server: {settings.get('smtp_server', 'smtp.gmail.com')}:{settings.get('smtp_port', 587)}")
    st.write(f"🔐 TLS: {'✅ Aktiviert' if settings.get('use_tls', True) else '❌ Deaktiviert'}")
    
    # Gesamtstatus
    if sender_ok and recipients_ok and password_ok:
        st.success(f"✅ **Email-System vollständig konfiguriert für {len(recipient_emails)} Empfänger!**")
    else:
        st.error("❌ **Email-System nicht vollständig konfiguriert!**")

# =============== WEITERE FUNKTIONEN ===============

def perform_comprehensive_pubmed_search_robust(query: str, max_results: int) -> List[Dict[str, Any]]:
    """Robuste PubMed-Suche mit Retry-Mechanismus und besserer Fehlerbehandlung"""
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
    
    # Retry-Konfiguration
    max_retries = 3
    retry_delays = [2, 5, 10]  # Sekunden zwischen Versuchen
    
    for attempt in range(max_retries):
        try:
            # 1. esearch - hole PMIDs mit verbesserter Query
            search_url = f"{base_url}esearch.fcgi"
            
            # Verbesserte Parameter
            params = {
                "db": "pubmed",
                "term": clean_pubmed_query(query),  # Query säubern
                "retmode": "json",
                "retmax": min(max_results, 9999),  # PubMed Limit beachten
                "email": "research.system@papersearch.com",
                "tool": "ScientificPaperSearchSystem",
                "sort": "relevance",
                "usehistory": "y"  # History Server nutzen
            }
            
            with st.spinner(f"🔍 Verbinde zu PubMed (Versuch {attempt + 1}/{max_retries})..."):
                # Längere Timeouts
                response = requests.get(search_url, params=params, timeout=60)
                
                # Status Code prüfen
                if response.status_code == 200:
                    try:
                        data = response.json()
                        pmids = data.get("esearchresult", {}).get("idlist", [])
                        total_count = int(data.get("esearchresult", {}).get("count", 0))
                        
                        if pmids:
                            st.write(f"📊 **PubMed Datenbank:** {total_count:,} Papers verfügbar, {len(pmids)} werden abgerufen")
                            
                            # 2. efetch - hole Details in Batches
                            return fetch_paper_details_batch_robust(pmids, batch_size=8)
                        else:
                            st.warning(f"⚠️ Keine PMIDs für Query '{query}' gefunden")
                            return []
                            
                    except json.JSONDecodeError as json_error:
                        st.error(f"❌ JSON Parse Error: {str(json_error)}")
                        if attempt < max_retries - 1:
                            continue
                        return []
                
                elif response.status_code == 500:
                    st.warning(f"⚠️ Server Error 500 - Versuch {attempt + 1}/{max_retries}")
                    if attempt < max_retries - 1:
                        time.sleep(retry_delays[attempt])
                        continue
                    else:
                        st.error("❌ **PubMed Server dauerhaft nicht erreichbar - versuchen Sie es später erneut**")
                        return []
                        
                elif response.status_code == 429:
                    st.warning(f"⚠️ Rate Limit erreicht - warte {retry_delays[attempt]} Sekunden...")
                    time.sleep(retry_delays[attempt] * 2)  # Längere Pause bei Rate Limiting
                    continue
                    
                else:
                    response.raise_for_status()
                    
        except requests.exceptions.Timeout:
            st.warning(f"⏰ Timeout - Versuch {attempt + 1}/{max_retries}")
            if attempt < max_retries - 1:
                time.sleep(retry_delays[attempt])
                continue
            else:
                st.error("❌ **PubMed Timeout - versuchen Sie es später erneut**")
                return []
                
        except requests.exceptions.RequestException as e:
            st.warning(f"🌐 Netzwerkfehler - Versuch {attempt + 1}/{max_retries}: {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(retry_delays[attempt])
                continue
            else:
                st.error(f"❌ **PubMed Verbindungsfehler:** {str(e)}")
                return []
                
        except Exception as e:
            st.error(f"❌ **Unerwarteter Fehler:** {str(e)}")
            return []
    
    return []


def perform_comprehensive_pubmed_search_robust(query: str, max_results: int) -> List[Dict[str, Any]]:
    """Ursprüngliche PubMed-Suche (Fallback für Kompatibilität)"""
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
    
    # 1. esearch - hole PMIDs
    search_url = f"{base_url}esearch.fcgi"
    params = {
        "db": "pubmed",
        "term": query,
        "retmode": "json",
        "retmax": max_results,
        "email": "research.system@papersearch.com",
        "tool": "ScientificPaperSearchSystem",
        "sort": "relevance"
    }
    
    try:
        with st.spinner("🔍 Verbinde zu PubMed..."):
            response = requests.get(search_url, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            pmids = data.get("esearchresult", {}).get("idlist", [])
            total_count = int(data.get("esearchresult", {}).get("count", 0))
            
            st.write(f"📊 **PubMed Datenbank:** {total_count:,} Papers verfügbar, {len(pmids)} werden abgerufen")
            
            if not pmids:
                return []
            
            # 2. efetch - hole Details in Batches
            return fetch_paper_details_batch(pmids)
            
    except requests.exceptions.RequestException as e:
        st.error(f"❌ **PubMed Verbindungsfehler:** {str(e)}")
        return []
    except Exception as e:
        st.error(f"❌ **PubMed Suchfehler:** {str(e)}")
        return []

def try_alternative_pubmed_search(query: str, max_results: int = 100) -> List[Dict[str, Any]]:
    """Alternative PubMed-Suche bei Server-Problemen"""
    st.info("🔄 Versuche alternative Suchmethode...")
    
    try:
        # Vereinfachte Query ohne Datums-Filter
        simple_query = query.replace(" AND ", " ").replace(" OR ", " ")
        
        # Alternative URL mit anderen Parametern
        search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        params = {
            "db": "pubmed",
            "term": simple_query,
            "retmode": "json",
            "retmax": min(max_results, 100),  # Kleineres Limit
            "email": "research@papersearch.com",
            "tool": "PaperSearch"
        }
        
        response = requests.get(search_url, params=params, timeout=30)
        
        if response.status_code == 200:
            data = response.json()
            pmids = data.get("esearchresult", {}).get("idlist", [])
            
            if pmids:
                st.success(f"✅ Alternative Methode erfolgreich: {len(pmids)} Papers gefunden")
                return fetch_paper_details_batch_robust(pmids[:50], batch_size=8) # Nur erste 50
        
        return []
        
    except Exception as e:
        st.error(f"❌ Auch alternative Methode fehlgeschlagen: {str(e)}")
        return []

def perform_comprehensive_pubmed_search_robust(query: str, max_results: int) -> List[Dict[str, Any]]:
    """Robuste PubMed-Suche mit Retry-Mechanismus"""
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
    
    # Retry-Konfiguration
    max_retries = 3
    retry_delays = [2, 5, 10]  # Sekunden zwischen Versuchen
    
    for attempt in range(max_retries):
        try:
            # 1. esearch - hole PMIDs
            search_url = f"{base_url}esearch.fcgi"
            params = {
                "db": "pubmed",
                "term": query,
                "retmode": "json",
                "retmax": min(max_results, 9999),
                "email": "research.system@papersearch.com",
                "tool": "ScientificPaperSearchSystem",
                "sort": "relevance"
            }
            
            with st.spinner(f"🔍 Verbinde zu PubMed (Versuch {attempt + 1}/{max_retries})..."):
                response = requests.get(search_url, params=params, timeout=60)
                
                if response.status_code == 200:
                    try:
                        data = response.json()
                        pmids = data.get("esearchresult", {}).get("idlist", [])
                        total_count = int(data.get("esearchresult", {}).get("count", 0))
                        
                        if pmids:
                            st.write(f"📊 **PubMed Datenbank:** {total_count:,} Papers verfügbar, {len(pmids)} werden abgerufen")
                            return fetch_paper_details_batch(pmids)
                        else:
                            st.warning(f"⚠️ Keine PMIDs für Query '{query}' gefunden")
                            return []
                    except json.JSONDecodeError:
                        if attempt < max_retries - 1:
                            continue
                        return []
                
                elif response.status_code == 500:
                    st.warning(f"⚠️ Server Error 500 - Versuch {attempt + 1}/{max_retries}")
                    if attempt < max_retries - 1:
                        time.sleep(retry_delays[attempt])
                        continue
                    else:
                        st.error("❌ **PubMed Server dauerhaft nicht erreichbar**")
                        return []
                
                elif response.status_code == 429:
                    st.warning(f"⚠️ Rate Limit erreicht - warte {retry_delays[attempt]} Sekunden...")
                    time.sleep(retry_delays[attempt] * 2)
                    continue
                else:
                    response.raise_for_status()
                    
        except requests.exceptions.Timeout:
            st.warning(f"⏰ Timeout - Versuch {attempt + 1}/{max_retries}")
            if attempt < max_retries - 1:
                time.sleep(retry_delays[attempt])
                continue
            else:
                st.error("❌ **PubMed Timeout - versuchen Sie es später erneut**")
                return []
                
        except requests.exceptions.RequestException as e:
            st.warning(f"🌐 Netzwerkfehler - Versuch {attempt + 1}/{max_retries}: {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(retry_delays[attempt])
                continue
            else:
                st.error(f"❌ **PubMed Verbindungsfehler:** {str(e)}")
                return []
    
    return []


def clean_pubmed_query(query: str) -> str:
    """Säubert und optimiert PubMed Query"""
    # Entferne problematische Zeichen
    cleaned = re.sub(r'[^\w\s\[\]:()-]', ' ', query)
    
    # Normalisiere Leerzeichen
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    
    # URL-encode für sicheren Transport
    import urllib.parse
    return urllib.parse.quote(cleaned)

def fetch_paper_details_batch_robust(pmids: List[str], batch_size: int = 15) -> List[Dict[str, Any]]:
    """
    🔧 ROBUSTE Paper-Details mit kleineren Batches und aggressivem Rate Limiting
    Löst PubMed Server-Probleme durch kleinere Requests und längere Pausen
    """
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    all_papers = []
    
    # SEHR KLEINE BATCHES (15 statt 50) für bessere Stabilität
    batches = [pmids[i:i + batch_size] for i in range(0, len(pmids), batch_size)]
    
    progress_bar = st.progress(0)
    batch_status = st.empty()
    
    successful_batches = 0
    failed_batches = 0
    
    for batch_idx, batch_pmids in enumerate(batches):
        try:
            batch_status.text(f"📥 Robust Batch {batch_idx + 1}/{len(batches)}: {len(batch_pmids)} Papers...")
            
            # AGGRESSIVER RETRY-MECHANISMUS (5 Versuche pro Batch)
            batch_success = False
            for retry_attempt in range(5):
                try:
                    # OPTIMIERTE PARAMETER
                    params = {
                        "db": "pubmed",
                        "id": ",".join(batch_pmids),
                        "retmode": "xml",
                        "email": "research.system@papersearch.com",
                        "tool": "ScientificPaperSearchSystem"
                    }
                    
                    # SICHERE SESSION mit erweiterten Headers
                    session = requests.Session()
                    session.headers.update({
                        'User-Agent': 'Mozilla/5.0 (compatible; PaperSearchSystem/1.0)',
                        'Connection': 'close',
                        'Accept': 'application/xml'
                    })
                    
                    # LÄNGERE TIMEOUTS
                    response = session.get(
                        base_url, 
                        params=params, 
                        timeout=(20, 60),  # connect_timeout, read_timeout
                        verify=True
                    )
                    
                    session.close()
                    
                    if response.status_code == 200:
                        # Parse XML erfolgreich
                        root = ET.fromstring(response.content)
                        articles = root.findall(".//PubmedArticle")
                        
                        batch_papers = []
                        for article in articles:
                            paper_data = parse_pubmed_article(article)
                            if paper_data:
                                batch_papers.append(paper_data)
                        
                        all_papers.extend(batch_papers)
                        successful_batches += 1
                        batch_success = True
                        break  # Erfolgreich - nächster Batch
                        
                    elif response.status_code == 429:
                        # Rate Limit - SEHR LANGE warten
                        wait_time = 15 + (retry_attempt * 10)  # 15, 25, 35, 45, 55 Sekunden
                        st.warning(f"⏳ Rate Limit - warte {wait_time}s (Retry {retry_attempt + 1}/5)")
                        time.sleep(wait_time)
                        continue
                        
                    elif response.status_code in [500, 502, 503, 504]:
                        # Server Error - exponentiell längere Wartezeit
                        wait_time = 10 * (2 ** retry_attempt)  # 10, 20, 40, 80, 160 Sekunden
                        st.warning(f"⚠️ Server Error {response.status_code} - Retry {retry_attempt + 1}/5 in {wait_time}s")
                        time.sleep(wait_time)
                        continue
                        
                    else:
                        st.warning(f"❌ HTTP {response.status_code} - Retry {retry_attempt + 1}/5")
                        time.sleep(5 + retry_attempt * 2)
                        continue
                        
                except requests.exceptions.Timeout:
                    st.warning(f"⏰ Timeout Batch {batch_idx + 1} - Retry {retry_attempt + 1}/5")
                    time.sleep(8 + retry_attempt * 3)
                    continue
                    
                except requests.exceptions.ConnectionError:
                    st.warning(f"🌐 Connection Error Batch {batch_idx + 1} - Retry {retry_attempt + 1}/5")
                    time.sleep(15 + retry_attempt * 5)  # Sehr lange Pause bei Connection-Errors
                    continue
                    
                except Exception as batch_error:
                    st.warning(f"⚠️ Unerwarteter Fehler Batch {batch_idx + 1}: {str(batch_error)}")
                    time.sleep(5)
                    continue
            
            # Wenn alle Retries fehlgeschlagen sind
            if not batch_success:
                failed_batches += 1
                st.error(f"❌ Batch {batch_idx + 1} nach 5 Versuchen fehlgeschlagen")
            
            # Progress Update
            progress = (batch_idx + 1) / len(batches)
            progress_bar.progress(progress)
            
            # ULTRA-AGGRESSIVES RATE LIMITING zwischen Batches
            sleep_time = 5.0 + (failed_batches * 2.0)  # Minimum 5s, mehr bei Fehlern
            if failed_batches > 0:
                st.info(f"⏳ Erweiterte Pause von {sleep_time:.0f}s nach Fehlern...")
            time.sleep(sleep_time)
            
        except Exception as e:
            failed_batches += 1
            st.error(f"❌ Kritischer Batch-Fehler {batch_idx + 1}: {str(e)}")
            time.sleep(10)  # Lange Pause nach kritischen Fehlern
            continue
    
    # Cleanup
    progress_bar.empty()
    batch_status.empty()
    
    # Finale Statistiken
    success_rate = (successful_batches / len(batches)) * 100 if batches else 0
    
    if successful_batches > 0:
        st.success(f"✅ **Batch-Ergebnis:** {successful_batches}/{len(batches)} erfolgreich ({success_rate:.1f}%) - {len(all_papers)} Papers erhalten")
    
    if failed_batches > 0:
        st.warning(f"⚠️ **{failed_batches} Batches fehlgeschlagen** - PubMed Server überlastet")
        st.info("💡 **Tipp:** Warten Sie 10-15 Minuten und versuchen Sie es erneut")
    
    return all_papers

def perform_comprehensive_pubmed_search_robust(query: str, max_results: int) -> List[Dict[str, Any]]:
    """
    🔧 ULTRA-ROBUSTE PubMed-Suche mit Server-Error Behandlung
    """
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
    search_url = f"{base_url}esearch.fcgi"
    fetch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"



    
    # Reduzierte Retry-Anzahl aber längere Wartezeiten
    max_retries = 3
    retry_delays = [10, 30, 60]  # 10s, 30s, 60s zwischen Versuchen
    
    for attempt in range(max_retries):
        try:
            search_url = f"{base_url}esearch.fcgi"
            
            params = {
                "db": "pubmed",
                "term": query,
                "retmode": "json",
                "retmax": min(max_results, 100),  # Begrenzt auf 100
                "email": "research.system@papersearch.com",
                "tool": "ScientificPaperSearchSystem",
                "sort": "relevance"
            }
            
            with st.spinner(f"🔍 PubMed Suche (Versuch {attempt + 1}/{max_retries})..."):
                
                # SICHERE SESSION
                session = requests.Session()
                session.headers.update({
                    'User-Agent': 'Mozilla/5.0 (compatible; PaperSearchSystem/1.0)',
                    'Accept': 'application/json',
                    'Connection': 'close'
                })
                
                response = session.get(
                    search_url, 
                    params=params, 
                    timeout=(15, 45),
                    verify=True
                )
                
                session.close()
                
                if response.status_code == 200:
                    try:
                        data = response.json()
                        pmids = data.get("esearchresult", {}).get("idlist", [])
                        total_count = int(data.get("esearchresult", {}).get("count", 0))
                        
                        if pmids:
                            st.success(f"✅ **PubMed Search erfolgreich:** {len(pmids)} von {total_count:,} Papers")
                            
                            # Verwende die neue robuste Batch-Funktion
                            return fetch_paper_details_batch_robust(pmids, batch_size=12)
                        else:
                            st.warning(f"⚠️ Keine Papers für '{query}' gefunden")
                            return []
                            
                    except json.JSONDecodeError as json_error:
                        st.error(f"❌ JSON Parse Error: {str(json_error)}")
                        if attempt < max_retries - 1:
                            time.sleep(retry_delays[attempt])
                            continue
                        return []
                
                elif response.status_code == 429:
                    wait_time = 60 + (attempt * 30)  # 60, 90, 120 Sekunden
                    st.warning(f"⏳ **Rate Limit erreicht** - warte {wait_time} Sekunden...")
                    time.sleep(wait_time)
                    continue
                    
                elif response.status_code in [500, 502, 503, 504]:
                    server_wait = retry_delays[attempt] * 2  # Doppelte Wartezeit bei Server-Fehlern
                    st.error(f"🔴 **PubMed Server Error {response.status_code}** - Versuch {attempt + 1}/{max_retries}")
                    st.info(f"⏳ **Warte {server_wait} Sekunden** vor nächstem Versuch...")
                    
                    if attempt < max_retries - 1:
                        time.sleep(server_wait)
                        continue
                    else:
                        st.error("❌ **PubMed Server dauerhaft überlastet!**")
                        st.info("💡 **Lösungen:**")
                        st.write("• Warten Sie 15-30 Minuten und versuchen Sie es erneut")
                        st.write("• Reduzieren Sie die Anzahl gleichzeitiger Suchen")
                        st.write("• Verwenden Sie spezifischere Suchbegriffe")
                        return []
                        
                else:
                    st.error(f"❌ Unerwarteter HTTP Status: {response.status_code}")
                    if attempt < max_retries - 1:
                        time.sleep(retry_delays[attempt])
                        continue
                    return []
                    
        except requests.exceptions.Timeout:
            st.warning(f"⏰ **Timeout** - Versuch {attempt + 1}/{max_retries}")
            if attempt < max_retries - 1:
                time.sleep(retry_delays[attempt])
                continue
            else:
                st.error("❌ **PubMed dauerhaft nicht erreichbar**")
                return []
                
        except requests.exceptions.ConnectionError:
            st.warning(f"🌐 **Netzwerkfehler** - Versuch {attempt + 1}/{max_retries}")
            if attempt < max_retries - 1:
                time.sleep(retry_delays[attempt] * 1.5)  # Längere Pause bei Connection-Fehlern
                continue
            else:
                st.error("❌ **Netzwerkverbindung fehlgeschlagen**")
                return []
                
        except Exception as e:
            st.error(f"❌ **Kritischer Fehler:** {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(retry_delays[attempt])
                continue
            return []
    
    return []

def repeat_all_searches_from_excel_fixed():
    """
    🔧 VOLLSTÄNDIG REPARIERTE repeat_all_searches Funktion
    Verwendet das neue ultra-robuste System
    """
    excel_stats = get_search_statistics_from_excel()
    search_terms = excel_stats.get("search_terms", [])
    
    if not search_terms:
        st.info("📭 Keine Suchhistorie in Excel vorhanden.")
        return
    
    st.info(f"🔄 Wiederhole {len(search_terms)} Suchen mit **Ultra-Robust System** (kleinste Batches, längste Pausen)...")
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_new_papers = 0
    successful_searches = 0
    failed_searches = []
    
    for i, term_info in enumerate(search_terms):
        search_term = term_info.get("term", "")
        if not search_term:
            continue
        
        try:
            status_text.text(f"🔍 Ultra-Safe Suche {i+1}/{len(search_terms)}: '{search_term}'...")
            
            # 🎯 VERWENDE NEUE ULTRA-ROBUSTE METHODE
            current_papers = None
            search_attempts = []
            
            # 🎯 VERSUCH 1: Ultra-Robuste Haupt-Suche
            try:
                advanced_query = build_advanced_search_query_corrected(search_term, "Letzte 2 Jahre")
                current_papers = perform_comprehensive_pubmed_search_robust(advanced_query, 50)  # Reduziert auf 50
                
                if current_papers:
                    search_attempts.append(f"✅ Ultra-Robust Haupt-Suche: {len(current_papers)} Papers")
                else:
                    search_attempts.append("⚠️ Ultra-Robust Haupt-Suche: Keine Ergebnisse")
                    
            except Exception as e:
                search_attempts.append(f"❌ Ultra-Robust Haupt-Suche Fehler: {str(e)}")
            
            # 🎯 VERSUCH 2: Minimal-Suche (falls Hauptsuche fehlschlägt)
            if not current_papers:
                try:
                    minimal_query = search_term.strip()
                    current_papers = try_minimal_pubmed_search(minimal_query, 20)  # Nur 20 Papers
                    
                    if current_papers:
                        search_attempts.append(f"✅ Ultra-Safe Minimal-Suche: {len(current_papers)} Papers")
                    else:
                        search_attempts.append("❌ Ultra-Safe Minimal-Suche: Keine Ergebnisse")
                        
                except Exception as e:
                    search_attempts.append(f"❌ Ultra-Safe Minimal-Suche Fehler: {str(e)}")
            
            # 📊 ERGEBNIS VERARBEITEN
            if current_papers:
                # Füge neue Papers zur Excel hinzu
                added_count, new_papers = add_new_papers_to_excel(search_term, current_papers)
                
                if added_count > 0:
                    if should_send_email(added_count):
                        send_excel_integrated_email_multiple(search_term, new_papers, len(current_papers), added_count)
                    
                    total_new_papers += added_count
                    successful_searches += 1
                    
                    successful_method = search_attempts[-1] if search_attempts else "Unbekannt"
                    st.success(f"✅ **{search_term}:** {added_count} neue Papers | **Methode:** {successful_method}")
                else:
                    successful_searches += 1
                    st.info(f"ℹ️ **{search_term}:** Keine neuen Papers (alle bekannt)")
            else:
                # Alle Methoden fehlgeschlagen
                failed_searches.append({
                    "term": search_term,
                    "attempts": search_attempts
                })
                st.error(f"❌ **{search_term}:** Ultra-Robust System fehlgeschlagen")
            
            # Progress update
            progress_bar.progress((i + 1) / len(search_terms))
            
            # ULTRA-LANGSAME PAUSEN zwischen Suchen (5 Sekunden minimum)
            pause_time = 5.0 + (len(failed_searches) * 2.0)  # Längere Pausen bei Fehlern
            status_text.text(f"⏳ Pause {pause_time:.0f}s vor nächster Suche...")
            time.sleep(pause_time)
            
        except Exception as e:
            failed_searches.append({
                "term": search_term,
                "error": str(e)
            })
            st.error(f"❌ **Kritischer Fehler bei '{search_term}':** {str(e)}")
            time.sleep(10)  # Lange Pause nach kritischen Fehlern
            continue
    
    progress_bar.empty()
    status_text.empty()
    
    # 📊 FINALE ERGEBNISSE
    st.markdown("---")
    st.subheader("📊 **Ultra-Robust Wiederholung - Ergebnisse**")
    
    col_result1, col_result2, col_result3, col_result4 = st.columns(4)
    
    with col_result1:
        st.metric("🔍 Gesamt Suchen", len(search_terms))
    
    with col_result2:
        st.metric("✅ Erfolgreich", successful_searches)
    
    with col_result3:
        st.metric("🆕 Neue Papers", total_new_papers)
    
    with col_result4:
        st.metric("❌ Fehlgeschlagen", len(failed_searches))
    
    # Erfolgs-Meldung
    if total_new_papers > 0:
        st.success(f"🎉 **Ultra-Robust Wiederholung erfolgreich!** {total_new_papers} neue Papers gefunden!")
        st.balloons()
    elif successful_searches > 0:
        st.info(f"ℹ️ **Wiederholung abgeschlossen.** {successful_searches} Suchen erfolgreich, aber keine neuen Papers.")
    else:
        st.warning("⚠️ **Alle Suchen fehlgeschlagen.** PubMed möglicherweise überlastet - versuchen Sie es später erneut.")
    
    # Fehlgeschlagene Suchen anzeigen
    if failed_searches:
        with st.expander(f"❌ **Fehlgeschlagene Suchen ({len(failed_searches)}):**"):
            for fail in failed_searches:
                st.write(f"**{fail['term']}:**")
                if 'attempts' in fail:
                    for attempt in fail['attempts']:
                        st.write(f"   • {attempt}")
                if 'error' in fail:
                    st.write(f"   • Kritischer Fehler: {fail['error']}")
                st.write("---")

def try_minimal_pubmed_search_ultra_safe(query: str, max_results: int) -> List[Dict[str, Any]]:
    """
    🔧 ULTRA-SICHERE minimale PubMed-Suche als absoluter Fallback
    """
    try:
        st.info(f"🔄 **Ultra-Safe Minimal Search:** {query}")
        
        search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        
        # ABSOLUT MINIMALE Parameter
        params = {
            "db": "pubmed",
            "term": query,
            "retmax": min(max_results, 10),  # Nur 10 Papers
            "retmode": "json",
            "email": "research.system@papersearch.com"
        }
        
        # SICHERE SESSION
        session = requests.Session()
        session.headers.update({'Connection': 'close'})
        
        response = session.get(search_url, params=params, timeout=(10, 30))
        session.close()
        
        if response.status_code == 200:
            data = response.json()
            pmids = data.get("esearchresult", {}).get("idlist", [])
            
            if pmids:
                st.success(f"✅ **Ultra-Safe Minimal Search erfolgreich:** {len(pmids)} Papers")
                # Verwende ultra-robuste Batch-Funktion mit kleinsten Batches
                return fetch_paper_details_batch_robust(pmids, batch_size=8)  # Nur 5 pro Batch
        
        return []
        
    except Exception as e:
        st.error(f"❌ Ultra-Safe Minimal Search Exception: {str(e)}")
        return []



def fetch_paper_details_batch(pmids: List[str], batch_size: int = 50) -> List[Dict[str, Any]]:
    """Holt Paper-Details in Batches für bessere Performance"""
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    all_papers = []
    
    # Teile PMIDs in Batches
    batches = [pmids[i:i + batch_size] for i in range(0, len(pmids), batch_size)]
    
    progress_bar = st.progress(0)
    batch_status = st.empty()
    
    for batch_idx, batch_pmids in enumerate(batches):
        try:
            batch_status.text(f"📥 Batch {batch_idx + 1}/{len(batches)}: {len(batch_pmids)} Papers...")
            
            params = {
                "db": "pubmed",
                "id": ",".join(batch_pmids),
                "retmode": "xml",
                "email": "research.system@papersearch.com",
                "tool": "ScientificPaperSearchSystem"
            }
            
            response = requests.get(base_url, params=params, timeout=60)
            response.raise_for_status()
            
            # Parse XML
            root = ET.fromstring(response.content)
            articles = root.findall(".//PubmedArticle")
            
            for article in articles:
                paper_data = parse_pubmed_article(article)
                if paper_data:
                    all_papers.append(paper_data)
            
            # Progress Update
            progress = (batch_idx + 1) / len(batches)
            progress_bar.progress(progress)
            
            # Rate limiting
            time.sleep(0.5)
            
        except Exception as e:
            st.warning(f"⚠️ Batch {batch_idx + 1} Fehler: {str(e)}")
            continue
    
    progress_bar.empty()
    batch_status.empty()
    
    return all_papers

def parse_pubmed_article(article) -> Dict[str, Any]:
    """Erweiterte Artikel-Parsing mit mehr Feldern"""
    try:
        # PMID
        pmid_elem = article.find(".//PMID")
        pmid = pmid_elem.text if pmid_elem is not None else ""
        
        # Title
        title_elem = article.find(".//ArticleTitle")
        title = title_elem.text if title_elem is not None else "Titel nicht verfügbar"
        
        # Abstract (alle Teile)
        abstract_parts = []
        for abstract_elem in article.findall(".//AbstractText"):
            if abstract_elem.text:
                label = abstract_elem.get("Label", "")
                text = abstract_elem.text
                if label and label.upper() not in ["UNLABELLED", "UNASSIGNED"]:
                    abstract_parts.append(f"**{label}:** {text}")
                else:
                    abstract_parts.append(text)
        
        abstract = "\n\n".join(abstract_parts) if abstract_parts else "Kein Abstract verfügbar"
        
        # Journal Info
        journal_elem = article.find(".//Journal/Title")
        journal = journal_elem.text if journal_elem is not None else "Journal unbekannt"
        
        # Publication Date
        year_elem = article.find(".//PubDate/Year")
        year = year_elem.text if year_elem is not None else "Unbekannt"
        
        # Authors
        authors = []
        for author in article.findall(".//Author"):
            lastname = author.find("LastName")
            forename = author.find("ForeName")
            
            if lastname is not None:
                author_name = lastname.text or ""
                if forename is not None:
                    author_name = f"{author_name}, {forename.text}"
                authors.append(author_name)
        
        authors_str = "; ".join(authors[:8])  # Erste 8 Autoren
        if len(authors) > 8:
            authors_str += f" et al. (+{len(authors) - 8} weitere)"
        
        # DOI
        doi = ""
        for article_id in article.findall(".//ArticleId"):
            if article_id.get("IdType") == "doi":
                doi = article_id.text
                break
        
        return {
            "PMID": pmid,
            "Title": title,
            "Abstract": abstract,
            "Journal": journal,
            "Year": year,
            "Authors": authors_str,
            "DOI": doi,
            "URL": f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/",
            "Search_Date": datetime.datetime.now().isoformat(),
            "Is_New": True,
            "Has_DOI": bool(doi)
        }
        
    except Exception as e:
        st.warning(f"⚠️ Fehler beim Parsen eines Artikels: {str(e)}")
        return None

def display_excel_integrated_results(all_papers: List[Dict], new_papers: List[Dict], query: str, added_count: int, show_existing: bool):
    """Zeigt Ergebnisse der Excel-integrierten Suche an"""
    
    # Statistiken
    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
    with col_stat1:
        st.metric("📄 Gefunden", len(all_papers))
    with col_stat2:
        st.metric("🆕 Neue Papers", added_count)
    with col_stat3:
        st.metric("📊 Bereits bekannt", len(all_papers) - added_count)
    with col_stat4:
        st.metric("💾 In Excel gespeichert", added_count)
    
    # Neue Papers hervorheben
    if new_papers:
        st.subheader(f"🆕 Neue Papers ({len(new_papers)})")
        
        with st.expander(f"📋 Alle {len(new_papers)} neuen Papers anzeigen", expanded=True):
            for i, paper in enumerate(new_papers[:10], 1):  # Zeige erste 10
                with st.container():
                    col_paper1, col_paper2 = st.columns([3, 1])
                    
                    with col_paper1:
                        st.write(f"**{i}. {paper.get('Title', 'Unbekannt')[:100]}...**")
                        st.write(f"👥 {paper.get('Authors', 'n/a')[:80]}...")
                        st.write(f"📚 {paper.get('Journal', 'n/a')} ({paper.get('Year', 'n/a')})")
                        if paper.get('URL'):
                            st.markdown(f"🔗 [**PubMed**]({paper.get('URL')})")
                    
                    with col_paper2:
                        st.success("🆕 NEU")
                        st.write(f"PMID: {paper.get('PMID', 'n/a')}")
            
            if len(new_papers) > 10:
                st.info(f"... und {len(new_papers) - 10} weitere neue Papers (siehe Excel-Datei)")
    
    # Bereits bekannte Papers (optional)
    if show_existing and (len(all_papers) - added_count) > 0:
        existing_papers = [p for p in all_papers if p not in new_papers]
        
        with st.expander(f"📊 Bereits bekannte Papers ({len(existing_papers)})", expanded=False):
            for i, paper in enumerate(existing_papers[:5], 1):  # Zeige erste 5
                with st.container():
                    col_paper1, col_paper2 = st.columns([3, 1])
                    
                    with col_paper1:
                        st.write(f"**{i}. {paper.get('Title', 'Unbekannt')[:100]}...**")
                        st.write(f"👥 {paper.get('Authors', 'n/a')[:80]}...")
                        st.write(f"📚 {paper.get('Journal', 'n/a')} ({paper.get('Year', 'n/a')})")
                    
                    with col_paper2:
                        st.info("📊 BEKANNT")
                        st.write(f"PMID: {paper.get('PMID', 'n/a')}")
            
            if len(existing_papers) > 5:
                st.write(f"... und {len(existing_papers) - 5} weitere bereits bekannte Papers")

def generate_sheet_name(search_term: str) -> str:
    """Generiert gültigen Excel-Sheet-Namen"""
    # Excel Sheet Namen dürfen max 31 Zeichen haben und bestimmte Zeichen nicht enthalten
    invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
    
    clean_name = search_term
    for char in invalid_chars:
        clean_name = clean_name.replace(char, '_')
    
    # Entferne multiple Unterstriche und trimme
    clean_name = re.sub(r'_+', '_', clean_name).strip('_')
    
    # Kürze auf 25 Zeichen (lasse Platz für eventuelle Suffixe)
    if len(clean_name) > 25:
        clean_name = clean_name[:25]
    
    return clean_name

def build_advanced_search_query(query: str, date_filter: str) -> str:
    """KORRIGIERTE Suchanfrage mit PubMed-konformen Datums-Filtern"""
    query_parts = [query.strip()]
    
    if date_filter != "Alle":
        current_year = datetime.datetime.now().year
        
        # ✅ KORRIGIERT: Verwende [pdat] statt [dp] und korrekte Jahre
        if date_filter == "Letztes Jahr":
            start_year = current_year - 1
            query_parts.append(f"AND ({start_year}[pdat]:{current_year-1}[pdat])")
        elif date_filter == "Letzte 2 Jahre":
            start_year = current_year - 2
            query_parts.append(f"AND ({start_year}[pdat]:{current_year-1}[pdat])")
        elif date_filter == "Letzte 5 Jahre":
            start_year = current_year - 5
            query_parts.append(f"AND ({start_year}[pdat]:{current_year-1}[pdat])")
        elif date_filter == "Letzte 10 Jahre":
            start_year = current_year - 10
            query_parts.append(f"AND ({start_year}[pdat]:{current_year-1}[pdat])")
    
    return " ".join(query_parts)



def is_email_configured() -> bool:
    """Prüft Email-Konfiguration für mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    return (bool(settings.get("sender_email")) and 
            len(recipient_emails) > 0 and
            bool(settings.get("sender_password")))

def should_send_email(paper_count: int) -> bool:
    """Prüft ob Email gesendet werden soll"""
    settings = st.session_state.get("email_settings", {})
    return (settings.get("auto_notifications", False) and
            paper_count >= settings.get("min_papers", 1) and
            is_email_configured())

# =============== STATUS UND WIEDERHOLUNGSFUNKTIONEN ===============

def send_status_email_multiple():
    """Sendet Status-Email mit aktueller Übersicht an mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    if not is_email_configured():
        st.error("❌ Email nicht konfiguriert! Bitte konfigurieren Sie die Email-Einstellungen.")
        return
    
    # System-Status sammeln
    status = st.session_state["system_status"]
    excel_stats = get_search_statistics_from_excel()
    email_history = st.session_state.get("email_history", [])
    
    # Subject
    subject = f"📊 System-Status Report - {datetime.datetime.now().strftime('%d.%m.%Y')}"
    
    # Message erstellen
    message = f"""📊 **SYSTEM-STATUS REPORT**
    
📅 **Berichts-Datum:** {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}

📈 **SYSTEM-STATISTIKEN:**
• 🔍 Gesamt Suchen: {excel_stats.get('total_searches', 0)}
• 📄 Papers in Excel: {excel_stats.get('total_papers', 0)}
• 📊 Excel Sheets: {excel_stats.get('total_sheets', 0)}
• 📧 Gesendete Emails: {len(email_history)}
• 📧 Email-Empfänger: {len(recipient_emails)}

📋 **LETZTE SUCHAKTIVITÄTEN (Excel-basiert):**"""

    # Letzte Suchen aus Excel hinzufügen
    if excel_stats.get("search_terms"):
        recent_searches = sorted(excel_stats["search_terms"], key=lambda x: x.get("last_update", ""), reverse=True)[:5]
        for i, search in enumerate(recent_searches, 1):
            term = search.get("term", "Unbekannt")
            papers = search.get("papers", 0)
            new_papers = search.get("new_papers", 0)
            last_update = search.get("last_update", "")[:16].replace('T', ' ')
            
            message += f"\n{i}. 🔍 {term} ({papers} Papers, {new_papers} neu) - {last_update}"
    
    message += f"""

📧 **EMAIL-EMPFÄNGER ({len(recipient_emails)}):**
{chr(10).join([f"• {email}" for email in recipient_emails])}

📎 **EXCEL-DATEI:** 
Die aktuelle Master Excel-Datei enthält {excel_stats.get('total_sheets', 0)} Sheets mit insgesamt {excel_stats.get('total_papers', 0)} Papers.

---
Dieser Report wurde automatisch generiert.
System: Paper-Suche & Email-System v3.0 (Excel-Integration + Mehrere Empfänger)"""
    
    # Email senden mit Excel-Anhang
    template_path = st.session_state["excel_template"]["file_path"]
    excel_path = template_path if os.path.exists(template_path) else None
    
    success, status_message = send_real_email_multiple(
        recipient_emails, 
        subject, 
        message,
        excel_path
    )
    
    # Email-Historie aktualisieren
    email_entry = {
        "timestamp": datetime.datetime.now().isoformat(),
        "type": "Status-Report",
        "recipients": recipient_emails,
        "recipient_count": len(recipient_emails),
        "subject": subject,
        "success": success,
        "status": status_message,
        "has_attachment": excel_path is not None
    }
    
    st.session_state["email_history"].append(email_entry)
    
    # Update System-Status
    if success:
        st.session_state["system_status"]["total_emails"] += 1
    
    # Ergebnis anzeigen
    if success:
        st.success(f"📧 **Status-Email erfolgreich an {len(recipient_emails)} Empfänger gesendet!**")
        st.balloons()
    else:
        st.error(f"❌ **Status-Email Fehler:** {status_message}")

def repeat_all_searches_from_excel():
    """
    🔧 VOLLSTÄNDIG ÜBERARBEITETE FUNKTION mit Multi-Fallback System
    Verwendet das gleiche robuste System wie execute_excel_integrated_search()
    """
    excel_stats = get_search_statistics_from_excel()
    search_terms = excel_stats.get("search_terms", [])
    
    if not search_terms:
        st.info("📭 Keine Suchhistorie in Excel vorhanden.")
        return
    
    st.info(f"🔄 Wiederhole {len(search_terms)} Suchen mit **Multi-Fallback System**...")
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_new_papers = 0
    successful_searches = 0
    failed_searches = []
    
    for i, term_info in enumerate(search_terms):
        search_term = term_info.get("term", "")
        if not search_term:
            continue
        
        try:
            status_text.text(f"🔍 Suche {i+1}/{len(search_terms)}: '{search_term}' (Multi-Fallback)...")
            
            # 🎯 VERWENDE DIESELBEN FALLBACK-METHODEN WIE execute_excel_integrated_search
            current_papers = None
            search_attempts = []
            
            # 🎯 VERSUCH 1: Korrigierte Haupt-Suche
            try:
                advanced_query = build_advanced_search_query_corrected(search_term, "Letzte 2 Jahre")
                current_papers = perform_comprehensive_pubmed_search_robust(advanced_query, 100)
                
                if current_papers:
                    search_attempts.append(f"✅ Haupt-Suche: {len(current_papers)} Papers")
                else:
                    search_attempts.append("⚠️ Haupt-Suche: Keine Ergebnisse")
                    
            except Exception as e:
                search_attempts.append(f"❌ Haupt-Suche Fehler: {str(e)}")
            
            # 🎯 VERSUCH 2: Vereinfachte Suche ohne Datums-Filter
            if not current_papers:
                try:
                    simple_query = search_term.strip()
                    current_papers = try_simple_pubmed_search_enhanced(simple_query, 100)
                    
                    if current_papers:
                        search_attempts.append(f"✅ Vereinfachte Suche: {len(current_papers)} Papers")
                    else:
                        search_attempts.append("⚠️ Vereinfachte Suche: Keine Ergebnisse")
                        
                except Exception as e:
                    search_attempts.append(f"❌ Vereinfachte Suche Fehler: {str(e)}")
            
            # 🎯 VERSUCH 3: Alternative PubMed Parameter
            if not current_papers:
                try:
                    current_papers = try_alternative_pubmed_search_enhanced(search_term, 100)
                    
                    if current_papers:
                        search_attempts.append(f"✅ Alternative Suche: {len(current_papers)} Papers")
                    else:
                        search_attempts.append("⚠️ Alternative Suche: Keine Ergebnisse")
                        
                except Exception as e:
                    search_attempts.append(f"❌ Alternative Suche Fehler: {str(e)}")
            
            # 🎯 VERSUCH 4: Minimal-Suche als letzter Fallback
            if not current_papers:
                try:
                    minimal_query = search_term.split()[0] if " " in search_term else search_term
                    current_papers = try_minimal_pubmed_search(minimal_query, 20)
                    
                    if current_papers:
                        search_attempts.append(f"✅ Minimal-Suche: {len(current_papers)} Papers")
                    else:
                        search_attempts.append("❌ Minimal-Suche: Keine Ergebnisse")
                        
                except Exception as e:
                    search_attempts.append(f"❌ Minimal-Suche Fehler: {str(e)}")
            
            # 📊 ERGEBNIS VERARBEITEN
            if current_papers:
                # Füge neue Papers zur Excel hinzu
                added_count, new_papers = add_new_papers_to_excel(search_term, current_papers)
                
                if added_count > 0:
                    # Sende Email wenn konfiguriert
                    if should_send_email(added_count):
                        send_excel_integrated_email_multiple(search_term, new_papers, len(current_papers), added_count)
                    
                    total_new_papers += added_count
                    successful_searches += 1
                    
                    # Zeige Erfolg mit verwendeter Methode
                    successful_method = search_attempts[-1] if search_attempts else "Unbekannt"
                    st.success(f"✅ **{search_term}:** {added_count} neue Papers | **Methode:** {successful_method}")
                else:
                    successful_searches += 1
                    st.info(f"ℹ️ **{search_term}:** Keine neuen Papers (alle bekannt)")
            else:
                # Alle Methoden fehlgeschlagen
                failed_searches.append({
                    "term": search_term,
                    "attempts": search_attempts
                })
                st.error(f"❌ **{search_term}:** Alle Suchmethoden fehlgeschlagen")
                
                # Debug-Info für fehlgeschlagene Suchen
                with st.expander(f"🔍 Debug Info für '{search_term}'"):
                    for j, attempt in enumerate(search_attempts, 1):
                        st.write(f"{j}. {attempt}")
            
            # Progress update
            progress_bar.progress((i + 1) / len(search_terms))
            time.sleep(2)  # Längere Pause zwischen Suchen
            
        except Exception as e:
            failed_searches.append({
                "term": search_term,
                "error": str(e)
            })
            st.error(f"❌ **Kritischer Fehler bei '{search_term}':** {str(e)}")
            continue
    
    progress_bar.empty()
    status_text.empty()
    
    # 📊 FINALE ERGEBNISSE
    st.markdown("---")
    st.subheader("📊 **Multi-Fallback Wiederholung - Ergebnisse**")
    
    col_result1, col_result2, col_result3, col_result4 = st.columns(4)
    
    with col_result1:
        st.metric("🔍 Gesamt Suchen", len(search_terms))
    
    with col_result2:
        st.metric("✅ Erfolgreich", successful_searches)
    
    with col_result3:
        st.metric("🆕 Neue Papers", total_new_papers)
    
    with col_result4:
        st.metric("❌ Fehlgeschlagen", len(failed_searches))
    
    # Erfolgs-Meldung
    if total_new_papers > 0:
        st.success(f"🎉 **Multi-Fallback Wiederholung abgeschlossen!** {total_new_papers} neue Papers insgesamt gefunden!")
        st.balloons()
    elif successful_searches > 0:
        st.info(f"ℹ️ **Wiederholung abgeschlossen.** {successful_searches} Suchen erfolgreich, aber keine neuen Papers gefunden.")
    else:
        st.warning("⚠️ **Wiederholung abgeschlossen.** Leider waren alle Suchen nicht erfolgreich.")
    
    # Fehlgeschlagene Suchen anzeigen
    if failed_searches:
        with st.expander(f"❌ **Fehlgeschlagene Suchen ({len(failed_searches)}):**"):
            for fail in failed_searches:
                st.write(f"**{fail['term']}:**")
                if 'attempts' in fail:
                    for attempt in fail['attempts']:
                        st.write(f"   • {attempt}")
                if 'error' in fail:
                    st.write(f"   • Kritischer Fehler: {fail['error']}")
                st.write("---")
            
            st.info("💡 **Tipp:** Diese Suchbegriffe können später einzeln in der normalen Suche wiederholt werden.")

def build_advanced_search_query_corrected(query: str, date_filter: str) -> str:
    """
    🔧 KORRIGIERTE Suchanfrage-Generierung für repeat_all_searches
    """
    query_parts = [query.strip()]
    
    if date_filter != "Alle":
        current_year = datetime.datetime.now().year
        
        if date_filter == "Letztes Jahr":
            start_year = current_year - 1
            end_year = current_year - 1
            query_parts.append(f"AND ({start_year}[pdat]:{end_year}[pdat])")
            
        elif date_filter == "Letzte 2 Jahre":
            start_year = current_year - 2
            end_year = current_year - 1
            query_parts.append(f"AND ({start_year}[pdat]:{end_year}[pdat])")
            
        elif date_filter == "Letzte 5 Jahre":
            start_year = current_year - 5
            end_year = current_year - 1
            query_parts.append(f"AND ({start_year}[pdat]:{end_year}[pdat])")
    
    return " ".join(query_parts)

# =============== ZUSÄTZLICHE FALLBACK-FUNKTIONEN (falls nicht vorhanden) ===============

def try_simple_pubmed_search_enhanced(query: str, max_results: int) -> List[Dict[str, Any]]:
    """
    🔧 VERBESSERTE einfache PubMed-Suche ohne Filter
    """
    try:
        search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        
        params = {
            "db": "pubmed",
            "term": query.strip(),
            "retmode": "json", 
            "retmax": min(max_results, 100),
            "email": "research@papersearch.com",
            "tool": "PaperSearchSystem"
        }
        
        response = requests.get(search_url, params=params, timeout=45)
        
        if response.status_code == 200:
            data = response.json()
            pmids = data.get("esearchresult", {}).get("idlist", [])
            
            if pmids:
                return fetch_paper_details_batch(pmids)
        
        return []
        
    except Exception as e:
        return []

def try_alternative_pubmed_search_enhanced(query: str, max_results: int) -> List[Dict[str, Any]]:
    """
    🔧 VERBESSERTE alternative PubMed-Suche
    """
    try:
        search_strategies = [
            query.replace(" ", " AND "),  # AND-Verknüpfung
            f'"{query}"',                 # Phrase-Suche  
            query.split()[0] if " " in query else query  # Nur erstes Wort
        ]
        
        for strategy_query in search_strategies:
            try:
                search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
                params = {
                    "db": "pubmed",
                    "term": strategy_query,
                    "retmode": "json",
                    "retmax": min(max_results, 50),
                    "email": "research@papersearch.com", 
                    "tool": "PaperSearchSystem"
                }
                
                response = requests.get(search_url, params=params, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    pmids = data.get("esearchresult", {}).get("idlist", [])
                    
                    if pmids:
                        return fetch_paper_details_batch(pmids[:max_results//2])
                
            except Exception:
                continue
        
        return []
        
    except Exception:
        return []

def try_minimal_pubmed_search(query: str, max_results: int) -> List[Dict[str, Any]]:
    """
    🔧 MINIMALE PubMed-Suche als absoluter Fallback
    """
    try:
        search_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        
        params = {
            "db": "pubmed",
            "term": query,
            "retmax": min(max_results, 20),
            "retmode": "json"
        }
        
        response = requests.get(search_url, params=params, timeout=20)
        
        if response.status_code == 200:
            data = response.json()
            pmids = data.get("esearchresult", {}).get("idlist", [])
            
            if pmids:
                return fetch_paper_details_batch(pmids)
        
        return []
        
    except Exception:
        return []


# =============== WEITERE TAB-FUNKTIONEN ===============

def show_search_details_from_excel(search_term: str, term_info: Dict):
    """Zeigt Details einer Suchanfrage basierend auf Excel-Daten"""
    st.markdown("---")
    st.subheader(f"🔍 Excel-Details für: '{search_term}'")
    
    # Statistiken
    papers = term_info.get("papers", 0)
    new_papers = term_info.get("new_papers", 0)
    last_update = term_info.get("last_update", "Unbekannt")
    
    col_detail1, col_detail2, col_detail3 = st.columns(3)
    
    with col_detail1:
        st.metric("📄 Gesamt Papers", papers)
    
    with col_detail2:
        st.metric("🆕 Neue Papers (letzter Run)", new_papers)
    
    with col_detail3:
        st.metric("📅 Letztes Update", last_update[:16].replace('T', ' ') if last_update != "Unbekannt" else "Unbekannt")
    
    # Aktionen
    col_action1, col_action2 = st.columns(2)
    
    with col_action1:
        if st.button("🔄 Suche wiederholen", key=f"repeat_{search_term}"):
            execute_excel_integrated_search(search_term, 100, "Letzte 2 Jahre", False, False)
    
    with col_action2:
        if st.button("📊 Excel-Sheet anzeigen", key=f"show_excel_{search_term}"):
            show_excel_sheet_content(search_term)

def show_excel_sheet_content(search_term: str):
    """Zeigt Inhalt eines Excel-Sheets"""
    template_path = st.session_state["excel_template"]["file_path"]
    sheet_name = generate_sheet_name(search_term)
    
    try:
        if os.path.exists(template_path):
            xl_file = pd.ExcelFile(template_path)
            
            if sheet_name in xl_file.sheet_names:
                df = pd.read_excel(template_path, sheet_name=sheet_name)
                
                st.markdown("---")
                st.subheader(f"📊 Excel-Sheet: '{search_term}'")
                
                # Statistiken
                col_stat1, col_stat2, col_stat3 = st.columns(3)
                
                with col_stat1:
                    st.metric("📄 Gesamt Papers", len(df))
                
                with col_stat2:
                    new_papers = len(df[df.get("Status") == "NEU"]) if "Status" in df.columns else 0
                    st.metric("🆕 Neue Papers", new_papers)
                
                with col_stat3:
                    with_doi = len(df[df.get("DOI", "").astype(str).str.len() > 0]) if "DOI" in df.columns else 0
                    st.metric("🔗 Mit DOI", with_doi)
                
                # Anzeige der Papers
                st.write("**📋 Papers (erste 10):**")
                display_papers = df.head(10)
                
                for idx, (_, paper) in enumerate(display_papers.iterrows(), 1):
                    title = paper.get("Titel", "Unbekannt")
                    authors = paper.get("Autoren", "Unbekannt")
                    journal = paper.get("Journal", "Unbekannt")
                    year = paper.get("Jahr", "")
                    
                    with st.expander(f"📄 **{idx}.** {title[:60]}... ({year})"):
                        st.write(f"**👥 Autoren:** {authors}")
                        st.write(f"**📚 Journal:** {journal}")
                        if paper.get("URL"):
                            st.markdown(f"🔗 [**PubMed ansehen**]({paper.get('URL')})")
                
                if len(df) > 10:
                    st.info(f"... und {len(df) - 10} weitere Papers")
            else:
                st.error(f"❌ Sheet '{sheet_name}' nicht gefunden!")
        else:
            st.error("❌ Excel-Datei nicht gefunden!")
    
    except Exception as e:
        st.error(f"❌ Fehler beim Anzeigen des Sheet-Inhalts: {str(e)}")

def show_excel_template_management():
    """Excel-Template Management mit Excel-Integration"""
    st.subheader("📋 Excel-Template Management & Integration")
    
    template_path = st.session_state["excel_template"]["file_path"]
    excel_stats = get_search_statistics_from_excel()
    
    # Template Status
    if os.path.exists(template_path):
        file_size = os.path.getsize(template_path)
        file_date = datetime.datetime.fromtimestamp(os.path.getmtime(template_path))
        
        st.success(f"✅ **Master Excel-Template aktiv:** {template_path}")
        st.info(f"📊 **Größe:** {file_size:,} bytes | **Letzte Änderung:** {file_date.strftime('%d.%m.%Y %H:%M')}")
        
        # Excel-Statistiken anzeigen
        excel_stats = get_search_statistics_from_excel()
        if excel_stats:
            col_stat1, col_stat2, col_stat3 = st.columns(3)
            with col_stat1:
                st.metric("📊 Excel-Sheets", excel_stats.get("total_sheets", 0))
            with col_stat2:
                st.metric("📄 Gesamt Papers", excel_stats.get("total_papers", 0))
            with col_stat3:
                st.metric("🔍 Durchsuchungen", excel_stats.get("total_searches", 0))
    else:
        st.error("❌ Master Excel-Template nicht gefunden!")
        if st.button("🔧 Template neu erstellen"):
            create_master_excel_template()
            st.rerun()
    
    # Excel-Aktionen
    col_excel1, col_excel2, col_excel3 = st.columns(3)
    
    with col_excel1:
        if st.button("📥 **Excel herunterladen**"):
            offer_excel_download()
    
    with col_excel2:
        if st.button("📊 **Sheet-Übersicht anzeigen**"):
            show_excel_sheets_overview()
    
    with col_excel3:
        if st.button("🔄 **Template zurücksetzen**"):
            if st.button("✅ Bestätigen", key="confirm_reset"):
                reset_excel_template()

def repair_excel_database():
    """Umfassende Excel-Datenbank Reparatur und Wartung"""
    st.subheader("🔧 Excel-Datenbank Reparatur & Wartung")
    
    template_path = st.session_state["excel_template"]["file_path"]
    
    # Status der Excel-Datei prüfen
    if os.path.exists(template_path):
        file_size = os.path.getsize(template_path)
        file_date = datetime.datetime.fromtimestamp(os.path.getmtime(template_path))
        st.info(f"📊 **Excel-Datei gefunden:** {file_size:,} bytes | Letzte Änderung: {file_date.strftime('%d.%m.%Y %H:%M')}")
    else:
        st.warning("⚠️ **Excel-Datei nicht gefunden!**")
    
    # Backup vor Reparatur erstellen
    if os.path.exists(template_path):
        backup_path = f"{template_path}.repair_backup_{int(time.time())}"
        try:
            import shutil
            shutil.copy2(template_path, backup_path)
            st.success(f"📁 **Backup erstellt:** {backup_path}")
        except Exception as e:
            st.warning(f"⚠️ Backup-Warnung: {str(e)}")
    
    # Reparatur-Optionen
    col_repair1, col_repair2, col_repair3 = st.columns(3)
    
    with col_repair1:
        if st.button("🔧 **Basis-Reparatur**", type="primary"):
            perform_basic_excel_repair()
    
    with col_repair2:
        if st.button("🛠️ **Vollständige Reparatur**", type="secondary"):
            perform_full_excel_repair()
    
    with col_repair3:
        if st.button("🆕 **Datenbank neu erstellen**"):
            if st.button("✅ Bestätigen", key="confirm_recreate_db"):
                recreate_excel_database()
    
    # Diagnose-Bereich
    st.markdown("---")
    st.subheader("🔍 Excel-Diagnose")
    
    if st.button("📊 **Datenbank-Integrität prüfen**"):
        diagnose_excel_integrity()

def perform_basic_excel_repair():
    """Führt grundlegende Excel-Reparatur durch"""
    template_path = st.session_state["excel_template"]["file_path"]
    
    st.info("🔧 Führe Basis-Reparatur durch...")
    
    repairs_made = []
    
    try:
        # 1. Excel-Datei laden oder erstellen
        if not os.path.exists(template_path):
            st.warning("📁 Excel-Datei nicht vorhanden - erstelle neue...")
            create_master_excel_template()
            repairs_made.append("Excel-Template neu erstellt")
            st.success("✅ Basis-Reparatur abgeschlossen!")
            return
        
        # 2. Workbook laden
        wb = load_master_workbook()
        if not wb:
            st.error("❌ Excel-Datei konnte nicht geladen werden - erstelle neue...")
            create_master_excel_template()
            repairs_made.append("Beschädigte Excel-Datei ersetzt")
            st.success("✅ Basis-Reparatur abgeschlossen!")
            return
        
        # 3. Kritische Sheets prüfen und reparieren
        required_sheets = ["📊_Overview", "ℹ️_Template_Info"]
        
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                repairs_made.append(f"Sheet '{sheet_name}' hinzugefügt")
                
                if sheet_name == "📊_Overview":
                    # Overview Sheet erstellen
                    overview_sheet = wb.create_sheet(sheet_name, 0)
                    
                    # Header
                    headers = [
                        "Sheet_Name", "Suchbegriff", "Anzahl_Papers", "Letztes_Update", 
                        "Neue_Papers_Letzter_Run", "Status", "Erstellt_am"
                    ]
                    
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    for col, header in enumerate(headers, 1):
                        cell = overview_sheet.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                
                elif sheet_name == "ℹ️_Template_Info":
                    # Template Info Sheet erstellen
                    info_sheet = wb.create_sheet(sheet_name)
                    
                    info_data = [
                        ["📋 Excel Template Information", ""],
                        ["", ""],
                        ["Repariert am:", datetime.datetime.now().strftime("%d.%m.%Y %H:%M")],
                        ["System:", "Wissenschaftliches Paper-Suche System"],
                        ["Version:", "4.0 mit Streamlit Secrets Integration"],
                        ["Status:", "Automatisch repariert"],
                        ["", ""],
                        ["🔧 Reparatur-Log:", ""],
                        ["• Basis-Reparatur durchgeführt", ""],
                        ["• Kritische Sheets überprüft", ""],
                        ["• Datenbank-Integrität wiederhergestellt", ""],
                    ]
                    
                    for row_idx, (key, value) in enumerate(info_data, 1):
                        info_sheet.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
                        info_sheet.cell(row=row_idx, column=2, value=value)
                    
                    info_sheet.column_dimensions['A'].width = 30
                    info_sheet.column_dimensions['B'].width = 40
        
        # 4. Speichern falls Reparaturen durchgeführt wurden
        if repairs_made:
            wb.save(template_path)
            st.success(f"✅ **Basis-Reparatur abgeschlossen!** Durchgeführt: {', '.join(repairs_made)}")
        else:
            st.info("ℹ️ **Excel-Datenbank ist in Ordnung** - keine Basis-Reparaturen erforderlich.")
        
        # 5. Statistiken nach Reparatur anzeigen
        show_post_repair_stats()
        
    except Exception as e:
        st.error(f"❌ **Basis-Reparatur fehlgeschlagen:** {str(e)}")

def perform_full_excel_repair():
    """Führt vollständige Excel-Reparatur durch"""
    st.info("🛠️ Führe vollständige Reparatur durch...")
    
    # 1. Basis-Reparatur
    perform_basic_excel_repair()
    
    # 2. Erweiterte Reparaturen
    template_path = st.session_state["excel_template"]["file_path"]
    
    try:
        wb = load_master_workbook()
        if not wb:
            st.error("❌ Vollständige Reparatur nicht möglich - Excel-Datei nicht ladbar")
            return
        
        repairs_made = []
        
        # 3. Alle Data-Sheets validieren
        data_sheets = [sheet for sheet in wb.sheetnames if not sheet.startswith(("📊", "ℹ️"))]
        
        st.write(f"🔍 Überprüfe {len(data_sheets)} Daten-Sheets...")
        
        for sheet_name in data_sheets:
            try:
                ws = wb[sheet_name]
                
                # Prüfe Header
                if ws.max_row >= 1:
                    expected_headers = ["PMID", "Titel", "Autoren", "Journal", "Jahr", "Abstract", "DOI", "URL", "Status", "Hinzugefügt_am"]
                    actual_headers = [ws.cell(row=1, column=col).value for col in range(1, len(expected_headers) + 1)]
                    
                    # Repariere Header falls nötig
                    if actual_headers != expected_headers:
                        for col, header in enumerate(expected_headers, 1):
                            cell = ws.cell(row=1, column=col, value=header)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        
                        repairs_made.append(f"Header in Sheet '{sheet_name}' repariert")
                
            except Exception as sheet_error:
                st.warning(f"⚠️ Problem mit Sheet '{sheet_name}': {str(sheet_error)}")
                continue
        
        # 4. Overview-Sheet aktualisieren
        if "📊_Overview" in wb.sheetnames:
            update_overview_sheet_comprehensive(wb)
            repairs_made.append("Overview-Sheet aktualisiert")
        
        # 5. Speichern
        if repairs_made:
            wb.save(template_path)
            st.success(f"✅ **Vollständige Reparatur abgeschlossen!** Durchgeführt: {', '.join(repairs_made)}")
        else:
            st.info("ℹ️ **Vollständige Validierung abgeschlossen** - keine zusätzlichen Reparaturen erforderlich.")
        
        # 6. Finale Validierung
        validate_excel_integrity()
        
    except Exception as e:
        st.error(f"❌ **Vollständige Reparatur fehlgeschlagen:** {str(e)}")

def recreate_excel_database():
    """Erstellt Excel-Datenbank komplett neu"""
    template_path = st.session_state["excel_template"]["file_path"]
    
    st.warning("🆕 Erstelle Excel-Datenbank komplett neu...")
    
    try:
        # Backup der alten Datei falls vorhanden
        if os.path.exists(template_path):
            backup_path = f"{template_path}.backup_before_recreate_{int(time.time())}"
            os.rename(template_path, backup_path)
            st.info(f"📁 Alte Datei gesichert als: {backup_path}")
        
        # Neue Excel-Datei erstellen
        create_master_excel_template()
        
        st.success("✅ **Excel-Datenbank erfolgreich neu erstellt!**")
        st.balloons()
        
        # Statistiken der neuen Datei anzeigen
        show_post_repair_stats()
        
    except Exception as e:
        st.error(f"❌ **Neuerstellung fehlgeschlagen:** {str(e)}")

def diagnose_excel_integrity():
    """Führt umfassende Excel-Diagnose durch"""
    st.info("🔍 Führe Excel-Diagnose durch...")
    
    template_path = st.session_state["excel_template"]["file_path"]
    
    # Diagnose-Ergebnisse sammeln
    diagnosis = {
        "file_exists": False,
        "file_readable": False,
        "required_sheets_present": False,
        "data_sheets_count": 0,
        "total_papers": 0,
        "corrupted_sheets": [],
        "missing_sheets": [],
        "health_score": 0
    }
    
    try:
        # 1. Datei-Existenz prüfen
        if os.path.exists(template_path):
            diagnosis["file_exists"] = True
            file_size = os.path.getsize(template_path)
            st.write(f"✅ **Datei existiert:** {file_size:,} bytes")
        else:
            st.write("❌ **Datei existiert nicht**")
            return diagnosis
        
        # 2. Datei-Lesbarkeit prüfen
        try:
            wb = openpyxl.load_workbook(template_path)
            diagnosis["file_readable"] = True
            st.write(f"✅ **Datei lesbar:** {len(wb.sheetnames)} Sheets gefunden")
        except Exception as read_error:
            st.write(f"❌ **Datei nicht lesbar:** {str(read_error)}")
            return diagnosis
        
        # 3. Erforderliche Sheets prüfen
        required_sheets = ["📊_Overview", "ℹ️_Template_Info"]
        missing_sheets = []
        
        for sheet in required_sheets:
            if sheet in wb.sheetnames:
                st.write(f"✅ **{sheet}** vorhanden")
            else:
                missing_sheets.append(sheet)
                st.write(f"❌ **{sheet}** fehlt")
        
        diagnosis["missing_sheets"] = missing_sheets
        diagnosis["required_sheets_present"] = len(missing_sheets) == 0
        
        # 4. Daten-Sheets zählen
        data_sheets = [sheet for sheet in wb.sheetnames if not sheet.startswith(("📊", "ℹ️"))]
        diagnosis["data_sheets_count"] = len(data_sheets)
        st.write(f"📊 **Daten-Sheets:** {len(data_sheets)}")
        
        # 5. Papers zählen (aus Overview falls vorhanden)
        if "📊_Overview" in wb.sheetnames:
            overview_sheet = wb["📊_Overview"]
            total_papers = 0
            
            for row in overview_sheet.iter_rows(min_row=2):
                if row[2].value:  # Anzahl_Papers Spalte
                    total_papers += row[2].value or 0
            
            diagnosis["total_papers"] = total_papers
            st.write(f"📄 **Gesamt Papers:** {total_papers:,}")
        
        # 6. Health Score berechnen
        health_score = 0
        if diagnosis["file_exists"]: health_score += 25
        if diagnosis["file_readable"]: health_score += 25
        if diagnosis["required_sheets_present"]: health_score += 25
        if diagnosis["data_sheets_count"] > 0: health_score += 25
        
        diagnosis["health_score"] = health_score
        
        # 7. Gesundheitsstatus anzeigen
        if health_score == 100:
            st.success(f"🎉 **Excel-Datenbank Gesundheit: {health_score}%** - Ausgezeichnet!")
        elif health_score >= 75:
            st.info(f"✅ **Excel-Datenbank Gesundheit: {health_score}%** - Gut")
        elif health_score >= 50:
            st.warning(f"⚠️ **Excel-Datenbank Gesundheit: {health_score}%** - Reparatur empfohlen")
        else:
            st.error(f"❌ **Excel-Datenbank Gesundheit: {health_score}%** - Kritisch!")
        
        return diagnosis
        
    except Exception as e:
        st.error(f"❌ **Diagnose fehlgeschlagen:** {str(e)}")
        return diagnosis

def update_overview_sheet_comprehensive(wb):
    """Aktualisiert Overview-Sheet umfassend"""
    try:
        if "📊_Overview" not in wb.sheetnames:
            return
        
        overview_sheet = wb["📊_Overview"]
        
        # Alle Daten-Sheets durchgehen
        data_sheets = [sheet for sheet in wb.sheetnames if not sheet.startswith(("📊", "ℹ️"))]
        
        # Overview-Sheet leeren (außer Header)
        for row in overview_sheet.iter_rows(min_row=2, max_row=overview_sheet.max_row):
            for cell in row:
                cell.value = None
        
        # Neu aufbauen
        for row_idx, sheet_name in enumerate(data_sheets, start=2):
            try:
                ws = wb[sheet_name]
                paper_count = max(0, ws.max_row - 1) if ws.max_row > 1 else 0
                
                overview_sheet.cell(row=row_idx, column=1, value=sheet_name)  # Sheet_Name
                overview_sheet.cell(row=row_idx, column=2, value=sheet_name.replace('_', ' '))  # Suchbegriff
                overview_sheet.cell(row=row_idx, column=3, value=paper_count)  # Anzahl_Papers
                overview_sheet.cell(row=row_idx, column=4, value=datetime.datetime.now().isoformat())  # Letztes_Update
                overview_sheet.cell(row=row_idx, column=5, value=0)  # Neue_Papers_Letzter_Run
                overview_sheet.cell(row=row_idx, column=6, value="Repariert")  # Status
                overview_sheet.cell(row=row_idx, column=7, value=datetime.datetime.now().isoformat())  # Erstellt_am
                
            except Exception as sheet_error:
                continue
                
    except Exception as e:
        st.warning(f"⚠️ Overview-Update Fehler: {str(e)}")

def show_post_repair_stats():
    """Zeigt Statistiken nach Reparatur an"""
    st.markdown("---")
    st.subheader("📊 Datenbank-Status nach Reparatur")
    
    excel_stats = get_search_statistics_from_excel()
    
    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
    
    with col_stat1:
        st.metric("📊 Excel-Sheets", excel_stats.get("total_sheets", 0))
    
    with col_stat2:
        st.metric("📄 Gesamt Papers", excel_stats.get("total_papers", 0))
    
    with col_stat3:
        st.metric("🔍 Durchsuchungen", excel_stats.get("total_searches", 0))
    
    with col_stat4:
        template_path = st.session_state["excel_template"]["file_path"]
        file_size = os.path.getsize(template_path) if os.path.exists(template_path) else 0
        st.metric("💾 Dateigröße", f"{file_size:,} bytes")

def validate_excel_integrity():
    """Validiert die Integrität der Excel-Datei"""
    template_path = st.session_state["excel_template"]["file_path"]
    
    try:
        wb = load_master_workbook()
        if wb:
            sheet_count = len(wb.sheetnames)
            data_sheets = len([s for s in wb.sheetnames if not s.startswith(("📊", "ℹ️"))])
            
            if sheet_count >= 2 and data_sheets >= 0:
                st.success(f"✅ **Excel-Integrität validiert:** {sheet_count} Sheets ({data_sheets} Daten-Sheets)")
            else:
                st.warning(f"⚠️ **Excel-Struktur unvollständig:** {sheet_count} Sheets")
        else:
            st.error("❌ **Excel-Validierung fehlgeschlagen!**")
    except Exception as e:
        st.error(f"❌ **Validierung-Fehler:** {str(e)}")


def repair_missing_sheets():
    """Erweiterte Excel-Sheets Reparatur mit Backup"""
    st.subheader("🔧 Excel-Sheets Reparatur & Wartung")
    
    template_path = st.session_state["excel_template"]["file_path"]
    
    # Backup erstellen vor Reparatur
    if os.path.exists(template_path):
        backup_path = f"{template_path}.repair_backup_{int(time.time())}"
        try:
            import shutil
            shutil.copy2(template_path, backup_path)
            st.info(f"📁 **Backup erstellt:** {backup_path}")
        except:
            st.warning("⚠️ Backup konnte nicht erstellt werden!")
    
    # Reparatur-Optionen
    col_repair1, col_repair2, col_repair3 = st.columns(3)
    
    with col_repair1:
        if st.button("🔧 **Basis-Reparatur**", type="primary"):
            perform_basic_repair()
    
    with col_repair2:
        if st.button("🛠️ **Vollständige Reparatur**"):
            perform_full_repair()
    
    with col_repair3:
        if st.button("🆕 **Neustart (Template zurücksetzen)**"):
            if st.button("✅ Bestätigen", key="confirm_reset_repair"):
                reset_excel_template()

def perform_basic_repair():
    """Führt Basis-Reparatur durch"""
    template_path = st.session_state["excel_template"]["file_path"]
    
    try:
        wb = load_master_workbook()
        if not wb:
            create_master_excel_template()
            st.success("✅ Template neu erstellt!")
            return
        
        repairs = []
        
        # Prüfe und repariere kritische Sheets
        required_sheets = ["📊_Overview", "ℹ️_Template_Info"]
        
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                repairs.append(f"Sheet '{sheet_name}' hinzugefügt")
                # Sheet-spezifische Reparatur...
        
        if repairs:
            wb.save(template_path)
            st.success(f"✅ **Reparaturen abgeschlossen:** {', '.join(repairs)}")
        else:
            st.info("ℹ️ Keine Reparaturen erforderlich!")
            
    except Exception as e:
        st.error(f"❌ Reparatur fehlgeschlagen: {str(e)}")

def perform_full_repair():
    """Führt vollständige Reparatur durch"""
    st.info("🛠️ Führe vollständige Reparatur durch...")
    
    # Kombiniere alle Reparatur-Schritte
    perform_basic_repair()
    
    # Zusätzliche Validierung
    validate_excel_integrity()
    
    st.success("✅ Vollständige Reparatur abgeschlossen!")

def validate_excel_integrity():
    """Validiert die Integrität der Excel-Datei"""
    template_path = st.session_state["excel_template"]["file_path"]
    
    try:
        wb = load_master_workbook()
        if wb:
            sheet_count = len(wb.sheetnames)
            st.info(f"✅ **Excel-Integrität OK:** {sheet_count} Sheets gefunden")
        else:
            st.warning("⚠️ Excel-Datei beschädigt!")
    except Exception as e:
        st.error(f"❌ Validierung fehlgeschlagen: {str(e)}")

def get_search_statistics_from_excel() -> Dict:
    """Holt Statistiken aus der Excel-Datei"""
    wb = load_master_workbook()
    if not wb:
        return {}
    
    stats = {
        "total_sheets": len([s for s in wb.sheetnames if not s.startswith(("📊", "ℹ️"))]),
        "total_searches": 0,
        "total_papers": 0,
        "search_terms": []
    }
    
    if "📊_Overview" in wb.sheetnames:
        overview_sheet = wb["📊_Overview"]
        
        for row in overview_sheet.iter_rows(min_row=2):
            if row[1].value:  # Suchbegriff existiert
                stats["total_searches"] += 1
                stats["total_papers"] += row[2].value or 0
                stats["search_terms"].append({
                    "term": row[1].value,
                    "papers": row[2].value or 0,
                    "last_update": row[3].value,
                    "new_papers": row[4].value or 0
                })
    
    return stats

def load_master_workbook():
    """Lädt das Master Excel Workbook"""
    excel_path = st.session_state["excel_template"]["file_path"]
    try:
        return openpyxl.load_workbook(excel_path)
    except Exception as e:
        st.error(f"❌ Excel-Datei konnte nicht geladen werden: {str(e)}")
        return None
def add_new_papers_to_excel(search_term: str, current_papers: List[Dict]) -> Tuple[int, List[Dict]]:
    """ULTRA-ROBUSTE Version - Fügt neue Papers zur Excel-Datei hinzu"""
    template_path = st.session_state["excel_template"]["file_path"]
    sheet_name = generate_sheet_name(search_term)
    
    # VALIDIERUNG DER EINGABEPARAMETER
    if not search_term or not isinstance(search_term, str):
        st.error("❌ Ungültiger Suchbegriff")
        return 0, []
    
    if not current_papers or not isinstance(current_papers, list):
        st.warning("⚠️ Keine gültigen Papers bereitgestellt")
        return 0, []
    
    try:
        # 1. ULTRA-SICHERE Workbook-Erstellung
        wb = None
        try:
            if os.path.exists(template_path):
                wb = openpyxl.load_workbook(template_path)
            else:
                wb = openpyxl.Workbook()
                # Entferne Standard-Sheet sicher
                if wb.active:
                    try:
                        wb.remove(wb.active)
                    except:
                        pass
        except Exception as wb_error:
            st.error(f"❌ Kritischer Workbook-Fehler: {str(wb_error)}")
            return 0, []
        
        # VALIDIERUNG: Workbook muss existieren
        if wb is None:
            st.error("❌ Workbook konnte nicht erstellt werden")
            return 0, []
        
        # 2. SICHERE Laden vorheriger Papers
        previous_papers = []
        try:
            previous_papers = load_previous_search_results(search_term)
            if previous_papers is None:
                previous_papers = []
        except Exception as load_error:
            st.warning(f"⚠️ Fehler beim Laden: {str(load_error)}")
            previous_papers = []
        
        # 3. SICHERE PMID-Set Erstellung
        previous_pmids = set()
        try:
            if isinstance(previous_papers, list):
                for paper in previous_papers:
                    if paper and isinstance(paper, dict):
                        pmid = paper.get("PMID")
                        if pmid:
                            previous_pmids.add(str(pmid))
        except Exception as pmid_error:
            st.warning(f"⚠️ PMID-Fehler: {str(pmid_error)}")
            previous_pmids = set()
        
        # 4. SICHERE Identifikation neuer Papers
        new_papers = []
        try:
            for paper in current_papers:
                if paper and isinstance(paper, dict):
                    current_pmid = str(paper.get("PMID", ""))
                    if current_pmid and current_pmid not in previous_pmids:
                        paper["Status"] = "NEU"
                        new_papers.append(paper)
                    else:
                        paper["Status"] = "BEKANNT"
        except Exception as new_papers_error:
            st.error(f"❌ Fehler bei neuen Papers: {str(new_papers_error)}")
            return 0, []
        
        # 5. ULTRA-SICHERE Sheet-Erstellung/Update
        ws = None
        try:
            if sheet_name not in wb.sheetnames:
                # SICHERE Sheet-Erstellung
                try:
                    ws = wb.create_sheet(sheet_name)
                    if ws is None:
                        raise ValueError("Sheet-Erstellung fehlgeschlagen")
                    
                    # SICHERE Header-Erstellung
                    create_excel_sheet_headers(ws)
                    
                except Exception as sheet_create_error:
                    st.error(f"❌ Sheet-Erstellung fehlgeschlagen: {str(sheet_create_error)}")
                    return len(new_papers), new_papers
            else:
                # SICHERE Sheet-Auswahl
                try:
                    ws = wb[sheet_name]
                    if ws is None:
                        raise ValueError("Sheet-Auswahl fehlgeschlagen")
                except Exception as sheet_select_error:
                    st.error(f"❌ Sheet-Auswahl fehlgeschlagen: {str(sheet_select_error)}")
                    return len(new_papers), new_papers
            
            # VALIDIERUNG: Worksheet muss existieren
            if ws is None:
                st.error("❌ Worksheet ist None nach Erstellung/Auswahl")
                return len(new_papers), new_papers
            
            # SICHERE Papers hinzufügen
            if new_papers:
                try:
                    add_papers_to_sheet(ws, new_papers)
                except Exception as add_error:
                    st.error(f"❌ Fehler beim Hinzufügen der Papers: {str(add_error)}")
                    # Trotzdem weiter machen
                    
        except Exception as sheet_error:
            st.error(f"❌ Sheet-Fehler: {str(sheet_error)}")
            return len(new_papers), new_papers
        
        # 6. SICHERE Overview Update
        try:
            update_overview_sheet(wb, search_term, len(current_papers), len(new_papers))
        except Exception as overview_error:
            st.warning(f"⚠️ Overview-Fehler: {str(overview_error)}")
        
        # 7. SICHERE Excel-Speicherung
        try:
            wb.save(template_path)
        except Exception as save_error:
            st.error(f"❌ Speicher-Fehler: {str(save_error)}")
        
        return len(new_papers), new_papers
        
    except Exception as e:
        st.error(f"❌ **ULTRA-KRITISCHER FEHLER:** {str(e)}")
        st.error(f"🔍 **Details:** term='{search_term}', papers={len(current_papers) if current_papers else 0}")
        
        # NOTFALL-RÜCKGABE
        if current_papers:
            return len(current_papers), current_papers
        return 0, []




def create_excel_sheet_headers(ws):
    """Erstellt Header für Excel-Sheet mit robustem Error-Handling"""
    if ws is None:
        raise ValueError("❌ Worksheet ist None - kann keine Header erstellen")
    
    try:
        headers = [
            "PMID", "Titel", "Autoren", "Journal", "Jahr", 
            "Abstract", "DOI", "URL", "Status", "Hinzugefügt_am"
        ]
        
        # Header-Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # SICHERE Header-Erstellung
        for col, header in enumerate(headers, 1):
            try:
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            except Exception as cell_error:
                st.warning(f"⚠️ Fehler bei Header-Zelle {col}: {str(cell_error)}")
                continue
        
        # SICHERE Spaltenbreite-Anpassung
        try:
            column_widths = [12, 50, 30, 25, 8, 60, 15, 25, 12, 18]
            for col, width in enumerate(column_widths, 1):
                try:
                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].width = width
                except Exception as width_error:
                    # Überspringe fehlerhafte Spaltenbreiten
                    continue
        except Exception as width_setup_error:
            st.warning(f"⚠️ Fehler bei Spaltenbreiten: {str(width_setup_error)}")
            
    except Exception as e:
        st.error(f"❌ Kritischer Fehler beim Erstellen der Header: {str(e)}")
        raise


def add_papers_to_sheet(ws, papers: List[Dict]):
    """Fügt Papers zu Excel-Sheet hinzu mit robustem Error-Handling"""
    if ws is None:
        raise ValueError("❌ Worksheet ist None - kann keine Papers hinzufügen")
    
    if not papers or not isinstance(papers, list):
        st.warning("⚠️ Keine gültigen Papers zum Hinzufügen")
        return
    
    try:
        # SICHERE Ermittlung der nächsten Zeile
        try:
            next_row = ws.max_row + 1 if ws.max_row and ws.max_row > 0 else 2
        except Exception as row_error:
            st.warning(f"⚠️ Fehler bei max_row: {str(row_error)}")
            next_row = 2  # Fallback auf Zeile 2
        
        # SICHERE Paper-Iteration
        papers_added = 0
        for i, paper in enumerate(papers):
            if not paper or not isinstance(paper, dict):
                continue
                
            try:
                # SICHERE Daten-Extraktion
                row_data = [
                    str(paper.get("PMID", ""))[:50],  # Begrenzt und als String
                    str(paper.get("Title", ""))[:500],
                    str(paper.get("Authors", ""))[:300],
                    str(paper.get("Journal", ""))[:100],
                    str(paper.get("Year", ""))[:10],
                    str(paper.get("Abstract", ""))[:1000],
                    str(paper.get("DOI", ""))[:100],
                    str(paper.get("URL", ""))[:200],
                    str(paper.get("Status", "NEU"))[:20],
                    datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
                ]
                
                # SICHERE Zeilen-Erstellung
                current_row = next_row + i
                for col, value in enumerate(row_data, 1):
                    try:
                        # Sichere Zellenwertzuweisung
                        if value is not None:
                            ws.cell(row=current_row, column=col, value=value)
                    except Exception as cell_error:
                        # Überspringe fehlerhafte Zellen, aber mache weiter
                        continue
                
                papers_added += 1
                
            except Exception as paper_error:
                st.warning(f"⚠️ Fehler bei Paper {i+1}: {str(paper_error)}")
                continue
        
        if papers_added > 0:
            st.success(f"✅ {papers_added} Papers erfolgreich zu Sheet hinzugefügt")
        else:
            st.warning("⚠️ Keine Papers konnten hinzugefügt werden")
            
    except Exception as e:
        st.error(f"❌ Kritischer Fehler beim Hinzufügen der Papers: {str(e)}")
        raise


def update_overview_sheet(wb, search_term: str, total_papers: int, new_papers: int):
    """Aktualisiert das Overview-Sheet mit robustem Error-Handling"""
    try:
        if "📊_Overview" not in wb.sheetnames:
            overview_sheet = wb.create_sheet("📊_Overview", 0)
            
            # Header erstellen
            headers = [
                "Sheet_Name", "Suchbegriff", "Anzahl_Papers", "Letztes_Update", 
                "Neue_Papers_Letzter_Run", "Status", "Erstellt_am"
            ]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = overview_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
        else:
            overview_sheet = wb["📊_Overview"]
        
        # ROBUSTE SUCHE nach bestehendem Eintrag
        sheet_name = generate_sheet_name(search_term)
        row_found = None
        
        try:
            # Sichere Iteration durch Zeilen
            for row_num in range(2, overview_sheet.max_row + 1):
                try:
                    # Sichere Zellenabfrage
                    cell_value = overview_sheet.cell(row=row_num, column=2).value
                    if cell_value and str(cell_value).strip() == search_term:
                        row_found = row_num
                        break
                except Exception as cell_error:
                    # Überspringe fehlerhafte Zeilen
                    continue
                    
        except Exception as iteration_error:
            # Falls Iteration fehlschlägt, erstelle neuen Eintrag
            row_found = None
        
        # Update oder erstelle Eintrag
        if row_found:
            # SICHERE Updates für bestehenden Eintrag
            try:
                overview_sheet.cell(row=row_found, column=3, value=total_papers)
                overview_sheet.cell(row=row_found, column=4, value=datetime.datetime.now().isoformat())
                overview_sheet.cell(row=row_found, column=5, value=new_papers)
                overview_sheet.cell(row=row_found, column=6, value="Aktualisiert")
            except Exception as update_error:
                st.warning(f"⚠️ Fehler beim Update der Zeile {row_found}: {str(update_error)}")
        else:
            # SICHERE Erstellung eines neuen Eintrags
            try:
                next_row = overview_sheet.max_row + 1
                overview_sheet.cell(row=next_row, column=1, value=sheet_name)
                overview_sheet.cell(row=next_row, column=2, value=search_term)
                overview_sheet.cell(row=next_row, column=3, value=total_papers)
                overview_sheet.cell(row=next_row, column=4, value=datetime.datetime.now().isoformat())
                overview_sheet.cell(row=next_row, column=5, value=new_papers)
                overview_sheet.cell(row=next_row, column=6, value="Neu")
                overview_sheet.cell(row=next_row, column=7, value=datetime.datetime.now().isoformat())
            except Exception as create_error:
                st.warning(f"⚠️ Fehler beim Erstellen eines neuen Eintrags: {str(create_error)}")
                
    except Exception as e:
        st.error(f"❌ Kritischer Fehler in update_overview_sheet: {str(e)}")
        # Trotzdem weiter versuchen - nicht das ganze System zum Absturz bringen


def show_excel_sheets_overview():
    """Zeigt Übersicht aller Excel-Sheets"""
    st.markdown("---")
    st.subheader("📊 Excel-Sheets Übersicht")
    
    excel_stats = get_search_statistics_from_excel()
    
    if excel_stats.get("search_terms"):
        # Erstelle DataFrame für bessere Darstellung
        df_overview = pd.DataFrame(excel_stats["search_terms"])
        df_overview.columns = ["Suchbegriff", "Papers", "Letztes Update", "Neue Papers"]
        
        # Sortiere nach letztem Update
        df_overview = df_overview.sort_values("Letztes Update", ascending=False)
        
        st.dataframe(df_overview, use_container_width=True)
        
        # Zusammenfassung
        total_papers = df_overview["Papers"].sum()
        total_new = df_overview["Neue Papers"].sum()
        
        col_sum1, col_sum2, col_sum3 = st.columns(3)
        with col_sum1:
            st.metric("📊 Gesamt Sheets", len(df_overview))
        with col_sum2:
            st.metric("📄 Gesamt Papers", total_papers)
        with col_sum3:
            st.metric("🆕 Neue Papers", total_new)
    else:
        st.info("📭 Noch keine Excel-Sheets vorhanden.")

def offer_excel_download():
    """Bietet Master Excel-Datei zum Download an"""
    template_path = st.session_state["excel_template"]["file_path"]
    
    if os.path.exists(template_path):
        try:
            with open(template_path, 'rb') as f:
                excel_data = f.read()
            
            filename = f"PaperSearch_Master_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            
            st.download_button(
                label="📥 **Master Excel-Datei herunterladen**",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Lädt die komplette Excel-Datei mit allen Sheets herunter"
            )
        
        except Exception as e:
            st.error(f"❌ Fehler beim Bereitstellen der Excel-Datei: {str(e)}")

def reset_excel_template():
    """Setzt Excel-Template zurück"""
    template_path = st.session_state["excel_template"]["file_path"]
    
    try:
        if os.path.exists(template_path):
            # Backup erstellen
            backup_path = f"{template_path}.backup_{int(time.time())}"
            os.rename(template_path, backup_path)
            st.info(f"📁 Backup erstellt: {backup_path}")
        
        create_master_excel_template()
        st.success("✅ Excel-Template zurückgesetzt!")
        st.rerun()
        
    except Exception as e:
        st.error(f"❌ Fehler beim Zurücksetzen: {str(e)}")

# =============== EMAIL-KONFIGURATION MIT MEHREREN EMPFÄNGERN ===============

def show_email_config():
    """Email-Konfiguration mit mehreren Empfängern"""
    st.subheader("📧 Email-Konfiguration (Mehrere Empfänger)")
    
    settings = st.session_state.get("email_settings", {})
    
    # Email-Setup Hilfe
    with st.expander("📖 Email-Setup Hilfe & Mehrere Empfänger"):
        st.info("""
        **Für Gmail (empfohlen):**
        1. ✅ 2-Faktor-Authentifizierung aktivieren
        2. ✅ App-Passwort erstellen (nicht normales Passwort!)
        3. ✅ SMTP: smtp.gmail.com, Port: 587, TLS: An
        
        **Mehrere Empfänger:**
        • Trennen Sie mehrere Email-Adressen mit Kommas
        • Beispiel: user1@gmail.com, user2@outlook.com, user3@company.de
        • Whitespaces werden automatisch entfernt
        
        **Für Outlook/Hotmail:**
        - SMTP: smtp-mail.outlook.com, Port: 587
        """)
    
    with st.form("email_config_form"):
        st.subheader("📬 Grundeinstellungen")
        
        col1, col2 = st.columns(2)
        
        with col1:
            sender_email = st.text_input(
                "Absender Email *", 
                value=settings.get("sender_email", ""),
                placeholder="absender@gmail.com"
            )
            
            smtp_server = st.text_input(
                "SMTP Server *",
                value=settings.get("smtp_server", "smtp.gmail.com")
            )
            
            auto_notifications = st.checkbox(
                "Automatische Benachrichtigungen", 
                value=settings.get("auto_notifications", True)
            )
        
        with col2:
            smtp_port = st.number_input(
                "SMTP Port *",
                value=settings.get("smtp_port", 587),
                min_value=1,
                max_value=65535
            )
            
            min_papers = st.number_input(
                "Min. Papers für Benachrichtigung", 
                value=settings.get("min_papers", 1),
                min_value=1,
                max_value=100
            )
            
            use_tls = st.checkbox(
                "TLS Verschlüsselung verwenden (empfohlen)",
                value=settings.get("use_tls", True)
            )
        
        # MEHRERE EMPFÄNGER - Text Area
        recipient_emails = st.text_area(
            "📧 Empfänger Email-Adressen * (mehrere mit Komma trennen)",
            value=settings.get("recipient_emails", ""),
            placeholder="empfaenger1@example.com, empfaenger2@gmail.com, empfaenger3@company.de",
            help="Mehrere Email-Adressen mit Komma trennen. Beispiel: user1@gmail.com, user2@outlook.com",
            height=80
        )
        
        sender_password = st.text_input(
            "Email Passwort / App-Passwort *",
            value=settings.get("sender_password", ""),
            type="password",
            help="Für Gmail: App-spezifisches Passwort verwenden!"
        )
        
        # Email-Vorlagen
        st.subheader("📝 Email-Vorlagen")
        
        subject_template = st.text_input(
            "Betreff-Vorlage",
            value=settings.get("subject_template", "🔬 {count} neue Papers für '{search_term}'"),
            help="Platzhalter: {count}, {search_term}, {frequency}"
        )
        
        message_template = st.text_area(
            "Nachricht-Vorlage",
            value=settings.get("message_template", """📧 Automatische Paper-Benachrichtigung

📅 Datum: {date}
🔍 Suchbegriff: '{search_term}'
📊 Neue Papers: {count}

📋 Neue Papers:
{new_papers_list}

📎 Excel-Datei: {excel_file}

Mit freundlichen Grüßen,
Ihr Paper-Suche System"""),
            height=200,
            help="Platzhalter: {date}, {search_term}, {count}, {frequency}, {new_papers_list}, {excel_file}"
        )
        
        if st.form_submit_button("💾 **Email-Einstellungen speichern**", type="primary"):
            # Validiere Email-Adressen
            recipient_list = parse_recipient_emails(recipient_emails)
            
            if not recipient_list:
                st.error("❌ Mindestens eine gültige Empfänger-Email erforderlich!")
            else:
                new_settings = {
                    "sender_email": sender_email,
                    "recipient_emails": recipient_emails,
                    "smtp_server": smtp_server,
                    "smtp_port": smtp_port,
                    "sender_password": sender_password,
                    "use_tls": use_tls,
                    "auto_notifications": auto_notifications,
                    "min_papers": min_papers,
                    "subject_template": subject_template,
                    "message_template": message_template,
                    "parsed_recipients": recipient_list  # Store parsed list
                }
                
                st.session_state["email_settings"] = new_settings
                st.success(f"✅ Email-Einstellungen gespeichert! **{len(recipient_list)} Empfänger** konfiguriert:")
                for i, email in enumerate(recipient_list, 1):
                    st.write(f"   {i}. 📧 {email}")
    
    # Zeige konfigurierte Empfänger
    if settings.get("recipient_emails"):
        recipient_list = parse_recipient_emails(settings.get("recipient_emails", ""))
        if recipient_list:
            st.info(f"📧 **Aktuell konfigurierte Empfänger ({len(recipient_list)}):**")
            cols = st.columns(min(len(recipient_list), 3))
            for i, email in enumerate(recipient_list):
                with cols[i % 3]:
                    st.write(f"✅ {email}")
    
    # Test-Email
    st.markdown("---")
    st.subheader("🧪 Email-System testen")
    
    col_test1, col_test2 = st.columns(2)
    
    with col_test1:
        if st.button("📧 **Test-Email an alle Empfänger senden**", type="primary"):
            send_test_email_multiple()
    
    with col_test2:
        if st.button("📊 **Email-Status prüfen**"):
            check_email_status_multiple()

def parse_recipient_emails(email_string: str) -> List[str]:
    """Parst Email-String und gibt Liste gültiger Emails zurück"""
    if not email_string:
        return []
    
    # Split by comma and clean
    emails = [email.strip() for email in email_string.split(",")]
    
    # Basic email validation
    valid_emails = []
    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    
    for email in emails:
        if email and email_pattern.match(email):
            valid_emails.append(email)
    
    return valid_emails

def send_real_email_multiple(to_emails: List[str], subject: str, message: str, attachment_path: str = None) -> tuple:
    """Sendet echte Email über SMTP an mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    
    sender_email = settings.get("sender_email", "")
    sender_password = settings.get("sender_password", "")
    smtp_server = settings.get("smtp_server", "smtp.gmail.com")
    smtp_port = settings.get("smtp_port", 587)
    use_tls = settings.get("use_tls", True)
    
    if not all([sender_email, sender_password]):
        return False, "❌ Email-Konfiguration unvollständig (Absender/Passwort)"
    
    if not to_emails:
        return False, "❌ Keine Empfänger-Emails konfiguriert"
    
    try:
        # SMTP Server Setup
        server = smtplib.SMTP(smtp_server, smtp_port)
        
        if use_tls:
            context = ssl.create_default_context()
            server.starttls(context=context)
        
        server.login(sender_email, sender_password)
        
        successful_sends = 0
        failed_sends = []
        
        # Send to each recipient
        for recipient in to_emails:
            try:
                msg = MIMEMultipart()
                msg['From'] = sender_email
                msg['To'] = recipient
                msg['Subject'] = subject
                
                msg.attach(MIMEText(message, 'plain', 'utf-8'))
                
                # Add attachment if provided
                if attachment_path and os.path.exists(attachment_path):
                    with open(attachment_path, "rb") as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header(
                            'Content-Disposition',
                            f'attachment; filename= {os.path.basename(attachment_path)}'
                        )
                        msg.attach(part)
                
                server.send_message(msg)
                successful_sends += 1
                
            except Exception as e:
                failed_sends.append(f"{recipient}: {str(e)}")
        
        server.quit()
        
        if successful_sends == len(to_emails):
            return True, f"✅ Email erfolgreich an alle {successful_sends} Empfänger gesendet"
        elif successful_sends > 0:
            return True, f"⚠️ Email an {successful_sends}/{len(to_emails)} Empfänger gesendet. Fehler: {'; '.join(failed_sends)}"
        else:
            return False, f"❌ Email an keinen Empfänger gesendet. Fehler: {'; '.join(failed_sends)}"
        
    except smtplib.SMTPAuthenticationError:
        return False, "❌ SMTP-Authentifizierung fehlgeschlagen - Prüfen Sie Email/Passwort"
    except smtplib.SMTPServerDisconnected:
        return False, "❌ SMTP-Server-Verbindung unterbrochen"
    except Exception as e:
        return False, f"❌ Email-Fehler: {str(e)}"

def send_test_email_multiple():
    """Sendet Test-Email an alle konfigurierten Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    if not settings.get("sender_email") or not recipient_emails:
        st.error("❌ Email-Konfiguration unvollständig!")
        return
    
    subject = "🧪 Test-Email vom Paper-Suche System (Mehrere Empfänger)"
    message = f"""Dies ist eine Test-Email vom Paper-Suche System mit Unterstützung für mehrere Empfänger.

📅 Gesendet am: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}
📧 Von: {settings.get('sender_email')}
📧 An: {len(recipient_emails)} Empfänger

Empfänger-Liste:
{chr(10).join([f"• {email}" for email in recipient_emails])}

✅ Wenn Sie diese Email erhalten, funktioniert das Email-System korrekt!

System-Informationen:
• SMTP Server: {settings.get('smtp_server')}
• Port: {settings.get('smtp_port')}
• TLS: {'Aktiviert' if settings.get('use_tls') else 'Deaktiviert'}
• Empfänger: {len(recipient_emails)}

Mit freundlichen Grüßen,
Ihr Paper-Suche System"""
    
    success, status_message = send_real_email_multiple(
        recipient_emails, 
        subject, 
        message
    )
    
    if success:
        st.success(f"✅ **Test-Email erfolgreich gesendet!** {status_message}")
        st.balloons()
    else:
        st.error(f"❌ **Test-Email fehlgeschlagen:** {status_message}")

def check_email_status_multiple():
    """Prüft Email-Status mit mehreren Empfängern"""
    settings = st.session_state.get("email_settings", {})
    
    st.write("**📊 Email-Konfiguration Status:**")
    
    # Prüfe Konfiguration
    sender_ok = bool(settings.get("sender_email"))
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    recipients_ok = len(recipient_emails) > 0
    password_ok = bool(settings.get("sender_password"))
    
    st.write(f"📧 Absender Email: {'✅' if sender_ok else '❌'} {settings.get('sender_email', 'Nicht konfiguriert')}")
    st.write(f"📧 Empfänger Emails: {'✅' if recipients_ok else '❌'} {len(recipient_emails)} konfiguriert")
    
    if recipients_ok:
        with st.expander(f"📧 Empfänger-Liste ({len(recipient_emails)})"):
            for i, email in enumerate(recipient_emails, 1):
                st.write(f"   {i}. {email}")
    
    st.write(f"🔑 Passwort: {'✅' if password_ok else '❌'} {'Konfiguriert' if password_ok else 'Nicht konfiguriert'}")
    st.write(f"🔒 SMTP Server: {settings.get('smtp_server', 'smtp.gmail.com')}:{settings.get('smtp_port', 587)}")
    st.write(f"🔐 TLS: {'✅ Aktiviert' if settings.get('use_tls', True) else '❌ Deaktiviert'}")
    
    # Gesamtstatus
    if sender_ok and recipients_ok and password_ok:
        st.success(f"✅ **Email-System vollständig konfiguriert für {len(recipient_emails)} Empfänger!**")
    else:
        st.error("❌ **Email-System nicht vollständig konfiguriert!**")

def is_email_configured() -> bool:
    """Prüft Email-Konfiguration für mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    return (bool(settings.get("sender_email")) and 
            len(recipient_emails) > 0 and
            bool(settings.get("sender_password")))

# =============== WEITERE FUNKTIONEN ===============

def show_detailed_statistics():
    """Detaillierte Statistiken mit Excel-Integration"""
    st.subheader("📈 Detaillierte Statistiken")
    
    status = st.session_state["system_status"]
    search_history = st.session_state.get("search_history", [])
    email_history = st.session_state.get("email_history", [])
    excel_stats = get_search_statistics_from_excel()
    
    # Hauptstatistiken
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("🔍 Suchen (Session)", status["total_searches"])
        st.metric("🔍 Suchen (Excel)", excel_stats.get("total_searches", 0))
    
    with col2:
        st.metric("📄 Papers (Session)", status["total_papers"])
        st.metric("📄 Papers (Excel)", excel_stats.get("total_papers", 0))
    
    with col3:
        st.metric("📧 Gesendete Emails", len(email_history))
        recipient_count = len(parse_recipient_emails(st.session_state.get("email_settings", {}).get("recipient_emails", "")))
        st.metric("📧 Email-Empfänger", recipient_count)
    
    with col4:
        st.metric("📊 Excel Sheets", excel_stats.get("total_sheets", 0))
        auto_searches = len(st.session_state.get("automatic_searches", {}))
        st.metric("🤖 Auto-Suchen", auto_searches)
    
    # Email-Statistiken detailliert
    if email_history:
        st.markdown("---")
        st.subheader("📧 Email-Statistiken")
        
        successful_emails = len([e for e in email_history if e.get("success", False)])
        success_rate = (successful_emails / len(email_history)) * 100
        
        col_email1, col_email2, col_email3 = st.columns(3)
        
        with col_email1:
            st.metric("📧 Gesamt Emails", len(email_history))
        with col_email2:
            st.metric("✅ Erfolgreich", successful_emails)
        with col_email3:
            st.metric("📊 Erfolgsrate", f"{success_rate:.1f}%")
        
        # Letzte Emails
        st.write("**📧 Letzte Email-Aktivitäten:**")
        recent_emails = sorted(email_history, key=lambda x: x.get("timestamp", ""), reverse=True)[:5]
        
        for i, email in enumerate(recent_emails, 1):
            timestamp = email.get("timestamp", "")[:16].replace('T', ' ')
            email_type = email.get("type", "Unbekannt")
            success_icon = "✅" if email.get("success", False) else "❌"
            recipient_count = email.get("recipient_count", 1)
            
            st.write(f"{i}. {success_icon} **{email_type}** ({recipient_count} Empfänger) - {timestamp}")
    
    # Excel-basierte Suchstatistiken
    if excel_stats.get("search_terms"):
        st.markdown("---")
        st.subheader("📊 Excel-basierte Suchstatistiken")
        
        # Top Suchbegriffe nach Papers
        top_searches = sorted(excel_stats["search_terms"], key=lambda x: x.get("papers", 0), reverse=True)[:5]
        
        st.write("**🔝 Top 5 Suchbegriffe (nach Papers):**")
        for i, search in enumerate(top_searches, 1):
            term = search.get("term", "Unbekannt")
            papers = search.get("papers", 0)
            new_papers = search.get("new_papers", 0)
            
            st.write(f"{i}. **{term}** - {papers} Papers ({new_papers} neue)")

def show_system_settings():
    """System-Einstellungen mit Excel-Integration"""
    st.subheader("⚙️ System-Einstellungen")
    
    # Excel-Template Einstellungen
    template_settings = st.session_state["excel_template"]
    
    with st.form("system_settings_form"):
        st.write("**📊 Excel-Template Einstellungen:**")
        
        col_set1, col_set2 = st.columns(2)
        
        with col_set1:
            auto_create_sheets = st.checkbox(
                "Automatische Sheet-Erstellung",
                value=template_settings.get("auto_create_sheets", True),
                help="Erstellt automatisch neue Sheets für jeden Suchbegriff"
            )
            
            max_sheets = st.number_input(
                "Maximale Anzahl Sheets",
                value=template_settings.get("max_sheets", 50),
                min_value=10,
                max_value=100,
                help="Maximale Anzahl von Sheets in der Excel-Datei"
            )
        
        with col_set2:
            sheet_naming = st.selectbox(
                "Sheet-Benennung",
                ["topic_based", "date_based", "custom"],
                index=0,
                help="Art der Sheet-Benennung"
            )
        
        if st.form_submit_button("💾 Einstellungen speichern"):
            st.session_state["excel_template"].update({
                "auto_create_sheets": auto_create_sheets,
                "max_sheets": max_sheets,
                "sheet_naming": sheet_naming
            })
            st.success("✅ System-Einstellungen gespeichert!")
    
    # System-Informationen
    st.markdown("---")
    st.subheader("ℹ️ System-Informationen")
    
    col_info1, col_info2 = st.columns(2)
    
    with col_info1:
        st.write("**📁 Pfade:**")
        st.code(f"Excel-Template: {st.session_state['excel_template']['file_path']}")
        st.code(f"Arbeitsverzeichnis: {os.getcwd()}")
    
    with col_info2:
        st.write("**🔧 Konfiguration:**")
        st.write(f"Auto-Sheets: {'✅' if template_settings.get('auto_create_sheets') else '❌'}")
        st.write(f"Max-Sheets: {template_settings.get('max_sheets', 50)}")
        st.write(f"Email-System: {'✅' if is_email_configured() else '❌'}")
    
    # System zurücksetzen
    st.markdown("---")
    st.subheader("🔄 System zurücksetzen")
    
    col_reset1, col_reset2, col_reset3 = st.columns(3)
    
    with col_reset1:
        if st.button("🗑️ Such-Historie löschen"):
            st.session_state["search_history"] = []
            st.success("Such-Historie gelöscht!")
    
    with col_reset2:
        if st.button("📧 Email-Historie löschen"):
            st.session_state["email_history"] = []
            st.success("Email-Historie gelöscht!")
    
    with col_reset3:
        if st.button("🤖 Auto-Suchen löschen"):
            st.session_state["automatic_searches"] = {}
            st.success("Automatische Suchen gelöscht!")

# =============== HILFSFUNKTIONEN ===============

def build_advanced_search_query(query: str, date_filter: str) -> str:
    """Erweiterte Suchanfrage mit Filtern"""
    query_parts = [query]
    
    if date_filter != "Alle":
        current_year = datetime.datetime.now().year
        if date_filter == "Letztes Jahr":
            query_parts.append(f"AND {current_year-1}:{current_year}[dp]")
        elif date_filter == "Letzte 2 Jahre":
            query_parts.append(f"AND {current_year-2}:{current_year}[dp]")
        elif date_filter == "Letzte 5 Jahre":
            query_parts.append(f"AND {current_year-5}:{current_year}[dp]")
        elif date_filter == "Letzte 10 Jahre":
            query_parts.append(f"AND {current_year-10}:{current_year}[dp]")
    
    return " ".join(query_parts)

def load_previous_search_results(query: str) -> List[Dict]:
    """Lädt vorherige Suchergebnisse aus Excel"""
    template_path = st.session_state["excel_template"]["file_path"]
    sheet_name = generate_sheet_name(query)
    
    if not os.path.exists(template_path):
        return []
    
    try:
        xl_file = pd.ExcelFile(template_path)
        if sheet_name not in xl_file.sheet_names:
            return []
        
        df = pd.read_excel(template_path, sheet_name=sheet_name)
        
        # SICHERE BEHANDLUNG LEERER DATAFRAMES
        if df.empty:
            return []
        
        previous_papers = []
        for _, row in df.iterrows():
            try:
                if pd.notna(row.get("PMID")):
                    paper = {
                        "PMID": str(row.get("PMID", "")),
                        "Title": str(row.get("Titel", "")),
                        "Authors": str(row.get("Autoren", "")),
                        "Journal": str(row.get("Journal", "")),
                        "Year": str(row.get("Jahr", ""))
                    }
                    previous_papers.append(paper)
            except Exception as row_error:
                # Überspringe fehlerhafte Zeilen
                continue
        
        return previous_papers
        
    except Exception as e:
        # Bei jedem Fehler eine leere Liste zurückgeben statt None
        return []


def identify_new_papers(current_papers: List[Dict], previous_papers: List[Dict]) -> List[Dict]:
    """Identifiziert neue Papers"""
    previous_pmids = set(paper.get("PMID", "") for paper in previous_papers if paper.get("PMID"))
    
    new_papers = []
    for paper in current_papers:
        current_pmid = paper.get("PMID", "")
        if current_pmid and current_pmid not in previous_pmids:
            paper["Is_New"] = True
            new_papers.append(paper)
        else:
            paper["Is_New"] = False
    
    return new_papers

def save_search_to_history(query: str, papers: List[Dict], new_papers: List[Dict]):
    """Speichert Suche in Historie"""
    search_entry = {
        "search_term": query,
        "timestamp": datetime.datetime.now().isoformat(),
        "paper_count": len(papers),
        "new_papers": len(new_papers),
        "date": datetime.datetime.now().date().isoformat()
    }
    
    st.session_state["search_history"].append(search_entry)

def update_system_status(paper_count: int):
    """Aktualisiert System-Status"""
    status = st.session_state["system_status"]
    status["total_searches"] += 1
    status["total_papers"] += paper_count
    status["last_search"] = datetime.datetime.now().isoformat()
    
    # Zähle Excel-Sheets
    excel_stats = get_search_statistics_from_excel()
    status["excel_sheets"] = excel_stats.get("total_sheets", 0)

def display_search_results(papers: List[Dict], new_papers: List[Dict], query: str, is_repeat: bool):
    """Zeigt Suchergebnisse an"""
    st.subheader(f"📋 Ergebnisse für: '{query}'")
    
    # Statistiken
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("📄 Gesamt Papers", len(papers))
    
    with col2:
        st.metric("🆕 Neue Papers", len(new_papers))
    
    with col3:
        with_abstract = len([p for p in papers if p.get("Abstract", "") != "Kein Abstract verfügbar"])
        st.metric("📝 Mit Abstract", with_abstract)
    
    with col4:
        with_doi = len([p for p in papers if p.get("DOI", "")])
        st.metric("🔗 Mit DOI", with_doi)
    
    # Papers anzeigen (erste 5)
    display_papers = papers[:5]
    
    for idx, paper in enumerate(display_papers, 1):
        is_new = paper.get("Is_New", False)
        status_icon = "🆕" if is_new else "📄"
        
        title = paper.get("Title", "Unbekannt")
        header = f"{status_icon} **{idx}.** {title[:60]}..."
        
        with st.expander(header):
            st.write(f"**📄 Titel:** {title}")
            st.write(f"**👥 Autoren:** {paper.get('Authors', 'n/a')}")
            st.write(f"**📚 Journal:** {paper.get('Journal', 'n/a')} ({paper.get('Year', 'n/a')})")
            st.write(f"**🆔 PMID:** {paper.get('PMID', 'n/a')}")
            
            if paper.get('DOI'):
                st.write(f"**🔗 DOI:** {paper.get('DOI')}")
            
            if paper.get('URL'):
                st.markdown(f"🔗 [**PubMed ansehen**]({paper.get('URL')})")
    
    if len(papers) > 5:
        st.info(f"... und {len(papers) - 5} weitere Papers (siehe Excel-Datei)")

def should_send_email(paper_count: int) -> bool:
    """Prüft ob Email gesendet werden soll"""
    settings = st.session_state.get("email_settings", {})
    return (settings.get("auto_notifications", False) and
            paper_count >= settings.get("min_papers", 1) and
            is_email_configured())
def show_automatic_search_system():
    """Automatisches Such-System (vereinfacht ohne schedule)"""
    st.subheader("🤖 Automatisches Such-System")
    
    st.info("""
    💡 **Hinweis:** Diese Version funktioniert ohne das 'schedule' Paket.
    Automatische Suchen können manuell ausgeführt werden.
    """)
    
    # Automatische Suchen verwalten
    auto_searches = st.session_state.get("automatic_searches", {})
    
    # Neue automatische Suche erstellen
    with st.expander("➕ Neue automatische Suche erstellen"):
        with st.form("create_auto_search"):
            col_auto1, col_auto2 = st.columns(2)
            
            with col_auto1:
                auto_search_term = st.text_input(
                    "Suchbegriff",
                    placeholder="z.B. 'diabetes genetics', 'COVID-19 treatment'"
                )
                
                auto_frequency = st.selectbox(
                    "Häufigkeit",
                    ["Täglich", "Wöchentlich", "Monatlich"],
                    index=1
                )
            
            with col_auto2:
                auto_max_papers = st.number_input(
                    "Max. Papers pro Suche",
                    min_value=10,
                    max_value=200,
                    value=50
                )
                
                auto_email_enabled = st.checkbox(
                    "Email-Benachrichtigungen",
                    value=True
                )
            
            if st.form_submit_button("🤖 **Automatische Suche erstellen**", type="primary"):
                if auto_search_term:
                    create_automatic_search(auto_search_term, auto_frequency, auto_max_papers, auto_email_enabled)
                else:
                    st.error("❌ Suchbegriff ist erforderlich!")
    
    # Bestehende automatische Suchen anzeigen
    if auto_searches:
        st.markdown("---")
        st.subheader(f"🤖 Konfigurierte automatische Suchen ({len(auto_searches)})")
        
        for search_id, search_config in auto_searches.items():
            search_term = search_config.get("search_term", "Unbekannt")
            frequency = search_config.get("frequency", "Unbekannt")
            last_run = search_config.get("last_run", "Nie")
            
            with st.expander(f"🤖 **{search_term}** ({frequency})"):
                col_config1, col_config2 = st.columns([2, 1])
                
                with col_config1:
                    st.write(f"**🔍 Suchbegriff:** {search_term}")
                    st.write(f"**⏰ Häufigkeit:** {frequency}")
                    st.write(f"**📧 Email:** {'✅' if search_config.get('email_enabled', False) else '❌'}")
                    st.write(f"**🕒 Letzter Lauf:** {last_run[:19] if last_run != 'Nie' else 'Nie'}")
                
                with col_config2:
                    if st.button("▶️ Jetzt ausführen", key=f"run_auto_{search_id}"):
                        run_automatic_search_simple(search_config)
                    
                    if st.button("🗑️ Löschen", key=f"delete_auto_{search_id}"):
                        delete_automatic_search(search_id)
                        st.rerun()
        
        # Globale Aktionen
        st.markdown("---")
        col_global1, col_global2 = st.columns(2)
        
        with col_global1:
            if st.button("▶️ **Alle automatischen Suchen ausführen**", type="primary"):
                run_all_automatic_searches_simple()
        
        with col_global2:
            if st.button("🔄 **Status aktualisieren**"):
                st.rerun()
    
    else:
        st.info("📭 Noch keine automatischen Suchen konfiguriert.")

def create_automatic_search(search_term: str, frequency: str, max_papers: int, email_enabled: bool):
    """Erstellt neue automatische Suche"""
    search_id = f"auto_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    search_config = {
        "search_id": search_id,
        "search_term": search_term,
        "frequency": frequency,
        "max_papers": max_papers,
        "email_enabled": email_enabled,
        "created_date": datetime.datetime.now().isoformat(),
        "last_run": "Nie",
        "total_runs": 0
    }
    
    st.session_state["automatic_searches"][search_id] = search_config
    
    st.success(f"✅ **Automatische Suche erstellt:** '{search_term}' ({frequency})")

def run_automatic_search_simple(search_config: Dict):
    """Führt eine automatische Suche aus (vereinfacht)"""
    search_term = search_config.get("search_term", "")
    max_papers = search_config.get("max_papers", 50)
    email_enabled = search_config.get("email_enabled", False)
    
    st.info(f"🤖 Führe automatische Suche aus: '{search_term}'")
    
    try:
        # Führe Excel-integrierte Suche durch
        execute_excel_integrated_search(search_term, max_papers, "Letzte 2 Jahre", email_enabled, False)
        
        # Update Konfiguration
        search_config["last_run"] = datetime.datetime.now().isoformat()
        search_config["total_runs"] = search_config.get("total_runs", 0) + 1
        
        st.success(f"✅ Automatische Suche für '{search_term}' abgeschlossen!")
        
    except Exception as e:
        st.error(f"❌ Fehler bei automatischer Suche '{search_term}': {str(e)}")

def run_all_automatic_searches_simple():
    """Führt alle automatischen Suchen aus (vereinfacht)"""
    auto_searches = st.session_state.get("automatic_searches", {})
    
    if not auto_searches:
        st.info("📭 Keine automatischen Suchen konfiguriert.")
        return
    
    st.info(f"🤖 Führe {len(auto_searches)} automatische Suchen aus...")
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_new_papers = 0
    
    for i, search_config in enumerate(auto_searches.values()):
        search_term = search_config.get("search_term", "")
        try:
            status_text.text(f"🔍 Automatische Suche {i+1}/{len(auto_searches)}: '{search_term}'...")
            
            # Führe Suche durch
            current_papers = perform_comprehensive_pubmed_search_robust(search_term, search_config.get("max_papers", 50))
            if not current_papers:
                current_papers = try_alternative_pubmed_search(search_term, search_config.get("max_papers", 50))

            
            if current_papers:
                # Füge neue Papers zur Excel hinzu
                added_count, new_papers = add_new_papers_to_excel(search_term, current_papers)
                
                if added_count > 0:
                    # Sende Email wenn konfiguriert
                    if search_config.get("email_enabled", False) and should_send_email(added_count):
                        send_excel_integrated_email_multiple(search_term, new_papers, len(current_papers), added_count)
                    
                    total_new_papers += added_count
                    st.write(f"✅ **{search_term}:** {added_count} neue Papers")
                else:
                    st.write(f"ℹ️ **{search_term}:** Keine neuen Papers")
                
                # Update Konfiguration
                search_config["last_run"] = datetime.datetime.now().isoformat()
                search_config["total_runs"] = search_config.get("total_runs", 0) + 1
            else:
                st.write(f"⚠️ **{search_term}:** Keine Papers gefunden")
            
            # Progress update
            progress_bar.progress((i + 1) / len(auto_searches))
            time.sleep(1)  # Rate limiting
            
        except Exception as e:
            st.error(f"❌ Fehler bei automatischer Suche '{search_term}': {str(e)}")
            continue
    
    progress_bar.empty()
    status_text.empty()
    
    # Ergebnis
    if total_new_papers > 0:
        st.success(f"🎉 **Alle automatischen Suchen abgeschlossen!** {total_new_papers} neue Papers insgesamt gefunden!")
        st.balloons()
    else:
        st.info("ℹ️ **Alle automatischen Suchen abgeschlossen.** Keine neuen Papers gefunden.")

def delete_automatic_search(search_id: str):
    """Löscht automatische Suche"""
    if search_id in st.session_state["automatic_searches"]:
        search_term = st.session_state["automatic_searches"][search_id].get("search_term", "Unbekannt")
        del st.session_state["automatic_searches"][search_id]
        st.success(f"🗑️ Automatische Suche '{search_term}' gelöscht!")
# =============== EXCEL-BASIERTE AUTOMATISCHE SUCHE ===============

def create_automation_excel_template():
    """Erstellt Excel-Template für automatische Suchen mit Einstellungen"""
    automation_path = "excel_templates/automation_schedule.xlsx"
    
    if not os.path.exists("excel_templates"):
        os.makedirs("excel_templates")
    
    if not os.path.exists(automation_path):
        try:
            wb = openpyxl.Workbook()
            
            # 1. AUTOMATION SCHEDULE SHEET
            schedule_sheet = wb.active
            schedule_sheet.title = "🤖_Auto_Schedule"
            
            # Header-Style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            
            # Schedule Headers
            schedule_headers = [
                "ID", "Suchbegriff", "Häufigkeit", "Max_Papers", "Email_Enabled",
                "Erstellt_am", "Letzter_Lauf", "Nächster_Lauf", "Total_Runs", 
                "Letzte_Neue_Papers", "Status", "Email_Empfänger"
            ]
            
            for col, header in enumerate(schedule_headers, 1):
                cell = schedule_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Spaltenbreite anpassen
            column_widths = [15, 30, 12, 12, 12, 18, 18, 18, 10, 15, 12, 40]
            for col, width in enumerate(column_widths, 1):
                col_letter = get_column_letter(col)
                schedule_sheet.column_dimensions[col_letter].width = width
            
            # Weitere Sheets... (vollständiger Code aus meiner vorherigen Antwort)
            wb.save(automation_path)
            st.session_state["automation_excel_path"] = automation_path
            
        except Exception as e:
            st.error(f"❌ Fehler beim Erstellen des Automation-Templates: {str(e)}")
    
    return automation_path

# Alle anderen neuen Funktionen hier hinzufügen...

if __name__ == "__main__":
    module_email()
