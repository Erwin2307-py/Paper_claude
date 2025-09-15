# modules/excel_manager.py - Persistent Excel File Management for Streamlit Cloud
import streamlit as st
import openpyxl
import pandas as pd
import os
import io
import json
import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path

class PersistentExcelManager:
    """Manages Excel files persistently for Streamlit Cloud deployment"""

    def __init__(self):
        self.ensure_excel_files()

    def ensure_excel_files(self):
        """Ensures all required Excel files exist, creates them if missing"""
        # Create directories if they don't exist
        os.makedirs("data", exist_ok=True)
        os.makedirs("modules", exist_ok=True)

        # Ensure core Excel files exist
        self.ensure_genes_excel()
        self.ensure_papers_template()
        self.ensure_snp_excel()

    def ensure_genes_excel(self):
        """Creates genes.xlsx if it doesn't exist"""
        genes_path = "modules/genes.xlsx"

        if not os.path.exists(genes_path):
            wb = openpyxl.Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create sample gene sheets
            gene_data = {
                "Oncogenes": [
                    "BRCA1", "BRCA2", "TP53", "EGFR", "MYC", "RAS", "PIK3CA",
                    "AKT1", "ERBB2", "ALK", "BRAF", "KRAS", "NRAS", "HRAS"
                ],
                "Tumor_Suppressors": [
                    "TP53", "RB1", "APC", "PTEN", "VHL", "NF1", "NF2", "CDKN2A",
                    "DCC", "BRCA1", "BRCA2", "MLH1", "MSH2", "MSH6"
                ],
                "DNA_Repair": [
                    "BRCA1", "BRCA2", "ATM", "ATR", "CHEK1", "CHEK2", "MLH1",
                    "MSH2", "MSH6", "PMS2", "XRCC1", "XRCC3", "RAD51", "PARP1"
                ],
                "Cell_Cycle": [
                    "TP53", "RB1", "CDKN1A", "CDKN1B", "CDKN2A", "CDKN2B",
                    "CCND1", "CCNE1", "CDK4", "CDK6", "E2F1", "MYC"
                ]
            }

            for sheet_name, genes in gene_data.items():
                ws = wb.create_sheet(title=sheet_name)
                ws['A1'] = 'Gene Symbol'
                ws['B1'] = 'Category'
                ws['C1'] = 'Description'

                for i, gene in enumerate(genes, 2):
                    ws[f'A{i}'] = gene
                    ws[f'B{i}'] = sheet_name.replace('_', ' ')
                    ws[f'C{i}'] = f"{gene} - {sheet_name.replace('_', ' ')} related gene"

            wb.save(genes_path)
            st.success(f"✅ Created genes.xlsx with {len(gene_data)} gene categories")

    def ensure_papers_template(self):
        """Creates paper analysis template if it doesn't exist"""
        template_path = "data/vorlage_paperqa2.xlsx"

        if not os.path.exists(template_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Paper_Analysis_Template"

            # Headers for paper analysis
            headers = [
                "Title", "Authors", "Journal", "Year", "DOI", "Abstract",
                "Keywords", "Methodology", "Results", "Conclusions",
                "Relevance_Score", "Category", "Notes", "Date_Added"
            ]

            for i, header in enumerate(headers, 1):
                ws.cell(row=1, column=i, value=header)

            # Add example row
            example_data = [
                "Example Paper Title",
                "Smith, J. et al.",
                "Nature",
                "2024",
                "10.1038/example",
                "This is an example abstract...",
                "gene expression, cancer, biomarkers",
                "RNA-seq analysis",
                "Significant findings...",
                "Important conclusions...",
                "8.5",
                "Oncology",
                "High relevance for current research",
                datetime.datetime.now().strftime("%Y-%m-%d")
            ]

            for i, value in enumerate(example_data, 1):
                ws.cell(row=2, column=i, value=value)

            wb.save(template_path)
            st.success(f"✅ Created paper analysis template: {template_path}")

    def ensure_snp_excel(self):
        """Creates SNP data file if it doesn't exist"""
        snp_path = "modules/snp.xlsx"

        if not os.path.exists(snp_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "SNP_Data"

            # Headers for SNP data
            headers = [
                "SNP_ID", "Chromosome", "Position", "Gene", "Alleles",
                "MAF", "Clinical_Significance", "Disease_Association",
                "Reference", "Notes"
            ]

            for i, header in enumerate(headers, 1):
                ws.cell(row=1, column=i, value=header)

            # Add example SNP data
            example_snps = [
                ["rs429358", "19", "45411941", "APOE", "C/T", "0.137", "Pathogenic", "Alzheimer's Disease", "PMID:12345", "Major risk factor"],
                ["rs7412", "19", "45412079", "APOE", "C/T", "0.078", "Protective", "Alzheimer's Disease", "PMID:67890", "Protective variant"],
                ["rs1801282", "3", "12393125", "PPARG", "C/G", "0.117", "Benign", "Type 2 Diabetes", "PMID:11111", "Metabolism related"],
            ]

            for i, snp_data in enumerate(example_snps, 2):
                for j, value in enumerate(snp_data, 1):
                    ws.cell(row=i, column=j, value=value)

            wb.save(snp_path)
            st.success(f"✅ Created SNP data file: {snp_path}")

    def get_available_gene_sheets(self) -> List[str]:
        """Returns list of available gene sheets"""
        genes_path = "modules/genes.xlsx"
        if os.path.exists(genes_path):
            try:
                xls = pd.ExcelFile(genes_path)
                return xls.sheet_names
            except Exception as e:
                st.error(f"Error reading gene sheets: {e}")
                return []
        return []

    def load_genes_from_sheet(self, sheet_name: str) -> List[str]:
        """Loads genes from specified sheet"""
        genes_path = "modules/genes.xlsx"
        try:
            df = pd.read_excel(genes_path, sheet_name=sheet_name)
            if 'Gene Symbol' in df.columns:
                return df['Gene Symbol'].dropna().tolist()
            else:
                # Fallback to first column
                return df.iloc[:, 0].dropna().tolist()
        except Exception as e:
            st.error(f"Error loading genes from {sheet_name}: {e}")
            return []

    def create_persistent_paper_database(self, file_path: str = "data/master_papers.xlsx"):
        """Creates persistent paper database for research tracking"""
        if os.path.exists(file_path):
            return file_path

        try:
            wb = openpyxl.Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create Overview sheet
            overview = wb.create_sheet("Overview")
            overview_headers = [
                "Search_Term", "Total_Papers", "Last_Updated", "Sheet_Name",
                "Relevance_Score", "Status", "Notes"
            ]
            for i, header in enumerate(overview_headers, 1):
                overview.cell(row=1, column=i, value=header)

            # Create template sheet for papers
            template = wb.create_sheet("Template")
            template_headers = [
                "Title", "Authors", "Journal", "Year", "DOI", "Abstract",
                "Keywords", "PubMed_ID", "Citations", "Relevance_Score",
                "Category", "Methodology", "Key_Findings", "Notes", "Date_Added"
            ]
            for i, header in enumerate(template_headers, 1):
                template.cell(row=1, column=i, value=header)

            wb.save(file_path)
            st.success(f"✅ Created persistent paper database: {file_path}")
            return file_path

        except Exception as e:
            st.error(f"Error creating paper database: {e}")
            return None

    def add_paper_to_database(self, paper_data: Dict, search_term: str,
                            file_path: str = "data/master_papers.xlsx"):
        """Adds paper data to persistent database"""
        try:
            if not os.path.exists(file_path):
                self.create_persistent_paper_database(file_path)

            wb = openpyxl.load_workbook(file_path)

            # Create or get sheet for search term
            safe_sheet_name = search_term.replace('/', '_').replace('\\', '_')[:31]
            if safe_sheet_name not in wb.sheetnames:
                # Copy template structure
                if "Template" in wb.sheetnames:
                    template = wb["Template"]
                    new_sheet = wb.copy_worksheet(template)
                    new_sheet.title = safe_sheet_name
                    # Clear example data, keep headers
                    for row in new_sheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.value = None
                else:
                    new_sheet = wb.create_sheet(safe_sheet_name)

            sheet = wb[safe_sheet_name]

            # Add paper data to next available row
            next_row = sheet.max_row + 1
            paper_values = [
                paper_data.get('title', ''),
                paper_data.get('authors', ''),
                paper_data.get('journal', ''),
                paper_data.get('year', ''),
                paper_data.get('doi', ''),
                paper_data.get('abstract', ''),
                paper_data.get('keywords', ''),
                paper_data.get('pubmed_id', ''),
                paper_data.get('citations', ''),
                paper_data.get('relevance_score', ''),
                paper_data.get('category', ''),
                paper_data.get('methodology', ''),
                paper_data.get('key_findings', ''),
                paper_data.get('notes', ''),
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]

            for i, value in enumerate(paper_values, 1):
                sheet.cell(row=next_row, column=i, value=value)

            # Update overview sheet
            if "Overview" in wb.sheetnames:
                overview = wb["Overview"]
                # Find or create overview entry
                overview_updated = False
                for row in overview.iter_rows(min_row=2):
                    if row[0].value == search_term:
                        row[1].value = (row[1].value or 0) + 1  # Increment paper count
                        row[2].value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        overview_updated = True
                        break

                if not overview_updated:
                    next_overview_row = overview.max_row + 1
                    overview_values = [
                        search_term, 1, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        safe_sheet_name, "", "Active", ""
                    ]
                    for i, value in enumerate(overview_values, 1):
                        overview.cell(row=next_overview_row, column=i, value=value)

            wb.save(file_path)
            return True

        except Exception as e:
            st.error(f"Error adding paper to database: {e}")
            return False

    def get_database_stats(self, file_path: str = "data/master_papers.xlsx") -> Dict:
        """Returns statistics about the paper database"""
        stats = {
            "total_sheets": 0,
            "total_papers": 0,
            "search_terms": [],
            "last_updated": None
        }

        if not os.path.exists(file_path):
            return stats

        try:
            wb = openpyxl.load_workbook(file_path)

            # Count sheets (exclude Overview and Template)
            data_sheets = [name for name in wb.sheetnames if name not in ["Overview", "Template"]]
            stats["total_sheets"] = len(data_sheets)

            # Get data from Overview sheet if available
            if "Overview" in wb.sheetnames:
                overview = wb["Overview"]
                for row in overview.iter_rows(min_row=2):
                    if row[0].value:  # Search term exists
                        stats["search_terms"].append({
                            "term": row[0].value,
                            "papers": row[1].value or 0,
                            "last_update": row[2].value,
                            "sheet_name": row[3].value
                        })
                        stats["total_papers"] += (row[1].value or 0)

                        if row[2].value and (not stats["last_updated"] or row[2].value > stats["last_updated"]):
                            stats["last_updated"] = row[2].value

            return stats

        except Exception as e:
            st.error(f"Error getting database stats: {e}")
            return stats

def initialize_excel_manager():
    """Initialize Excel Manager in Streamlit session state"""
    if "excel_manager" not in st.session_state:
        st.session_state["excel_manager"] = PersistentExcelManager()
        st.success("✅ Excel Manager initialized - All required files are available")

    return st.session_state["excel_manager"]

def show_excel_manager_dashboard():
    """Display Excel Manager dashboard"""
    st.subheader("📊 Excel File Management Dashboard")

    manager = initialize_excel_manager()

    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric("Gene Categories", len(manager.get_available_gene_sheets()))

    with col2:
        stats = manager.get_database_stats()
        st.metric("Research Topics", len(stats["search_terms"]))

    with col3:
        st.metric("Total Papers", stats["total_papers"])

    # File status
    with st.expander("📁 File Status", expanded=True):
        files_to_check = [
            ("modules/genes.xlsx", "Gene Database"),
            ("modules/snp.xlsx", "SNP Database"),
            ("data/vorlage_paperqa2.xlsx", "Paper Template"),
            ("data/master_papers.xlsx", "Research Database")
        ]

        for file_path, description in files_to_check:
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path) / 1024  # KB
                st.success(f"✅ {description}: {file_path} ({file_size:.1f} KB)")
            else:
                st.warning(f"⚠️ {description}: {file_path} (Missing)")

    # Gene sheets overview
    if st.button("🔄 Refresh All Files"):
        manager.ensure_excel_files()
        st.rerun()

    return manager