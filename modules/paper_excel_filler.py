# modules/paper_excel_filler.py - Automatisierte Excel-Ausfüllung für Paper-Daten
import streamlit as st
import requests
import xml.etree.ElementTree as ET
import pandas as pd
import datetime
import time
import os
import shutil
import json
import re
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook
from dataclasses import dataclass

# Lokale Imports
from modules.unified_paper_search import Paper


@dataclass
class ExcelFillData:
    """Struktur für Excel-Ausfüllung aus Paper-Daten"""
    paper_title: str
    authors: str
    journal: str
    year: str
    abstract: str
    doi: str
    pubmed_id: str
    chatgpt_rating: float
    chatgpt_summary: str
    keywords: str
    gene_name: str = ""
    rs_numbers: List[str] = None
    phenotypes: List[str] = None
    genotypes: List[str] = None
    frequencies: List[str] = None

    def __post_init__(self):
        if self.rs_numbers is None:
            self.rs_numbers = []
        if self.phenotypes is None:
            self.phenotypes = []
        if self.genotypes is None:
            self.genotypes = []
        if self.frequencies is None:
            self.frequencies = []


class PaperExcelFiller:
    """Automatisierte Excel-Ausfüllung für wissenschaftliche Paper mit Claude API"""

    def __init__(self):
        self.claude_api_key = self._get_claude_api_key()

    def _get_claude_api_key(self) -> Optional[str]:
        """Holt Claude API Key aus Streamlit Secrets"""
        try:
            if hasattr(st, 'secrets'):
                # Versuche verschiedene mögliche Speicherorte
                claude_key = st.secrets.get("claude", {}).get("api_key")
                if claude_key:
                    return claude_key

                # Fallback: Anthropic section
                anthropic_key = st.secrets.get("anthropic", {}).get("api_key")
                if anthropic_key:
                    return anthropic_key

        except Exception:
            pass

        # Environment variable als letzter Fallback
        return os.getenv("CLAUDE_API_KEY") or os.getenv("ANTHROPIC_API_KEY")

    def copy_template(self, paper_title: str, output_dir: str = "Excel") -> str:
        """Kopiert Excel-Vorlage für die Ausfüllung"""
        # Stelle sicher, dass Output-Ordner existiert
        os.makedirs(output_dir, exist_ok=True)

        # Suche nach verfügbaren Vorlagen
        template_paths = [
            "vorlage_paperqa2.xlsx",
            "modules/vorlage_paperqa2.xlsx",
            "vorlage_gene.xlsx",
            "modules/vorlage_gene.xlsx"
        ]

        src_file = None
        for path in template_paths:
            if os.path.exists(path):
                src_file = path
                break

        if not src_file:
            raise FileNotFoundError(f"Keine Excel-Vorlage gefunden! Suchpfade: {template_paths}")

        # Erstelle eindeutigen Dateinamen
        safe_title = re.sub(r'[^\w\s-]', '', paper_title.replace(' ', '_'))[:50]
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        dst_file = os.path.join(output_dir, f"paper_analysis_{safe_title}_{timestamp}.xlsx")

        # Kopiere Vorlage
        shutil.copy2(src_file, dst_file)

        # Verifiziere Kopie
        if not os.path.exists(dst_file):
            raise IOError(f"Fehler beim Kopieren der Vorlage von {src_file} nach {dst_file}")

        return dst_file

    def extract_gene_from_paper(self, paper: Paper) -> Optional[str]:
        """Extrahiert Gene-Namen aus Paper-Daten"""
        # Suche nach Gene-Namen in Titel, Abstract und Keywords
        text_to_search = f"{paper.title} {paper.abstract} {paper.keywords}".lower()

        # Bekannte Gene-Namen (kann erweitert werden)
        common_genes = [
            "brca1", "brca2", "tp53", "egfr", "myc", "ras", "pik3ca",
            "akt1", "erbb2", "alk", "braf", "kras", "nras", "hras",
            "rb1", "apc", "pten", "vhl", "nf1", "nf2", "cdkn2a",
            "atm", "atr", "chek1", "chek2", "mlh1", "msh2", "msh6"
        ]

        # Suche nach Genen im Text
        for gene in common_genes:
            if gene in text_to_search:
                return gene.upper()

        # Fallback: Versuche GenBank-ähnliche Patterns zu finden
        gene_pattern = r'\b[A-Z]{2,6}\d*\b'
        matches = re.findall(gene_pattern, paper.title + " " + paper.abstract)
        if matches:
            return matches[0]

        return None

    def fetch_rs_numbers_for_gene(self, gene: str, max_results: int = 3) -> List[str]:
        """Holt rs-Nummern für ein Gen über dbSNP"""
        if not gene:
            return []

        try:
            # dbSNP Suche
            ids_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
            response = requests.get(ids_url, params={
                "db": "snp",
                "term": f"{gene}[Gene Name]",
                "retmax": max_results,
                "retmode": "xml"
            }, timeout=10)

            if response.status_code != 200:
                return [f"rs{gene.lower()}1"]

            # Parse XML Response
            root = ET.fromstring(response.text)
            snp_ids = [id_elem.text for id_elem in root.findall(".//Id")]

            if not snp_ids:
                return [f"rs{gene.lower()}1"]

            # Hole Details zu den SNPs
            fetch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
            detail_response = requests.get(fetch_url, params={
                "db": "snp",
                "id": ",".join(snp_ids),
                "retmode": "xml"
            }, timeout=10)

            if detail_response.status_code == 200:
                detail_root = ET.fromstring(detail_response.text)
                rs_numbers = []
                for rs_elem in detail_root.findall(".//Rs"):
                    rs_id = rs_elem.get("rsId")
                    if rs_id:
                        rs_numbers.append(f"rs{rs_id}")

                return rs_numbers[:max_results] if rs_numbers else [f"rs{gene.lower()}1"]

            return [f"rs{gene.lower()}1"]

        except Exception as e:
            st.warning(f"⚠️ Fehler beim Abrufen der rs-Nummern: {str(e)}")
            return [f"rs{gene.lower()}1"]

    def analyze_paper_with_claude(self, paper: Paper, gene: str = None) -> Dict[str, Any]:
        """Analysiert Paper mit Claude API für detaillierte Erkenntnisse"""
        if not self.claude_api_key:
            return self._get_fallback_analysis(paper, gene)

        # Erstelle intelligenten Prompt basierend auf Paper-Daten
        prompt = f"""
Analysiere das folgende wissenschaftliche Paper und extrahiere strukturierte Informationen:

**Paper Details:**
Titel: {paper.title}
Autoren: {paper.authors}
Journal: {paper.journal}
Jahr: {paper.year}
Abstract: {paper.abstract[:2000]}...
Keywords: {paper.keywords}
{'Gen: ' + gene if gene else ''}

**Aufgabe:**
Erstelle eine JSON-Antwort mit folgender Struktur:

{{
    "key_findings": ["Haupterkenntnisse 1", "Haupterkenntnisse 2", "Haupterkenntnisse 3"],
    "methodology": "Kurze Beschreibung der verwendeten Methoden",
    "clinical_relevance": "Klinische Bedeutung der Ergebnisse",
    "limitations": "Hauptlimitationen der Studie",
    "gene_analysis": {{
        "primary_gene": "Hauptgen (falls identifiziert)",
        "genetic_variants": ["Variante 1", "Variante 2"],
        "phenotype_associations": ["Phänotyp 1", "Phänotyp 2"],
        "population_data": "Populationsdaten falls verfügbar"
    }},
    "quality_score": 8.5,
    "research_category": "Grundlagenforschung/Klinische Studie/Review/etc.",
    "future_directions": "Zukünftige Forschungsrichtungen"
}}

Sei präzise und nutze nur Informationen aus dem Abstract und den verfügbaren Daten.
"""

        headers = {
            "Content-Type": "application/json",
            "x-api-key": self.claude_api_key,
            "anthropic-version": "2023-06-01"
        }

        payload = {
            "model": "claude-3-haiku-20240307",
            "max_tokens": 1500,
            "messages": [{"role": "user", "content": prompt}]
        }

        try:
            response = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers=headers,
                json=payload,
                timeout=30
            )
            response.raise_for_status()

            # Parse Response
            response_data = response.json()
            content_text = response_data["content"][0]["text"]

            # Extrahiere JSON aus Response
            json_match = re.search(r'\{.*\}', content_text, re.DOTALL)
            if json_match:
                analysis_data = json.loads(json_match.group())
                return analysis_data
            else:
                st.warning("⚠️ Claude API Response konnte nicht geparst werden")
                return self._get_fallback_analysis(paper, gene)

        except requests.exceptions.RequestException as e:
            st.error(f"❌ Claude API Fehler: {str(e)}")
            return self._get_fallback_analysis(paper, gene)
        except json.JSONDecodeError as e:
            st.error(f"❌ JSON Parse Fehler: {str(e)}")
            return self._get_fallback_analysis(paper, gene)
        except Exception as e:
            st.error(f"❌ Unerwarteter Fehler bei Claude Analyse: {str(e)}")
            return self._get_fallback_analysis(paper, gene)

    def _get_fallback_analysis(self, paper: Paper, gene: str = None) -> Dict[str, Any]:
        """Fallback-Analyse wenn Claude API nicht verfügbar ist"""
        return {
            "key_findings": ["Analysiere Paper manuell", "Keine automatische Analyse verfügbar", "Claude API Key erforderlich"],
            "methodology": "Methodenanalyse nicht verfügbar",
            "clinical_relevance": "Klinische Relevanz manuell bewerten",
            "limitations": "Limitationen aus Abstract extrahieren",
            "gene_analysis": {
                "primary_gene": gene or "Nicht identifiziert",
                "genetic_variants": ["Manuelle Analyse erforderlich"],
                "phenotype_associations": ["Aus Abstract extrahieren"],
                "population_data": "Nicht verfügbar"
            },
            "quality_score": paper.chatgpt_rating or 5.0,
            "research_category": "Unbestimmt",
            "future_directions": "Aus Paper-Diskussion extrahieren"
        }

    def fill_excel_with_paper_data(self, xlsx_path: str, paper: Paper, analysis_data: Dict[str, Any]) -> bool:
        """Füllt Excel-Datei mit Paper-Daten aus"""
        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active

            # Grunddaten ausfüllen
            current_date = datetime.datetime.now().strftime("%d.%m.%Y")

            # Basis-Informationen (sicher ausfüllen)
            try:
                ws["D5"] = paper.title[:100] if paper.title else "Unbekannt"
                ws["D6"] = paper.authors[:100] if paper.authors else "Unbekannt"
                ws["D7"] = paper.journal[:100] if paper.journal else "Unbekannt"
                ws["D8"] = paper.year if paper.year else "Unbekannt"
                ws["I2"] = current_date  # Datum der Analyse
            except Exception as e:
                # Fallback: Try different cell locations or skip problematic fields
                st.warning(f"⚠️ Fehler beim Ausfüllen der Basis-Informationen: {str(e)}")
                pass

            # DOI und PMID
            ws["D9"] = paper.doi if paper.doi else "Nicht verfügbar"
            ws["D10"] = paper.pubmed_id if paper.pubmed_id else "Nicht verfügbar"

            # ChatGPT/Claude Bewertung
            ws["D11"] = f"{paper.chatgpt_rating}/10" if paper.chatgpt_rating else "Nicht bewertet"

            # Analyse-Ergebnisse
            gene_info = analysis_data.get("gene_analysis", {})

            # Gen-Informationen (Zeilen 13-15)
            ws["D13"] = gene_info.get("primary_gene", "Nicht identifiziert")
            ws["D14"] = ", ".join(gene_info.get("genetic_variants", [])[:3])
            ws["D15"] = ", ".join(gene_info.get("phenotype_associations", [])[:3])

            # Haupterkenntnisse (Zeilen 17-19)
            key_findings = analysis_data.get("key_findings", [])
            for i, finding in enumerate(key_findings[:3]):
                ws[f"D{17+i}"] = finding[:200]  # Begrenzte Länge

            # Methodologie und klinische Relevanz (Zeilen 21-22)
            ws["D21"] = analysis_data.get("methodology", "")[:300]
            ws["D22"] = analysis_data.get("clinical_relevance", "")[:300]

            # Qualitätsbewertung und Kategorie (Zeilen 24-25)
            ws["D24"] = analysis_data.get("quality_score", "Nicht bewertet")
            ws["D25"] = analysis_data.get("research_category", "Unbestimmt")

            # Abstract (gekürzt, Zeile 27)
            ws["D27"] = paper.abstract[:500] + "..." if len(paper.abstract) > 500 else paper.abstract

            # Limitationen und Zukunftsperspektiven (Zeilen 29-30)
            ws["D29"] = analysis_data.get("limitations", "")[:300]
            ws["D30"] = analysis_data.get("future_directions", "")[:300]

            # Speichere Datei
            wb.save(xlsx_path)
            return True

        except Exception as e:
            st.error(f"❌ Fehler beim Ausfüllen der Excel-Datei: {str(e)}")
            return False

    def process_paper_to_excel(self, paper: Paper, output_dir: str = "Excel") -> Optional[str]:
        """Verarbeitet ein Paper komplett zu einer ausgefüllten Excel-Datei"""
        try:
            # 1. Template kopieren
            with st.spinner("📄 Kopiere Excel-Vorlage..."):
                xlsx_path = self.copy_template(paper.title, output_dir)

            # 2. Gen extrahieren (falls vorhanden)
            gene = self.extract_gene_from_paper(paper)
            if gene:
                st.info(f"🧬 Identifiziertes Gen: **{gene}**")

            # 3. Claude-Analyse durchführen
            with st.spinner("🤖 Analysiere Paper mit Claude API..."):
                analysis_data = self.analyze_paper_with_claude(paper, gene)

            # 4. Excel ausfüllen
            with st.spinner("📊 Fülle Excel-Vorlage aus..."):
                success = self.fill_excel_with_paper_data(xlsx_path, paper, analysis_data)

            if success:
                st.success(f"✅ Excel-Datei erfolgreich erstellt: `{os.path.basename(xlsx_path)}`")
                return xlsx_path
            else:
                st.error("❌ Fehler beim Ausfüllen der Excel-Datei")
                return None

        except Exception as e:
            st.error(f"❌ Fehler bei der Paper-Verarbeitung: {str(e)}")
            return None


def show_paper_excel_interface(selected_papers: List[Paper]):
    """Streamlit Interface für Paper-zu-Excel Konvertierung"""
    if not selected_papers:
        st.info("📋 Keine Papers ausgewählt. Wählen Sie zuerst Papers in der Suche aus.")
        return

    st.subheader("📊 Excel-Ausfüllung für ausgewählte Papers")

    # Excel Filler initialisieren
    filler = PaperExcelFiller()

    # API-Status anzeigen
    api_status = "✅ Verfügbar" if filler.claude_api_key else "❌ Nicht konfiguriert"
    st.info(f"🤖 **Claude API Status:** {api_status}")

    if not filler.claude_api_key:
        st.warning("⚠️ **Claude API Key fehlt!** Fallback-Analyse wird verwendet. Konfigurieren Sie den API Key in den Streamlit Secrets.")

    # Paper-Auswahl
    st.write(f"📋 **Verfügbare Papers:** {len(selected_papers)}")

    # Einzelne Paper-Verarbeitung
    for i, paper in enumerate(selected_papers):
        with st.expander(f"📄 **{i+1}.** {paper.title[:80]}..."):
            col1, col2 = st.columns([3, 1])

            with col1:
                st.write(f"**👥 Autoren:** {paper.authors}")
                st.write(f"**📚 Journal:** {paper.journal} ({paper.year})")
                st.write(f"**⭐ ChatGPT Rating:** {paper.chatgpt_rating}/10" if paper.chatgpt_rating else "Nicht bewertet")
                st.write(f"**🔬 Quelle:** {paper.source}")

                if paper.chatgpt_summary:
                    st.write(f"**📝 Zusammenfassung:** {paper.chatgpt_summary[:200]}...")

            with col2:
                if st.button(f"📊 **Excel Erstellen**", key=f"excel_{i}", type="primary"):
                    with st.spinner("⏳ Verarbeite Paper..."):
                        excel_path = filler.process_paper_to_excel(paper)

                        if excel_path and os.path.exists(excel_path):
                            # Download-Button für die erstellte Datei
                            with open(excel_path, 'rb') as file:
                                st.download_button(
                                    label="📥 **Excel Herunterladen**",
                                    data=file.read(),
                                    file_name=os.path.basename(excel_path),
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"download_{i}"
                                )

    # Batch-Verarbeitung
    if len(selected_papers) > 1:
        st.markdown("---")
        st.subheader("🚀 Batch-Verarbeitung")

        if st.button("📊 **Alle Papers zu Excel verarbeiten**", type="primary"):
            progress_bar = st.progress(0)
            status_text = st.empty()

            successful_files = []
            failed_papers = []

            for i, paper in enumerate(selected_papers):
                progress = (i + 1) / len(selected_papers)
                progress_bar.progress(progress)
                status_text.text(f"Verarbeite Paper {i+1}/{len(selected_papers)}: {paper.title[:50]}...")

                excel_path = filler.process_paper_to_excel(paper)

                if excel_path:
                    successful_files.append(excel_path)
                else:
                    failed_papers.append(paper.title)

            # Ergebnisse anzeigen
            progress_bar.progress(1.0)
            status_text.text("✅ Batch-Verarbeitung abgeschlossen!")

            if successful_files:
                st.success(f"✅ **{len(successful_files)} Excel-Dateien erfolgreich erstellt!**")

                # ZIP-Download für alle Dateien anbieten
                if len(successful_files) > 1:
                    try:
                        import zipfile
                        import io

                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                            for file_path in successful_files:
                                if os.path.exists(file_path):
                                    zip_file.write(file_path, os.path.basename(file_path))

                        zip_buffer.seek(0)
                        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

                        st.download_button(
                            label="📥 **Alle Excel-Dateien als ZIP herunterladen**",
                            data=zip_buffer.getvalue(),
                            file_name=f"paper_excel_batch_{timestamp}.zip",
                            mime="application/zip"
                        )
                    except ImportError:
                        st.warning("⚠️ ZIP-Export nicht verfügbar. Laden Sie die Dateien einzeln herunter.")

            if failed_papers:
                st.error(f"❌ **{len(failed_papers)} Papers konnten nicht verarbeitet werden:**")
                for title in failed_papers:
                    st.write(f"• {title}")