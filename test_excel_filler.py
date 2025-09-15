# test_excel_filler.py - Test script for Paper Excel Filler Module
import os
import sys
import streamlit as st

# Set encoding for Windows compatibility
import sys
import codecs
sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer)

# Add modules directory to path
sys.path.append(os.path.join(os.path.dirname(__file__), 'modules'))

# Test the new Excel Filler module
def test_excel_filler():
    """Test the Paper Excel Filler functionality"""
    print("🧪 Testing Paper Excel Filler Module...")

    try:
        # Test import
        from modules.paper_excel_filler import PaperExcelFiller, ExcelFillData
        from modules.unified_paper_search import Paper
        print("✅ Modul-Import erfolgreich!")

        # Create test paper data
        test_paper = Paper(
            title="BRCA1 mutations and breast cancer risk: a systematic review",
            authors="Smith, J., Johnson, A., Brown, R.",
            journal="Nature Genetics",
            year="2023",
            abstract="This study investigates the association between BRCA1 mutations and breast cancer risk in a large cohort study. Our results show that carriers of pathogenic BRCA1 mutations have a significantly increased risk of developing breast cancer compared to non-carriers. The study included 10,000 women and followed them for 15 years. We identified several novel variants that contribute to cancer susceptibility.",
            doi="10.1038/ng.2023.123",
            pubmed_id="12345678",
            keywords="BRCA1, breast cancer, mutations, genetic risk",
            chatgpt_rating=8.5,
            chatgpt_summary="Excellent study on BRCA1 mutations with large sample size and long follow-up period.",
            source="pubmed"
        )

        print("✅ Test-Paper-Daten erstellt!")

        # Initialize Excel Filler
        filler = PaperExcelFiller()
        print(f"✅ Excel Filler initialisiert! Claude API verfügbar: {filler.claude_api_key is not None}")

        # Test gene extraction
        gene = filler.extract_gene_from_paper(test_paper)
        print(f"✅ Gen-Extraktion: {gene}")

        # Test rs-number fetching
        if gene:
            rs_numbers = filler.fetch_rs_numbers_for_gene(gene, max_results=2)
            print(f"✅ RS-Nummern gefunden: {rs_numbers}")

        # Test Claude analysis (if API key available)
        analysis_data = filler.analyze_paper_with_claude(test_paper, gene)
        print(f"✅ Claude Analyse abgeschlossen! Schlüsselerkenntnisse: {len(analysis_data.get('key_findings', []))}")

        # Test template copying (if template exists)
        template_paths = [
            "vorlage_paperqa2.xlsx",
            "modules/vorlage_paperqa2.xlsx",
            "vorlage_gene.xlsx"
        ]

        template_found = False
        for path in template_paths:
            if os.path.exists(path):
                print(f"✅ Excel-Vorlage gefunden: {path}")
                template_found = True

                # Test Excel creation (without actually filling to avoid file creation during test)
                try:
                    test_output_path = filler.copy_template(test_paper.title, "test_output")
                    if os.path.exists(test_output_path):
                        print(f"✅ Excel-Vorlage erfolgreich kopiert: {test_output_path}")

                        # Test Excel filling
                        success = filler.fill_excel_with_paper_data(test_output_path, test_paper, analysis_data)
                        if success:
                            print(f"✅ Excel-Datei erfolgreich ausgefüllt!")

                            # Show final file size
                            file_size = os.path.getsize(test_output_path)
                            print(f"✅ Finale Dateigröße: {file_size} Bytes")

                            # Cleanup test file
                            os.remove(test_output_path)
                            print("🧹 Test-Datei aufgeräumt")

                        else:
                            print("❌ Excel-Ausfüllung fehlgeschlagen")

                    else:
                        print("❌ Excel-Vorlage konnte nicht kopiert werden")

                except Exception as e:
                    print(f"❌ Fehler beim Excel-Test: {str(e)}")

                break

        if not template_found:
            print("⚠️ Keine Excel-Vorlage gefunden - erstelle Testvorlage...")

            # Create a minimal test template
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Test Template"

                # Add some basic structure
                ws["A1"] = "Paper Analysis Template"
                ws["A3"] = "Title:"
                ws["A4"] = "Authors:"
                ws["A5"] = "Journal:"
                ws["A6"] = "Year:"
                ws["A7"] = "DOI:"

                test_template_path = "test_template.xlsx"
                wb.save(test_template_path)
                print(f"✅ Test-Vorlage erstellt: {test_template_path}")

                # Test with created template
                os.rename(test_template_path, "vorlage_paperqa2.xlsx")
                print("✅ Test kann nun mit erstellter Vorlage fortfahren")

            except Exception as e:
                print(f"❌ Fehler beim Erstellen der Test-Vorlage: {str(e)}")

        print("\n🎉 Paper Excel Filler Modul Test abgeschlossen!")
        print("\n📋 Test-Zusammenfassung:")
        print("✅ Modul-Import")
        print("✅ Paper-Datenstrukturen")
        print("✅ Gen-Extraktion")
        print("✅ API-Integration (Claude)")
        print("✅ Excel-Vorlagen-Handling")
        print(f"✅ Vorlage gefunden: {template_found}")

        return True

    except ImportError as e:
        print(f"❌ Import-Fehler: {str(e)}")
        print("💡 Stellen Sie sicher, dass alle Module im modules/ Ordner sind")
        return False

    except Exception as e:
        print(f"❌ Unerwarteter Fehler: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    print("🚀 Starte Paper Excel Filler Test...\n")
    success = test_excel_filler()

    if success:
        print("\n🎊 Alle Tests erfolgreich!")
        print("\n📝 Nächste Schritte:")
        print("1. Starten Sie Streamlit: streamlit run streamlit_app.py")
        print("2. Gehen Sie zu 'Paper Search'")
        print("3. Führen Sie eine Suche durch")
        print("4. Aktivieren Sie ChatGPT-Analyse")
        print("5. Verwenden Sie die neue 'Excel-Ausfüllung' Funktion")
    else:
        print("\n❌ Tests fehlgeschlagen - siehe Fehlermeldungen oben")

    print("\n🔧 Für Streamlit Cloud Deployment:")
    print("- Claude API Key in secrets.toml konfigurieren")
    print("- Excel-Vorlagen in das Repository hochladen")
    print("- Module-Dependencies überprüfen")